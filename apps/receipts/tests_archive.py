"""
Adversarial tests for the ARCHIVE feature of apps.receipts:
``SlipReview.archived / archived_at / archived_by``, the three-way ``archived``
list/ids/export filter (whose DEFAULT — uniquely in build_filters — ADDS a
clause), the PATCH review + bulk write paths, and the v4 export contract
(``decision`` column out; ``archived`` + ``comments`` in).

Sibling module to ``tests.py`` (2,200+ lines when this feature shipped): same
fixture, same conventions — it reuses ``ReceiptsFixtureMixin`` / ``BulkMixin``
from ``tests`` so the production-shaped register rows and the shared helpers
exist here too, and ``manage.py test apps.receipts`` discovers both modules.

Run:  manage.py test apps.receipts
"""
from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APIClient

from .models import SlipComment, SlipReview
from .tests import BulkMixin, ReceiptsFixtureMixin, body_has_blob, sha

# The v4 export contract, pinned as a literal (NOT read from views.EXPORT_COLUMNS —
# a test that derives its expectation from the code under test asserts nothing).
EXPORT_HEADER_V4 = ['date', 'supplier', 'total', 'category', 'xero_status', 'status_group', 'journal_number',
                    'synced', 'to_process', 'archived', 'note', 'comments', 'filename', 'sha256', 'view_url']

ARCHIVED_400 = 'archived must be a boolean (true/false, 1/0, yes/no, on/off)'
SET_ARCHIVED_400 = 'set_archived must be a boolean (true/false, 1/0, yes/no, on/off)'


class ArchiveMixin(ReceiptsFixtureMixin):
    def review_url(self, name):
        return reverse('receipts:review', args=[self.NAMES[name]])

    def patch(self, name, payload, client=None, **kwargs):
        return (client or self.client).patch(self.review_url(name), payload, format='json', **kwargs)

    def seed_archived(self, *names, **extra):
        """Archive fixture slips directly (the filter tests seed state like tests.py does;
        the write-path tests below go through the API instead)."""
        for n in names:
            SlipReview.objects.update_or_create(
                sha256=self.NAMES[n],
                defaults={'archived': True, 'archived_at': timezone.now(), 'archived_by': 'seeder', **extra},
            )

    def ids(self, **params):
        resp = self.client.get(reverse('receipts:list'), {'ids_only': '1', **params})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()


# --------------------------------------------------------------------------- #
# A1. Defaults: a slip is never born archived
# --------------------------------------------------------------------------- #
class ArchivedDefaultTests(ArchiveMixin, TestCase):
    def test_fresh_slipreview_defaults_to_unarchived(self):
        rv = SlipReview.objects.create(sha256=self.NAMES['a_fy26'])
        rv.refresh_from_db()
        self.assertFalse(rv.archived)
        self.assertIsNone(rv.archived_at)
        self.assertEqual(rv.archived_by, '')

    def test_slip_with_no_review_row_reports_archived_false_in_list(self):
        self.assertFalse(SlipReview.objects.exists())
        data = self.list_(page_size=200)
        self.assertEqual(data['count'], 12, 'no review rows exist, so nothing may be hidden')
        for row in data['results']:
            self.assertIs(row['review']['archived'], False, row['sha256'])
            self.assertIsNone(row['review']['archived_at'], row['sha256'])
            self.assertEqual(row['review']['archived_by'], '', row['sha256'])

    def test_detail_review_dict_carries_the_archive_keys_when_no_row_exists(self):
        resp = self.client.get(reverse('receipts:detail', args=[self.NAMES['b_fy26_auto']]))
        self.assertEqual(resp.status_code, 200)
        review = resp.json()['review']
        for key, expected in (('archived', False), ('archived_at', None), ('archived_by', '')):
            self.assertIn(key, review)
            self.assertEqual(review[key], expected)

    def test_review_row_created_by_a_non_archive_patch_stays_unarchived(self):
        # An ordinary review write must not accidentally flip the new column.
        resp = self.patch('a_fy26', {'to_process': True, 'note': 'look at this'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertIs(resp.json()['archived'], False)
        self.assertEqual(self.list_(page_size=200)['count'], 12, 'a plain review write hid a row from the list')


# --------------------------------------------------------------------------- #
# A2. The three-way `archived` filter (list mode + totals)
# --------------------------------------------------------------------------- #
class ArchivedFilterTests(ArchiveMixin, TestCase):
    def test_default_list_excludes_archived_and_the_sum_drops_by_the_archived_total(self):
        self.seed_archived('b_fy26_auto')  # total 1234.56 — the only slip whose absence moves the sum
        data = self.list_(page_size=200)
        self.assertEqual(data['count'], 11)
        self.assertNotIn('b_fy26_auto', self.names(data))
        # Not just the count: the working-set SUM must drop by exactly the archived row's total.
        expected = self.FIXTURE_SUM - Decimal('1234.56')
        self.assertEqual(Decimal(data['totals']['sum_total']), expected)
        self.assertEqual(data['totals'], {'count': 11, 'sum_total': '990.99'})

    def test_archived_true_returns_only_archived_rows_and_their_totals(self):
        self.seed_archived('b_fy26_auto', 'j_ts_null')
        data = self.list_(archived='true', page_size=200)
        self.assertEqual(self.names(data), {'b_fy26_auto', 'j_ts_null'})
        self.assertEqual(data['totals'], {'count': 2, 'sum_total': '1249.56'})  # 1234.56 + 15.00

    def test_archived_all_returns_both_and_restores_the_full_sum(self):
        self.seed_archived('b_fy26_auto')
        data = self.list_(archived='all', page_size=200)
        self.assertEqual(data['count'], 12)
        self.assertEqual(Decimal(data['totals']['sum_total']), self.FIXTURE_SUM)

    def test_archived_false_behaves_exactly_like_the_default(self):
        self.seed_archived('b_fy26_auto')
        default = self.list_(page_size=200)
        explicit = self.list_(archived='false', page_size=200)
        self.assertEqual(self.shas(explicit), self.shas(default))
        self.assertEqual(explicit['totals'], default['totals'])

    def test_hostile_and_documented_archived_values_land_on_the_right_branch(self):
        # Table-driven: absent/false-y/unrecognised -> EXCLUDE; true-y -> ONLY; 'all' (any case) -> BOTH.
        self.seed_archived('b_fy26_auto')
        exclude_names = set(self.NAMES) - {'b_fy26_auto'}
        table = [
            # (param value, expected fixture names)
            ('banana', exclude_names), ('2', exclude_names), ('', exclude_names),
            ('false', exclude_names), ('0', exclude_names), ('no', exclude_names),
            ('off', exclude_names), ('OFF', exclude_names), ('None', exclude_names),
            ('archived', exclude_names), ('-1', exclude_names), ('tru', exclude_names),
            ('TRUE', {'b_fy26_auto'}), ('true', {'b_fy26_auto'}), ('1', {'b_fy26_auto'}),
            ('yes', {'b_fy26_auto'}), ('Yes', {'b_fy26_auto'}), ('on', {'b_fy26_auto'}),
            ('all', set(self.NAMES)), ('All', set(self.NAMES)), ('ALL', set(self.NAMES)),
            (' all ', set(self.NAMES)),
        ]
        for value, expected in table:
            data = self.list_(archived=value, page_size=200)
            self.assertEqual(self.names(data), expected, f'archived={value!r}')
            self.assertEqual(data['totals']['count'], len(expected), f'totals.count for archived={value!r}')

    def test_archived_param_with_nul_byte_is_exclude_not_all_and_never_500(self):
        # 'a\x00ll' must NOT be read as 'all', and the NUL must never reach a SQL bind.
        self.seed_archived('b_fy26_auto')
        client = APIClient(raise_request_exception=False)
        client.credentials(**self.client._credentials)
        resp = client.get(reverse('receipts:list'), {'archived': 'a\x00ll', 'page_size': 200})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['count'], 11, "archived='a\\x00ll' must land on the exclude branch")

    def test_unarchived_review_row_is_never_excluded(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], archived=False, decision='CAPTURE')
        self.assertEqual(self.list_(page_size=200)['count'], 12)
        self.assertEqual(self.list_(archived='true')['count'], 0)

    def test_no_review_rows_at_all_empty_array_edge(self):
        # exclude branch binds an EMPTY array: `not (sha = any('{}'))` must return everything.
        self.assertEqual(self.list_(page_size=200)['count'], 12)
        self.assertEqual(self.list_(archived='true')['totals'], {'count': 0, 'sum_total': '0.00'})
        self.assertEqual(self.list_(archived='all')['count'], 12)

    def test_orphaned_archived_review_row_hides_nothing_and_matches_nothing(self):
        ghost = sha(999)  # archived review whose sha is NOT in the register
        SlipReview.objects.create(sha256=ghost, archived=True, archived_by='seeder')
        self.assertEqual(self.list_(page_size=200)['count'], 12)
        self.assertEqual(self.list_(archived='true')['count'], 0)

    def test_archived_composes_with_other_filters_by_and(self):
        # Hardware slips: a_fy26, d_fy26_nix, e_skip_neg. Archive a + d.
        self.seed_archived('a_fy26', 'd_fy26_nix')
        self.assertEqual(self.names(self.list_(category='Hardware')), {'e_skip_neg'})
        self.assertEqual(self.names(self.list_(category='Hardware', archived='true')), {'a_fy26', 'd_fy26_nix'})
        self.assertEqual(self.names(self.list_(category='Hardware', archived='all')),
                         {'a_fy26', 'd_fy26_nix', 'e_skip_neg'})

    def test_archiving_beats_the_to_process_filter(self):
        # A flagged slip that gets archived leaves the to_process working list too.
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True,
                                  archived=True, archived_at=timezone.now(), archived_by='seeder')
        self.assertEqual(self.list_(to_process='true')['count'], 0,
                         'an archived slip surfaced in the default to_process working list')
        self.assertEqual(self.names(self.list_(to_process='true', archived='true')), {'a_fy26'})
        self.assertEqual(self.names(self.list_(to_process='true', archived='all')), {'a_fy26'})


# --------------------------------------------------------------------------- #
# A3. ids_only honours the same exclusion (select-all must not touch archived)
# --------------------------------------------------------------------------- #
class ArchivedIdsOnlyTests(ArchiveMixin, TestCase):
    def test_ids_only_default_excludes_archived(self):
        self.seed_archived('b_fy26_auto', 'j_ts_null')
        data = self.ids()
        self.assertEqual(data['count'], 10)
        self.assertEqual(len(data['sha256s']), 10)
        self.assertNotIn(self.NAMES['b_fy26_auto'], data['sha256s'])
        self.assertNotIn(self.NAMES['j_ts_null'], data['sha256s'])
        self.assertEqual(data['sha256s'], self.shas(self.list_(page_size=200)),
                         'ids_only and row mode disagree about the archived exclusion')

    def test_ids_only_archived_true_and_all(self):
        self.seed_archived('b_fy26_auto', 'j_ts_null')
        only = self.ids(archived='true')
        # default ordering -slip_ts nulls last: b (2025-12-09) then j (NULL slip_ts)
        self.assertEqual(only['sha256s'], [self.NAMES['b_fy26_auto'], self.NAMES['j_ts_null']])
        self.assertEqual(only['count'], 2)
        self.assertEqual(self.ids(archived='all')['count'], 12)
        self.assertEqual(set(self.ids(archived='all')['sha256s']), set(self.NAMES.values()))

    def test_select_all_then_bulk_archive_never_rearchives_archived_rows(self):
        # The console flow: ids_only (select all N) -> bulk set_archived. The archived rows
        # must not be in the selection, so their stamps must survive untouched.
        self.seed_archived('b_fy26_auto')
        original = SlipReview.objects.get(sha256=self.NAMES['b_fy26_auto'])
        selection = self.ids()['sha256s']
        resp = self.client.post(reverse('receipts:bulk'),
                                {'sha256s': selection, 'set_archived': 'true'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 11)
        self.assertEqual(self.list_(page_size=200)['count'], 0, 'archive-all left rows in the working list')
        untouched = SlipReview.objects.get(sha256=self.NAMES['b_fy26_auto'])
        self.assertEqual(untouched.archived_at, original.archived_at)
        self.assertEqual(untouched.archived_by, 'seeder', "select-all silently re-archived an archived row")


# --------------------------------------------------------------------------- #
# A4. PATCH review: archive / unarchive / rejection semantics
# --------------------------------------------------------------------------- #
class ArchivePatchTests(ArchiveMixin, TestCase):
    def test_patch_archive_sets_flag_and_stamps_who_and_when(self):
        before = timezone.now()
        resp = self.patch('a_fy26', {'archived': True})
        after = timezone.now()
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertIs(body['archived'], True)
        self.assertEqual(body['archived_by'], 'reviewer', 'archived_by must be the authenticated user')
        stamp = dt.datetime.fromisoformat(body['archived_at'])
        self.assertTrue(before <= stamp <= after, f'archived_at {stamp} outside the request window')
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertTrue(rv.archived)
        self.assertEqual(rv.archived_at, stamp)
        self.assertEqual(rv.archived_by, 'reviewer')
        self.assertEqual(rv.updated_by, 'reviewer')

    def test_patch_unarchive_clears_both_stamps(self):
        self.seed_archived('a_fy26')
        resp = self.patch('a_fy26', {'archived': False})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertIs(body['archived'], False)
        self.assertIsNone(body['archived_at'])
        self.assertEqual(body['archived_by'], '')
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertFalse(rv.archived)
        self.assertIsNone(rv.archived_at)
        self.assertEqual(rv.archived_by, '')

    def test_patch_archived_coercions_match_the_documented_vocabulary(self):
        for raw in (True, 'true', 'TRUE', 1, '1', 'yes', 'Yes', 'on', 'ON'):
            SlipReview.objects.all().delete()
            resp = self.patch('a_fy26', {'archived': raw})
            self.assertEqual(resp.status_code, 200, (raw, resp.content[:200]))
            rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
            self.assertTrue(rv.archived, f'archived={raw!r} must archive')
            self.assertIsNotNone(rv.archived_at, raw)
            self.assertEqual(rv.archived_by, 'reviewer', raw)
        for raw in (False, 'false', 'FALSE', 0, '0', 'no', 'No', 'off', 'OFF', None):
            self.seed_archived('a_fy26')
            resp = self.patch('a_fy26', {'archived': raw})
            self.assertEqual(resp.status_code, 200, (raw, resp.content[:200]))
            rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
            self.assertFalse(rv.archived, f'archived={raw!r} must unarchive')
            self.assertIsNone(rv.archived_at, raw)
            self.assertEqual(rv.archived_by, '', raw)

    def test_patch_unrecognised_archived_is_400_with_the_documented_message_and_writes_nothing(self):
        for raw in ('banana', '2', 2, 'yess', '', [], {}, 'true false', 'all', 'null', [True], {'v': True}):
            resp = self.patch('a_fy26', {'archived': raw})
            self.assertEqual(resp.status_code, 400, (raw, resp.content[:200]))
            self.assertEqual(resp.json()['detail'], ARCHIVED_400, raw)
        self.assertFalse(SlipReview.objects.exists(), 'a rejected archived value wrote a review row')

    def test_patch_archived_all_is_a_filter_word_not_a_write_value(self):
        # 'all' is vocabulary for the QUERY param; a client leaking it into the WRITE
        # must get a 400, never a silent archive/unarchive.
        self.seed_archived('a_fy26')
        resp = self.patch('a_fy26', {'archived': 'all'})
        self.assertEqual(resp.status_code, 400)
        self.assertTrue(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).archived,
                        "a rejected archived='all' flipped the stored flag")

    def test_patch_bad_archived_with_valid_siblings_is_atomic(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], decision='CAPTURE', note='keep')
        resp = self.patch('a_fy26', {'decision': 'PERSONAL', 'note': 'clobber', 'archived': 'maybe'})
        self.assertEqual(resp.status_code, 400, resp.content[:300])
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertEqual(rv.decision, 'CAPTURE', 'a 400 PATCH applied its valid sibling fields')
        self.assertEqual(rv.note, 'keep')
        self.assertFalse(rv.archived)

    def test_archive_alone_does_not_disturb_to_process_decision_or_note(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True,
                                  decision='CAPTURE', note='keep me', updated_by='mc')
        resp = self.patch('a_fy26', {'archived': True})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertIs(body['to_process'], True)
        self.assertEqual(body['decision'], 'CAPTURE')
        self.assertEqual(body['note'], 'keep me')
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertTrue(rv.to_process)
        self.assertEqual(rv.decision, 'CAPTURE')
        self.assertEqual(rv.note, 'keep me')
        self.assertTrue(rv.archived)

    def test_rearchive_after_unarchive_never_shows_a_stale_stamp(self):
        self.patch('a_fy26', {'archived': True})
        first = SlipReview.objects.get(sha256=self.NAMES['a_fy26']).archived_at
        self.patch('a_fy26', {'archived': False})
        self.assertIsNone(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).archived_at)
        self.patch('a_fy26', {'archived': True})
        second = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertGreater(second.archived_at, first, 're-archive kept the stale first timestamp')
        self.assertEqual(second.archived_by, 'reviewer')

    def test_patch_omitting_decision_leaves_a_stored_decision_intact(self):
        # THE data-loss guard: the console stopped sending `decision`, so an archive (or any
        # partial PATCH) must never blank a previously-stored decision.
        self.assertEqual(self.patch('a_fy26', {'decision': 'MEAL_SKIP'}).status_code, 200)
        resp = self.patch('a_fy26', {'archived': True, 'to_process': True, 'note': 'archived it'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['decision'], 'MEAL_SKIP', 'a decision-less PATCH blanked the decision')
        self.assertEqual(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).decision, 'MEAL_SKIP')
        self.patch('a_fy26', {'archived': False})
        self.assertEqual(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).decision, 'MEAL_SKIP')

    def test_form_encoded_archive_patch_works(self):
        resp = self.client.patch(self.review_url('a_fy26'), data='archived=true',
                                 content_type='application/x-www-form-urlencoded')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertTrue(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).archived)

    def test_archived_slip_stays_patchable_and_detail_visible(self):
        # Archiving hides a row from the LIST, not from direct addressing.
        self.seed_archived('a_fy26')
        self.assertEqual(self.client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).status_code, 200)
        self.assertEqual(self.patch('a_fy26', {'note': 'still editable'}).status_code, 200)
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertTrue(rv.archived, 'a note-only PATCH silently unarchived the slip')


# --------------------------------------------------------------------------- #
# A5. Bulk set_archived
# --------------------------------------------------------------------------- #
class ArchiveBulkTests(BulkMixin, TestCase):
    def test_bulk_archive_mixed_batch_counts_updated_and_reports_unknown(self):
        a, b, c, ghost = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto'], self.NAMES['c_fy25_pend'], sha(999)
        before = timezone.now()
        resp = self.bulk({'sha256s': [a, b, ghost, c], 'set_archived': 'true'})
        after = timezone.now()
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 3, 'commented': 0, 'unknown': [ghost]})
        for s in (a, b, c):
            rv = SlipReview.objects.get(sha256=s)
            self.assertTrue(rv.archived, s)
            self.assertEqual(rv.archived_by, 'reviewer', s)
            self.assertTrue(before <= rv.archived_at <= after, s)
        self.assertFalse(SlipReview.objects.filter(sha256=ghost).exists(), 'bulk wrote a review for an unknown sha')
        self.assertEqual(self.list_(page_size=200)['count'], 9)

    def test_bulk_unarchive_clears_stamps_across_a_mixed_batch(self):
        a, b, d = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto'], self.NAMES['d_fy26_nix']
        SlipReview.objects.create(sha256=a, archived=True, archived_at=timezone.now(), archived_by='seeder')
        SlipReview.objects.create(sha256=b, archived=False)  # already unarchived
        # d has no review row at all
        resp = self.bulk({'sha256s': [a, b, d], 'set_archived': 'false'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 3, 'every known target counts in updated, even no-ops')
        for s in (a, b, d):
            rv = SlipReview.objects.get(sha256=s)
            self.assertFalse(rv.archived, s)
            self.assertIsNone(rv.archived_at, s)
            self.assertEqual(rv.archived_by, '', s)
        self.assertEqual(self.list_(page_size=200)['count'], 12)

    def test_bulk_bad_set_archived_is_400_before_any_write(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        for raw in ('banana', '2', 2, 'yess', '', [], {}, 'all', 'true false', [True]):
            resp = self.bulk({'sha256s': [a, b], 'set_archived': raw, 'note': 'sibling'})
            self.assertEqual(resp.status_code, 400, (raw, resp.content[:200]))
            self.assertEqual(resp.json()['detail'], SET_ARCHIVED_400, raw)
        self.assert_nothing_written()

    def test_bulk_bad_set_archived_never_flips_an_existing_flag(self):
        a = self.NAMES['a_fy26']
        SlipReview.objects.create(sha256=a, archived=True, archived_at=timezone.now(), archived_by='seeder')
        resp = self.bulk({'sha256s': [a], 'set_archived': 'maybe'})
        self.assertEqual(resp.status_code, 400)
        rv = SlipReview.objects.get(sha256=a)
        self.assertTrue(rv.archived, 'a rejected set_archived silently unarchived the row')
        self.assertEqual(rv.archived_by, 'seeder')

    def test_bulk_set_archived_null_is_an_explicit_clear(self):
        a = self.NAMES['a_fy26']
        SlipReview.objects.create(sha256=a, archived=True, archived_at=timezone.now(), archived_by='seeder')
        resp = self.bulk({'sha256s': [a], 'set_archived': None})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        rv = SlipReview.objects.get(sha256=a)
        self.assertFalse(rv.archived)
        self.assertIsNone(rv.archived_at)
        self.assertEqual(rv.archived_by, '')

    def test_bulk_archive_only_leaves_decisions_intact(self):
        # Bulk side of the data-loss guard: archiving a batch must not blank stored decisions.
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        SlipReview.objects.create(sha256=a, decision='CAPTURE', note='keep', to_process=True)
        resp = self.bulk({'sha256s': [a, b], 'set_archived': 'true'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        rv = SlipReview.objects.get(sha256=a)
        self.assertEqual(rv.decision, 'CAPTURE', 'bulk set_archived blanked a stored decision')
        self.assertEqual(rv.note, 'keep')
        self.assertTrue(rv.to_process)
        self.assertTrue(rv.archived)

    def test_bulk_archive_with_comment_applies_both(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        resp = self.bulk({'sha256s': [a, b], 'set_archived': 'true', 'comment': 'batch archived'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 2, 'commented': 2, 'unknown': []})
        self.assertEqual(SlipComment.objects.filter(text='batch archived').count(), 2)

    def test_bulk_all_unknown_set_archived_is_200_and_writes_nothing(self):
        ghosts = [sha(700), sha(701)]
        resp = self.bulk({'sha256s': ghosts, 'set_archived': 'true'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 0, 'commented': 0, 'unknown': ghosts})
        self.assert_nothing_written()


# --------------------------------------------------------------------------- #
# A6. Export: filter, header contract, comments count — csv AND xlsx
# --------------------------------------------------------------------------- #
class ArchiveExportTests(ArchiveMixin, TestCase):
    def rec_by_sha(self, rows):
        sidx = rows[0].index('sha256')
        return {r[sidx]: dict(zip(rows[0], r)) for r in rows[1:]}

    def test_csv_header_is_the_v4_contract(self):
        rows = self.csv_rows(self.export())
        self.assertEqual(rows[0], EXPORT_HEADER_V4)
        self.assertNotIn('decision', rows[0])
        self.assertIn('archived', rows[0])
        self.assertIn('comments', rows[0])

    def test_default_csv_export_contains_no_archived_row(self):
        self.seed_archived('b_fy26_auto')
        rows = self.csv_rows(self.export())
        self.assertEqual(len(rows), 11 + 1)
        self.assertNotIn(self.NAMES['b_fy26_auto'], self.rec_by_sha(rows))
        for rec in self.rec_by_sha(rows).values():
            self.assertEqual(rec['archived'], 'False')

    def test_csv_export_archived_true_and_all(self):
        self.seed_archived('b_fy26_auto')
        only = self.csv_rows(self.export(archived='true'))
        self.assertEqual(len(only), 1 + 1)
        self.assertEqual(self.rec_by_sha(only)[self.NAMES['b_fy26_auto']]['archived'], 'True')
        both = self.csv_rows(self.export(archived='all'))
        self.assertEqual(len(both), 12 + 1)
        recs = self.rec_by_sha(both)
        self.assertEqual(recs[self.NAMES['b_fy26_auto']]['archived'], 'True')
        self.assertEqual(recs[self.NAMES['a_fy26']]['archived'], 'False')

    def test_csv_comments_column_carries_the_real_comment_count(self):
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='one', author='mc')
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='two', author='mc')
        recs = self.rec_by_sha(self.csv_rows(self.export()))
        self.assertEqual(recs[self.NAMES['a_fy26']]['comments'], '2')
        self.assertEqual(recs[self.NAMES['d_fy26_nix']]['comments'], '0')

    def test_xlsx_export_header_filter_and_cell_types(self):
        from openpyxl import load_workbook
        self.seed_archived('b_fy26_auto')
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='one', author='mc')
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='two', author='mc')
        resp = self.export(format='xlsx')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        rows = list(load_workbook(io.BytesIO(resp.content)).active.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), EXPORT_HEADER_V4)
        self.assertEqual(len(rows) - 1, 11, 'default xlsx export must exclude archived rows')
        sidx = EXPORT_HEADER_V4.index('sha256')
        recs = {r[sidx]: dict(zip(EXPORT_HEADER_V4, r)) for r in rows[1:]}
        self.assertNotIn(self.NAMES['b_fy26_auto'], recs)
        self.assertIs(recs[self.NAMES['a_fy26']]['archived'], False)
        self.assertEqual(recs[self.NAMES['a_fy26']]['comments'], 2)
        # archived=all brings the archived row back, flagged
        resp_all = self.export(format='xlsx', archived='all')
        rows_all = list(load_workbook(io.BytesIO(resp_all.content)).active.iter_rows(values_only=True))
        self.assertEqual(len(rows_all) - 1, 12)
        recs_all = {r[sidx]: dict(zip(EXPORT_HEADER_V4, r)) for r in rows_all[1:]}
        self.assertIs(recs_all[self.NAMES['b_fy26_auto']]['archived'], True)

    def test_archived_all_export_still_never_leaks_file_bytes(self):
        self.seed_archived('c_fy25_pend')  # the blob row
        resp = self.export(archived='all')
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(body_has_blob(resp.content.decode('utf-8', 'replace')))


# --------------------------------------------------------------------------- #
# A7. Auth: every archive-touching path is 401 for anonymous callers
# --------------------------------------------------------------------------- #
class ArchiveAuthTests(ArchiveMixin, TestCase):
    def archive_requests(self, client):
        s = self.NAMES['a_fy26']
        return [
            ('list?archived=true', client.get(reverse('receipts:list'), {'archived': 'true'})),
            ('review PATCH archived', client.patch(reverse('receipts:review', args=[s]),
                                                   {'archived': True}, format='json')),
            ('bulk set_archived', client.post(reverse('receipts:bulk'),
                                              {'sha256s': [s], 'set_archived': 'true'}, format='json')),
            ('export?archived=all', client.get(reverse('receipts:export'), {'archived': 'all'})),
        ]

    def test_anonymous_caller_gets_401_on_every_archive_path_and_writes_nothing(self):
        self.seed_archived('b_fy26_auto')
        for name, resp in self.archive_requests(self.anon):
            self.assertEqual(resp.status_code, 401, f'{name}: expected exactly 401, got {resp.status_code}')
            text = resp.content.decode('utf-8', 'replace')
            for s in self.NAMES.values():
                self.assertNotIn(s, text, f'{name}: 401 body leaked a sha256')
            self.assertNotIn('Builders', text, f'{name}: 401 body leaked supplier data')
        self.assertEqual(SlipReview.objects.count(), 1, 'an anonymous archive write landed')
        self.assertTrue(SlipReview.objects.get(sha256=self.NAMES['b_fy26_auto']).archived)

    def test_garbage_bearer_gets_401_on_every_archive_path(self):
        bad = APIClient()
        bad.credentials(HTTP_AUTHORIZATION='Bearer not-a-real-token')
        for name, resp in self.archive_requests(bad):
            self.assertEqual(resp.status_code, 401, f'{name}: expected exactly 401, got {resp.status_code}')
        self.assertFalse(SlipReview.objects.exists())
