"""
Adversarial tests for apps.receipts (Audit -> Receipts review over whatsapp.klikk_slips).

``whatsapp.klikk_slips`` is NOT a Django model (external WhatsApp sync), so the
test database does not contain it. Every test class below creates it with raw
DDL inside the class-level transaction (Postgres DDL is transactional, so the
table is rolled back with the rest of the class fixture) and populates it with
production-shaped rows: NULL slip_ts, NULL ocr, non-numeric / negative / missing
``ocr->>'total'``, ``MATCHED (auto-recon ...)`` status variants, multi-line
journals, journal numbers that collide across Xero tenants, and a real byte blob
that must never surface in any JSON/CSV/XLSX response.

Run:  manage.py test apps.receipts
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import json
from decimal import Decimal
from unittest import mock
from urllib.parse import parse_qs, urlparse

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase, override_settings
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken, RefreshToken

from apps.audit.slip_view import slip_signature
from apps.xero.xero_core.models import XeroTenant
from apps.xero.xero_data.models import XeroJournals
from apps.xero.xero_metadata.models import XeroAccount, XeroContacts

from .models import DECISION_VALUES, SlipComment, SlipReview
from .services import KLIKK_TENANT_ID, SLIP_TZ, fy_label, fy_range

UTC = dt.timezone.utc
User = get_user_model()

KLIKK_SLIPS_DDL = """
CREATE SCHEMA IF NOT EXISTS whatsapp;
CREATE TABLE IF NOT EXISTS whatsapp.klikk_slips (
  sha256 text primary key,
  slip_ts timestamptz,
  filename text,
  source text,
  mime_ext text,
  byte_size bigint,
  file_bytes bytea,
  xero_status text not null default 'PENDING',
  xero_detail text,
  imported_at timestamptz not null default now(),
  synced_to_xero boolean not null default false,
  ocr jsonb,
  search_tsv tsvector,
  journal_number integer,
  xero_org text
);
"""

# A recognisable blob. If any byte of this shows up in a list/detail/export body, the
# endpoint is leaking ``file_bytes``.
BLOB = b'\xff\xd8\xff\xe0LEAKED_FILE_BYTES_SENTINEL_0123456789' * 20
BLOB_MARKERS = ('LEAKED_FILE_BYTES_SENTINEL', 'file_bytes', '/9j/4', '\\xff\\xd8')


def sha(n: int) -> str:
    """Deterministic 64-hex sha256-shaped key."""
    return f'{n:064x}'


def ts(y, m, d, hh=0, mm=0, ss=0, tz=UTC) -> dt.datetime:
    return dt.datetime(y, m, d, hh, mm, ss, tzinfo=tz)


def create_slips_table():
    with connection.cursor() as cur:
        cur.execute(KLIKK_SLIPS_DDL)


def insert_slip(sha256, *, slip_ts=None, filename='PHOTO.jpg', source='chat_export', mime_ext='jpg',
                byte_size=None, file_bytes=None, xero_status='PENDING', xero_detail=None, synced=False,
                ocr=None, journal_number=None, xero_org=None, tsv_text=None, ocr_raw=None):
    """
    Insert one production-shaped row. ``ocr`` is a dict (json-dumped); ``ocr_raw`` lets a
    test hand in a raw jsonb literal (e.g. a JSON string / array instead of an object).
    ``search_tsv`` is populated from ``tsv_text`` (or supplier+filename) to mirror the
    sync's trigger, since the ``q`` filter reads it.
    """
    if ocr_raw is None:
        ocr_raw = json.dumps(ocr) if ocr is not None else None
    if tsv_text is None:
        sup = (ocr or {}).get('supplier') if isinstance(ocr, dict) else None
        tsv_text = ' '.join(x for x in (sup, filename) if x)
    with connection.cursor() as cur:
        cur.execute(
            """
            insert into whatsapp.klikk_slips
              (sha256, slip_ts, filename, source, mime_ext, byte_size, file_bytes, xero_status, xero_detail,
               synced_to_xero, ocr, search_tsv, journal_number, xero_org)
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, to_tsvector('simple', %s), %s, %s)
            """,
            [sha256, slip_ts, filename, source, mime_ext,
             byte_size if byte_size is not None else (len(file_bytes) if file_bytes else None),
             file_bytes, xero_status, xero_detail, synced, ocr_raw, tsv_text or '', journal_number, xero_org],
        )


def make_tenant(tenant_id, name):
    return XeroTenant.objects.create(tenant_id=tenant_id, tenant_name=name)


def make_account(tenant, account_id, code, name, type_):
    # bulk_create: skips post_save. apps/xero/xero_metadata/signals.py writes
    # GlossaryRefreshRequest.organisation_id (IntegerField) from a UUID-string tenant pk and
    # raises ValueError on every ORM save() of an account/contact — pre-existing bug, not receipts'.
    return XeroAccount.objects.bulk_create(
        [XeroAccount(organisation=tenant, account_id=account_id, code=code, name=name, type=type_)])[0]


def make_contact(tenant, contacts_id, name):
    return XeroContacts.objects.bulk_create([XeroContacts(organisation=tenant, contacts_id=contacts_id, name=name)])[0]


def make_line(tenant, journal_number, account, amount, *, journal_id, date, description='', contact=None,
              journal_type='journal'):
    # journal_type: the mirror holds journal / manual_journal / system_journal / transaction in
    # production (the model's `choices` are not DB-enforced), and the join is discriminated by it.
    amount = Decimal(amount)
    return XeroJournals.objects.create(
        organisation=tenant, journal_id=journal_id, journal_number=journal_number, journal_type=journal_type,
        account=account, contact=contact, date=date, description=description,
        amount=amount, debit=amount if amount > 0 else Decimal('0'), credit=amount if amount < 0 else Decimal('0'),
        tax_amount=Decimal('0'),
    )


def body_has_blob(text: str) -> bool:
    return any(m in text for m in BLOB_MARKERS)


# --------------------------------------------------------------------------- #
# Shared fixture
# --------------------------------------------------------------------------- #
class ReceiptsFixtureMixin:
    """
    12 slips covering every production shape. Keyed by name for readability.

    name          sha      slip_ts (UTC)            status                         total      cat        journal / org     notes
    ----------    ------   ----------------------   ----------------------------   --------   --------   ---------------   -----
    a_fy26        1        2025-08-10 10:00         MATCHED                        259.00     Hardware   697 / Klikk       5-line journal, #697 collides with tenant B
    b_fy26_auto   2        2025-12-09 19:58         MATCHED (auto-recon 2026-08-17) 1234.56   Meals      962018 / Klikk    2-line journal, synced
    c_fy25_pend   3        2025-03-01 08:00         PENDING                        40.00      Fuel       -                 blob row (file_bytes set)
    d_fy26_nix    4        2026-02-14 12:00         NOT IN XERO                    99.99      Hardware   -
    e_skip_neg    5        2026-03-18 07:25         skipped (no/negative total)    -259.00    Hardware   -                 negative total => NULL numeric
    f_ocr_abc     6        2026-04-01 09:00         PENDING                        "abc"      Meals      -                 non-numeric
    g_ocr_empty   7        2026-04-02 09:00         PENDING                        ""         -          -                 empty string
    h_ocr_nototal 8        2026-04-03 09:00         PENDING                        (missing)  Meals      -                 no 'total' key
    i_ocr_null    9        2026-04-04 09:00         PENDING                        NULL ocr   -          -                 ocr IS NULL
    j_ts_null     10       NULL                     PENDING                        15.00      Meals      -                 slip_ts IS NULL (exists in prod)
    k_fy27        11       2026-07-15 06:00         PENDING                        500.00     Fuel       -                 FY27
    l_orphan_jn   12       2025-09-09 09:00         MATCHED                        77.00      Meals      555 / Klikk       journal number with NO lines anywhere
    """

    NAMES = {
        'a_fy26': sha(1), 'b_fy26_auto': sha(2), 'c_fy25_pend': sha(3), 'd_fy26_nix': sha(4),
        'e_skip_neg': sha(5), 'f_ocr_abc': sha(6), 'g_ocr_empty': sha(7), 'h_ocr_nototal': sha(8),
        'i_ocr_null': sha(9), 'j_ts_null': sha(10), 'k_fy27': sha(11), 'l_orphan_jn': sha(12),
    }
    NUMERIC_TOTALS = {'a_fy26': '259.00', 'b_fy26_auto': '1234.56', 'c_fy25_pend': '40.00', 'd_fy26_nix': '99.99',
                      'j_ts_null': '15.00', 'k_fy27': '500.00', 'l_orphan_jn': '77.00'}
    FIXTURE_SUM = sum(Decimal(v) for v in NUMERIC_TOTALS.values())  # 2225.55

    @classmethod
    def setUpTestData(cls):
        create_slips_table()
        S = cls.NAMES

        # ---- Xero fixtures: two tenants, journal #697 in BOTH (the cross-tenant trap) ----
        cls.klikk = make_tenant('tenant-klikk', 'Klikk (Pty) Ltd')
        cls.dfam = make_tenant('tenant-dfam', 'Dippenaar Family')

        k_bank = make_account(cls.klikk, 'acc-k-bank', '090', 'Investec Current', 'BANK')
        k_hw = make_account(cls.klikk, 'acc-k-hw', '429', 'Hardware & Consumables', 'EXPENSE')
        k_vat = make_account(cls.klikk, 'acc-k-vat', '820', 'VAT', 'CURRLIAB')
        k_meals = make_account(cls.klikk, 'acc-k-meals', '420', 'Entertainment', 'EXPENSE')
        d_bank = make_account(cls.dfam, 'acc-d-bank', '090', 'FNB Cheque', 'BANK')
        d_groc = make_account(cls.dfam, 'acc-d-groc', '450', 'Groceries', 'EXPENSE')

        builders = make_contact(cls.klikk, 'con-k-builders', 'Builders Warehouse')
        woolies = make_contact(cls.dfam, 'con-d-woolies', 'Woolworths')
        jonnys = make_contact(cls.klikk, 'con-k-jonnys', 'Jonnys Mozambican Restaurant')

        d697 = ts(2025, 8, 10, 0, 0)
        # Klikk #697: 5 lines (hardware 200 + hardware 25.22 + VAT 33.78 + bank -259 + a zero line)
        make_line(cls.klikk, 697, k_hw, '200.00', journal_id='k697-1', date=d697, description='Drill bits', contact=builders)
        make_line(cls.klikk, 697, k_hw, '25.22', journal_id='k697-2', date=d697, description='Screws', contact=builders)
        make_line(cls.klikk, 697, k_vat, '33.78', journal_id='k697-3', date=d697, description='VAT on Builders', contact=builders)
        make_line(cls.klikk, 697, k_bank, '-259.00', journal_id='k697-4', date=d697, description='Builders Warehouse card', contact=builders)
        make_line(cls.klikk, 697, k_hw, '0.00', journal_id='k697-5', date=d697, description='rounding')
        # Dippenaar Family #697: totally different transaction (the collision)
        d697b = ts(2024, 2, 2, 0, 0)
        make_line(cls.dfam, 697, d_groc, '812.40', journal_id='d697-1', date=d697b, description='Woolies groceries', contact=woolies)
        make_line(cls.dfam, 697, d_bank, '-812.40', journal_id='d697-2', date=d697b, description='Woolies groceries', contact=woolies)
        # Klikk #962018: 2 lines
        d962 = ts(2025, 12, 9, 0, 0)
        make_line(cls.klikk, 962018, k_meals, '1234.56', journal_id='k962018-1', date=d962, description='Team dinner', contact=jonnys)
        make_line(cls.klikk, 962018, k_bank, '-1234.56', journal_id='k962018-2', date=d962, description='Team dinner', contact=jonnys)

        # ---- Slips ----
        insert_slip(S['a_fy26'], slip_ts=ts(2025, 8, 10, 10, 0), filename='00000292-PHOTO-2025-08-10-12-00-00.jpg',
                    xero_status='MATCHED', xero_detail='journal_number=697', synced=True,
                    ocr={'supplier': 'Builders Warehouse', 'total': '259.00', 'category': 'Hardware',
                         'slip_date': '2025-08-10', 'payment_method': 'Card',
                         'items': [{'description': 'Drill bits', 'amount': '200.00'}, {'description': 'Screws', 'amount': '25.22'}]},
                    journal_number=697, xero_org='Klikk (Pty) Ltd')
        insert_slip(S['b_fy26_auto'], slip_ts=ts(2025, 12, 9, 19, 58, 48), filename='00000612-PHOTO-2025-12-09-21-58-48.jpg',
                    xero_status='MATCHED (auto-recon 2026-08-17)', xero_detail='Slip date=2025-12-09; Supplier=Jonnys Mozambican Restaurant;',
                    synced=True, ocr={'supplier': 'Jonnys Mozambican Restaurant', 'total': '1234.56', 'category': 'Meals',
                                      'slip_date': '2025-12-09', 'payment_method': 'Card'},
                    journal_number=962018, xero_org='Klikk (Pty) Ltd')
        insert_slip(S['c_fy25_pend'], slip_ts=ts(2025, 3, 1, 8, 0), filename='PHOTO-2025-03-01-10-00-00.jpg', source='archive5_2026',
                    xero_status='PENDING', file_bytes=BLOB,
                    ocr={'supplier': 'Engen Stellenbosch', 'total': '40.00', 'category': 'Fuel', 'slip_date': '2025-03-01'})
        insert_slip(S['d_fy26_nix'], slip_ts=ts(2026, 2, 14, 12, 0), filename='PHOTO-2026-02-14-14-00-00.jpg',
                    xero_status='NOT IN XERO', ocr={'supplier': 'Takealot', 'total': '99.99', 'category': 'Hardware'})
        insert_slip(S['e_skip_neg'], slip_ts=ts(2026, 3, 18, 7, 25, 45), filename='PHOTO-2026-03-18-09-25-45 3.jpg',
                    xero_status='skipped (no/negative total)', xero_detail='refund',
                    ocr={'supplier': 'Builders Warehouse', 'total': '-259.00', 'category': 'Hardware',
                         'payment_method': 'Card (Tap) - Refund'})
        insert_slip(S['f_ocr_abc'], slip_ts=ts(2026, 4, 1, 9, 0), filename='PHOTO-2026-04-01.jpg',
                    ocr={'supplier': 'Spur', 'total': 'abc', 'category': 'Meals'})
        insert_slip(S['g_ocr_empty'], slip_ts=ts(2026, 4, 2, 9, 0), filename='PHOTO-2026-04-02.jpg',
                    ocr={'supplier': 'Unknown', 'total': ''})
        insert_slip(S['h_ocr_nototal'], slip_ts=ts(2026, 4, 3, 9, 0), filename='PHOTO-2026-04-03.jpg',
                    ocr={'supplier': 'Nandos', 'category': 'Meals'})
        insert_slip(S['i_ocr_null'], slip_ts=ts(2026, 4, 4, 9, 0), filename='PHOTO-2026-04-04.pdf', mime_ext='pdf', ocr=None)
        insert_slip(S['j_ts_null'], slip_ts=None, filename='no-timestamp.jpg',
                    ocr={'supplier': 'KFC', 'total': '15.00', 'category': 'Meals'})
        insert_slip(S['k_fy27'], slip_ts=ts(2026, 7, 15, 6, 0), filename='PHOTO-2026-07-15.jpg',
                    ocr={'supplier': 'Shell', 'total': '500.00', 'category': 'Fuel'})
        insert_slip(S['l_orphan_jn'], slip_ts=ts(2025, 9, 9, 9, 0), filename='PHOTO-2025-09-09.jpg',
                    xero_status='MATCHED', synced=True, journal_number=555, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Wimpy', 'total': '77.00', 'category': 'Meals'})

        cls.user = User.objects.create_user(username='reviewer', email='reviewer@example.com', password='pw-irrelevant')

    # ---- helpers ----
    def setUp(self):
        # Every receipts endpoint now requires authentication, so the shared client
        # authenticates by default; tests that probe the anonymous path use self.anon
        # (or clear credentials explicitly).
        self.client = APIClient()
        self.auth()
        self.anon = APIClient()

    def auth(self):
        token = str(RefreshToken.for_user(self.user).access_token)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

    def list_(self, **params):
        resp = self.client.get(reverse('receipts:list'), params)
        self.assertEqual(resp.status_code, 200, resp.content[:500])
        return resp.json()

    def shas(self, data) -> list[str]:
        return [r['sha256'] for r in data['results']]

    def names(self, data) -> set[str]:
        inv = {v: k for k, v in self.NAMES.items()}
        return {inv.get(s, s) for s in self.shas(data)}

    def export(self, **params):
        return self.client.get(reverse('receipts:export'), params)

    def csv_rows(self, resp):
        return list(csv.reader(io.StringIO(resp.content.decode('utf-8'))))


# --------------------------------------------------------------------------- #
# 1. List: shape, filters, pagination, totals
# --------------------------------------------------------------------------- #
class ListShapeTests(ReceiptsFixtureMixin, TestCase):
    def test_unfiltered_list_returns_every_slip_once(self):
        data = self.list_(page_size=200)
        self.assertEqual(data['count'], 12)
        self.assertEqual(len(data['results']), 12)
        self.assertEqual(len(set(self.shas(data))), 12, 'duplicate rows in list')
        self.assertEqual(data['num_pages'], 1)
        self.assertEqual(data['page'], 1)

    def test_row_shape_and_no_file_bytes(self):
        data = self.list_(page_size=200)
        row = next(r for r in data['results'] if r['sha256'] == self.NAMES['a_fy26'])
        for key in ('sha256', 'slip_ts', 'filename', 'source', 'mime_ext', 'mime', 'is_pdf', 'byte_size', 'xero_status',
                    'status_group', 'xero_detail', 'synced_to_xero', 'journal_number', 'xero_org', 'fy', 'supplier',
                    'total', 'category', 'slip_date', 'payment_method', 'view_url', 'journal', 'review', 'comment_count'):
            self.assertIn(key, row)
        self.assertNotIn('file_bytes', row)
        self.assertNotIn('ocr', row, 'list must not ship the full ocr blob per row')
        self.assertEqual(row['total'], '259.00')
        self.assertEqual(row['fy'], 'FY26')
        self.assertEqual(row['status_group'], 'MATCHED')
        self.assertEqual(row['mime'], 'image/jpeg')
        self.assertFalse(row['is_pdf'])

    def test_blob_row_does_not_leak_bytes_in_list_or_detail(self):
        resp = self.client.get(reverse('receipts:list'), {'page_size': 200})
        self.assertFalse(body_has_blob(resp.content.decode('utf-8', 'replace')), 'file_bytes leaked into list JSON')
        # The blob row is ~1.2KB; the whole list must be well under raw-blob size * rows.
        detail = self.client.get(reverse('receipts:detail', args=[self.NAMES['c_fy25_pend']]))
        self.assertEqual(detail.status_code, 200)
        text = detail.content.decode('utf-8', 'replace')
        self.assertFalse(body_has_blob(text), 'file_bytes leaked into detail JSON')
        self.assertNotIn('file_bytes', detail.json())
        self.assertEqual(detail.json()['byte_size'], len(BLOB))

    def test_pdf_row_mime(self):
        data = self.list_(page_size=200)
        row = next(r for r in data['results'] if r['sha256'] == self.NAMES['i_ocr_null'])
        self.assertTrue(row['is_pdf'])
        self.assertEqual(row['mime'], 'application/pdf')
        self.assertIsNone(row['total'])
        self.assertIsNone(row['supplier'])

    def test_totals_cover_whole_filter_set_not_current_page(self):
        data = self.list_(page_size=3, page=2)
        self.assertEqual(data['count'], 12)
        self.assertEqual(data['num_pages'], 4)
        self.assertEqual(len(data['results']), 3)
        self.assertEqual(data['totals']['count'], 12)
        self.assertEqual(Decimal(data['totals']['sum_total']), self.FIXTURE_SUM)
        # Page 4 is the short page, totals must be unchanged.
        data4 = self.list_(page_size=3, page=4)
        self.assertEqual(len(data4['results']), 3)
        self.assertEqual(Decimal(data4['totals']['sum_total']), self.FIXTURE_SUM)

    def test_sum_total_ignores_non_numeric_and_negative_totals(self):
        # -259.00, "abc", "", missing key, NULL ocr must all be excluded, not crash.
        data = self.list_(page_size=200)
        self.assertEqual(Decimal(data['totals']['sum_total']), self.FIXTURE_SUM)
        by_sha = {r['sha256']: r for r in data['results']}
        self.assertIsNone(by_sha[self.NAMES['e_skip_neg']]['total'],
                          'negative total is rendered as a number; TOTAL_SQL regex was meant to null it')
        self.assertIsNone(by_sha[self.NAMES['f_ocr_abc']]['total'])
        self.assertIsNone(by_sha[self.NAMES['g_ocr_empty']]['total'])
        self.assertIsNone(by_sha[self.NAMES['h_ocr_nototal']]['total'])
        self.assertIsNone(by_sha[self.NAMES['i_ocr_null']]['total'])

    # --- pagination edges ---
    def test_pagination_edges(self):
        self.assertEqual(self.list_(page=0)['page'], 1)
        self.assertEqual(self.list_(page=-1)['page'], 1)
        self.assertEqual(self.list_(page='abc')['page'], 1)
        far = self.list_(page=99999, page_size=5)
        self.assertEqual(far['count'], 12)
        self.assertEqual(far['results'], [])
        self.assertEqual(far['num_pages'], 3)
        self.assertEqual(self.list_(page_size=0)['page_size'], 1)
        self.assertEqual(self.list_(page_size=-5)['page_size'], 1)
        self.assertEqual(self.list_(page_size=100000)['page_size'], 200)
        self.assertEqual(self.list_(page_size='x')['page_size'], 50)
        empty = self.list_(q='zzz-nothing-matches')
        self.assertEqual(empty['count'], 0)
        self.assertEqual(empty['num_pages'], 1)
        self.assertEqual(empty['results'], [])
        self.assertEqual(empty['totals'], {'count': 0, 'sum_total': '0.00'})

    def test_pages_partition_results_without_overlap_or_loss(self):
        seen = []
        for p in (1, 2, 3, 4, 5):
            seen += self.shas(self.list_(page=p, page_size=3, ordering='slip_ts'))
        self.assertEqual(len(seen), 12)
        self.assertEqual(len(set(seen)), 12)


class ListFilterTests(ReceiptsFixtureMixin, TestCase):
    def test_q_matches_supplier_via_tsv(self):
        self.assertEqual(self.names(self.list_(q='Builders')), {'a_fy26', 'e_skip_neg'})

    def test_q_matches_filename_ilike(self):
        self.assertEqual(self.names(self.list_(q='no-timestamp')), {'j_ts_null'})

    def test_q_is_case_insensitive_and_multiword(self):
        self.assertEqual(self.names(self.list_(q='jonnys mozambican')), {'b_fy26_auto'})
        self.assertEqual(self.names(self.list_(q='JONNYS')), {'b_fy26_auto'})

    def test_q_junk_and_sqli_do_not_500(self):
        for junk in ("'; DROP TABLE whatsapp.klikk_slips; --", "%", "_", "\\", "' or 1=1 --", "a & b | c ! d", "(((", "'"):
            data = self.list_(q=junk)
            self.assertIn('count', data)
        # Table still there (the DROP did not run).
        self.assertEqual(self.list_()['count'], 12)

    def test_synced_filter(self):
        self.assertEqual(self.names(self.list_(synced='true')), {'a_fy26', 'b_fy26_auto', 'l_orphan_jn'})
        self.assertEqual(self.list_(synced='false')['count'], 9)
        self.assertEqual(self.list_(synced='1')['count'], 3)
        self.assertEqual(self.list_(synced='maybe')['count'], 12, 'junk synced value must be ignored')

    def test_status_prefix_matching(self):
        self.assertEqual(self.names(self.list_(status='MATCHED')), {'a_fy26', 'b_fy26_auto', 'l_orphan_jn'},
                         'status=MATCHED must include "MATCHED (auto-recon ...)" rows')
        self.assertEqual(self.names(self.list_(status='matched')), {'a_fy26', 'b_fy26_auto', 'l_orphan_jn'})
        self.assertEqual(self.names(self.list_(status='SKIPPED')), {'e_skip_neg'})
        self.assertEqual(self.names(self.list_(status='skipped')), {'e_skip_neg'})
        self.assertEqual(self.names(self.list_(status='NOT IN XERO')), {'d_fy26_nix'})
        self.assertEqual(self.list_(status='PENDING')['count'], 7)
        self.assertEqual(self.list_(status='BANANA')['count'], 0)
        self.assertEqual(self.list_(status="'; drop table x; --")['count'], 0)

    def test_status_group_labels(self):
        groups = {r['sha256']: r['status_group'] for r in self.list_(page_size=200)['results']}
        self.assertEqual(groups[self.NAMES['b_fy26_auto']], 'MATCHED')
        self.assertEqual(groups[self.NAMES['e_skip_neg']], 'SKIPPED')
        self.assertEqual(groups[self.NAMES['d_fy26_nix']], 'NOT IN XERO')
        self.assertEqual(groups[self.NAMES['c_fy25_pend']], 'PENDING')

    def test_category_filter_is_exact_case_insensitive(self):
        self.assertEqual(self.names(self.list_(category='hardware')), {'a_fy26', 'd_fy26_nix', 'e_skip_neg'})
        self.assertEqual(self.names(self.list_(category='Fuel')), {'c_fy25_pend', 'k_fy27'})
        self.assertEqual(self.list_(category='Hard')['count'], 0, 'category is an exact match, not a prefix')
        self.assertEqual(self.list_(category="x'; drop table x; --")['count'], 0)

    def test_min_max_total(self):
        self.assertEqual(self.names(self.list_(min_total='100')), {'a_fy26', 'b_fy26_auto', 'k_fy27'})
        self.assertEqual(self.names(self.list_(min_total='259.01')), {'b_fy26_auto', 'k_fy27'})
        self.assertEqual(self.names(self.list_(min_total='259.00')), {'a_fy26', 'b_fy26_auto', 'k_fy27'}, 'bounds are inclusive')
        self.assertEqual(self.names(self.list_(max_total='40')), {'c_fy25_pend', 'j_ts_null'})
        self.assertEqual(self.names(self.list_(min_total='50', max_total='300')), {'a_fy26', 'd_fy26_nix', 'l_orphan_jn'})
        # Rows whose total is non-numeric / negative / missing never match a numeric bound.
        self.assertEqual(self.list_(min_total='-1000')['count'], 7)
        self.assertEqual(self.list_(max_total='0')['count'], 0)

    def test_min_total_junk_is_ignored_or_harmless(self):
        self.assertEqual(self.list_(min_total='abc')['count'], 12)
        self.assertEqual(self.list_(min_total='')['count'], 12)
        self.assertEqual(self.list_(min_total="1; drop table x")['count'], 12)
        # Decimal() accepts these; Postgres must not choke on them.
        for weird in ('NaN', 'sNaN', 'Infinity', '-Infinity', '1e400', '1e-400', '0x10'):
            data = self.list_(min_total=weird)
            self.assertIn('count', data)
            data = self.list_(max_total=weird)
            self.assertIn('count', data)

    def test_date_range_filters(self):
        self.assertEqual(self.names(self.list_(date_from='2026-04-01', date_to='2026-04-04')),
                         {'f_ocr_abc', 'g_ocr_empty', 'h_ocr_nototal', 'i_ocr_null'})
        self.assertEqual(self.names(self.list_(date_from='2026-07-01')), {'k_fy27'})
        self.assertEqual(self.names(self.list_(date_to='2025-03-01')), {'c_fy25_pend'})
        # NULL slip_ts is excluded by any date bound
        self.assertNotIn('j_ts_null', self.names(self.list_(date_from='1900-01-01')))
        # junk dates ignored
        self.assertEqual(self.list_(date_from='not-a-date')['count'], 12)
        self.assertEqual(self.list_(date_from='2026-13-45')['count'], 12)

    def test_ordering_whitelist_and_sqli(self):
        default = self.shas(self.list_(page_size=200))
        for bad in ('slip_ts; DROP TABLE x', '__proto__', 'file_bytes', '-file_bytes', 's.sha256', 'total desc', '', ' '):
            self.assertEqual(self.shas(self.list_(ordering=bad, page_size=200)), default,
                             f'ordering={bad!r} must fall back to the default ordering')
        self.assertEqual(self.list_()['count'], 12)

    def test_ordering_slip_ts_nulls_last_both_directions(self):
        asc = self.shas(self.list_(ordering='slip_ts', page_size=200))
        desc = self.shas(self.list_(ordering='-slip_ts', page_size=200))
        self.assertEqual(asc[-1], self.NAMES['j_ts_null'])
        self.assertEqual(desc[-1], self.NAMES['j_ts_null'])
        self.assertEqual(asc[0], self.NAMES['c_fy25_pend'])
        self.assertEqual(desc[0], self.NAMES['k_fy27'])
        self.assertEqual(asc[:-1], list(reversed(desc[:-1])))

    def test_ordering_total_with_non_numeric_values(self):
        rows = self.list_(ordering='-total', page_size=200)['results']
        totals = [r['total'] for r in rows]
        numeric = [Decimal(t) for t in totals if t is not None]
        self.assertEqual(numeric, sorted(numeric, reverse=True))
        self.assertEqual(totals[:7], ['1234.56', '500.00', '259.00', '99.99', '77.00', '40.00', '15.00'])
        self.assertTrue(all(t is None for t in totals[7:]), 'non-numeric totals must sort last')
        rows = self.list_(ordering='total', page_size=200)['results']
        self.assertEqual([r['total'] for r in rows][:7], ['15.00', '40.00', '77.00', '99.99', '259.00', '500.00', '1234.56'])

    def test_ordering_supplier_and_status(self):
        rows = self.list_(ordering='supplier', page_size=200)['results']
        sups = [r['supplier'] for r in rows if r['supplier']]
        self.assertEqual(sups, sorted(sups, key=str.lower))
        self.assertIsNone(rows[-1]['supplier'], 'NULL supplier should sort last')
        rows = self.list_(ordering='xero_status', page_size=200)['results']
        sts = [r['xero_status'] for r in rows]
        self.assertEqual(sts, sorted(sts))

    def test_combined_filters_and_totals_agree(self):
        data = self.list_(fy='FY26', status='MATCHED', synced='true', category='Hardware', min_total='100', ordering='-total')
        self.assertEqual(self.names(data), {'a_fy26'})
        self.assertEqual(data['totals'], {'count': 1, 'sum_total': '259.00'})
        data = self.list_(fy='FY26', status='PENDING')
        self.assertEqual(self.names(data), {'f_ocr_abc', 'g_ocr_empty', 'h_ocr_nototal', 'i_ocr_null'})
        self.assertEqual(data['totals'], {'count': 4, 'sum_total': '0.00'})

    def test_to_process_and_decision_filters(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True, decision='CAPTURE')
        SlipReview.objects.create(sha256=self.NAMES['d_fy26_nix'], to_process=True, decision='')
        SlipReview.objects.create(sha256=self.NAMES['f_ocr_abc'], to_process=False, decision='PERSONAL')
        self.assertEqual(self.names(self.list_(to_process='true')), {'a_fy26', 'd_fy26_nix'})
        self.assertEqual(self.list_(to_process='false')['count'], 10)
        self.assertEqual(self.list_(to_process='junk')['count'], 12)
        self.assertEqual(self.names(self.list_(decision='CAPTURE')), {'a_fy26'})
        self.assertEqual(self.names(self.list_(decision='personal')), {'f_ocr_abc'})
        undecided = self.names(self.list_(decision='UNDECIDED'))
        self.assertEqual(len(undecided), 10)
        self.assertIn('d_fy26_nix', undecided, 'a review row with decision="" is still undecided')
        self.assertNotIn('a_fy26', undecided)
        self.assertEqual(self.names(self.list_(decision='none')), undecided)
        # combination
        self.assertEqual(self.names(self.list_(to_process='true', decision='UNDECIDED')), {'d_fy26_nix'})
        self.assertEqual(self.names(self.list_(to_process='true', decision='CAPTURE', fy='FY26')), {'a_fy26'})
        # totals follow the filter
        self.assertEqual(self.list_(to_process='true')['totals'], {'count': 2, 'sum_total': '358.99'})

    def test_decision_junk_does_not_500(self):
        for junk in ('BANANA', "'; drop table x; --", '__proto__', ' '):
            self.assertIn('count', self.list_(decision=junk))

    def test_to_process_false_with_no_reviews_at_all(self):
        # empty array param -> `not (sha = any('{}'))` must still return everything.
        self.assertEqual(self.list_(to_process='false')['count'], 12)
        self.assertEqual(self.list_(to_process='true')['count'], 0)
        self.assertEqual(self.list_(decision='UNDECIDED')['count'], 12)
        self.assertEqual(self.list_(decision='CAPTURE')['count'], 0)

    def test_review_state_and_comment_count_attached(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True, decision='CAPTURE', note='n', updated_by='mc')
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='one', author='mc')
        SlipComment.objects.create(sha256=self.NAMES['a_fy26'], text='two', author='mc')
        rows = {r['sha256']: r for r in self.list_(page_size=200)['results']}
        a = rows[self.NAMES['a_fy26']]
        self.assertEqual(a['comment_count'], 2)
        self.assertEqual(a['review']['decision'], 'CAPTURE')
        self.assertTrue(a['review']['to_process'])
        self.assertEqual(a['review']['updated_by'], 'mc')
        b = rows[self.NAMES['b_fy26_auto']]
        self.assertEqual(b['comment_count'], 0)
        self.assertEqual(b['review'], {'to_process': False, 'decision': '', 'note': '', 'updated_by': '', 'updated_at': None})


# --------------------------------------------------------------------------- #
# 2. FY bucketing
# --------------------------------------------------------------------------- #
class FiscalYearTests(TestCase):
    """Pure helpers + SQL boundary behaviour. FY = Jul..Jun named by the ENDING year, read in SLIP_TZ (SAST)."""

    @classmethod
    def setUpTestData(cls):
        create_slips_table()
        SAST = SLIP_TZ
        cls.rows = {
            # --- SAST-expressed boundaries ---
            'fy25_last_minute_sast': (sha(101), ts(2025, 6, 30, 23, 59, tz=SAST), 'FY25'),
            'fy26_first_minute_sast': (sha(102), ts(2025, 7, 1, 0, 0, tz=SAST), 'FY26'),
            'fy26_last_minute_sast': (sha(103), ts(2026, 6, 30, 23, 59, tz=SAST), 'FY26'),
            'fy27_first_minute_sast': (sha(104), ts(2026, 7, 1, 0, 0, tz=SAST), 'FY27'),
            # --- The UTC/SAST straddle: 2025-06-30T22:30Z == 2025-07-01 00:30 SAST -> FY26 under SLIP_TZ ---
            'straddle_utc_jun30_is_sast_jul1': (sha(105), ts(2025, 6, 30, 22, 30), 'FY26'),
            # --- The reverse straddle: 2026-06-30T21:59:59Z == 2026-06-30 23:59:59 SAST -> still FY26 ---
            'straddle_utc_jun30_2159_is_sast_jun30': (sha(106), ts(2026, 6, 30, 21, 59, 59), 'FY26'),
            # --- 2026-06-30T22:00:00Z == 2026-07-01 00:00 SAST -> FY27 ---
            'straddle_utc_jun30_2200_is_sast_jul1': (sha(107), ts(2026, 6, 30, 22, 0, 0), 'FY27'),
            # --- NULL timestamp (exists in production) ---
            'null_ts': (sha(108), None, None),
        }
        for name, (s, when, _fy) in cls.rows.items():
            insert_slip(s, slip_ts=when, filename=f'{name}.jpg', ocr={'supplier': name, 'total': '1.00'})
        cls.user = User.objects.create_user(username='fy-reviewer', email='fy@example.com', password='pw-irrelevant')

    def setUp(self):
        # reads are now IsAuthenticated
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {RefreshToken.for_user(self.user).access_token}')

    def list_(self, **params):
        resp = self.client.get(reverse('receipts:list'), params)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()

    # --- pure helpers ---
    def test_fy_range(self):
        self.assertEqual(fy_range('FY26'), (dt.date(2025, 7, 1), dt.date(2026, 6, 30)))
        self.assertEqual(fy_range('fy26'), (dt.date(2025, 7, 1), dt.date(2026, 6, 30)))
        self.assertEqual(fy_range('FY2026'), (dt.date(2025, 7, 1), dt.date(2026, 6, 30)))
        self.assertEqual(fy_range(' FY25 '), (dt.date(2024, 7, 1), dt.date(2025, 6, 30)))
        for bad in ('2026', 'FY', 'FY2', 'FY12345', 'FY26; drop', '', None):
            with self.assertRaises(ValueError, msg=bad):
                fy_range(bad)

    def test_fy_label_pure(self):
        self.assertEqual(fy_label(dt.date(2025, 6, 30)), 'FY25')
        self.assertEqual(fy_label(dt.date(2025, 7, 1)), 'FY26')
        self.assertEqual(fy_label(dt.date(2026, 6, 30)), 'FY26')
        self.assertEqual(fy_label(dt.date(2026, 7, 1)), 'FY27')
        self.assertIsNone(fy_label(None))
        # aware datetimes are read in SAST
        self.assertEqual(fy_label(ts(2025, 6, 30, 22, 30)), 'FY26', '22:30Z on 30 Jun is 00:30 SAST on 1 Jul')
        self.assertEqual(fy_label(ts(2025, 6, 30, 21, 59)), 'FY25')
        self.assertEqual(fy_label(ts(2026, 6, 30, 22, 0)), 'FY27')

    # --- SQL + shaped row agree on every boundary ---
    def test_each_row_lands_in_the_documented_fy_in_list_and_filter(self):
        data = self.list_(page_size=200)
        self.assertEqual(data['count'], 8)
        by_sha = {r['sha256']: r for r in data['results']}
        for name, (s, _when, expected) in self.rows.items():
            self.assertEqual(by_sha[s]['fy'], expected, f'{name}: row.fy')
        for fy in ('FY25', 'FY26', 'FY27'):
            got = {r['sha256'] for r in self.list_(fy=fy, page_size=200)['results']}
            want = {s for (s, _w, e) in self.rows.values() if e == fy}
            self.assertEqual(got, want, f'fy={fy} filter set')

    def test_fy26_boundaries_explicit(self):
        got = {r['sha256'] for r in self.list_(fy='FY26')['results']}
        self.assertIn(sha(102), got)   # 2025-07-01 00:00 SAST
        self.assertIn(sha(103), got)   # 2026-06-30 23:59 SAST
        self.assertNotIn(sha(101), got)  # 2025-06-30 23:59 SAST
        self.assertNotIn(sha(104), got)  # 2026-07-01 00:00 SAST

    def test_utc_jun30_2230Z_lands_in_FY26_because_it_is_jul1_sast(self):
        got = {r['sha256'] for r in self.list_(fy='FY26')['results']}
        self.assertIn(sha(105), got)
        self.assertNotIn(sha(105), {r['sha256'] for r in self.list_(fy='FY25')['results']})

    def test_utc_jun30_2200Z_lands_in_FY27_because_it_is_jul1_sast(self):
        self.assertIn(sha(107), {r['sha256'] for r in self.list_(fy='FY27')['results']})
        self.assertNotIn(sha(107), {r['sha256'] for r in self.list_(fy='FY26')['results']})
        self.assertIn(sha(106), {r['sha256'] for r in self.list_(fy='FY26')['results']})

    def test_fy_and_date_filters_use_the_same_sast_day(self):
        # date_from/date_to must agree with the fy bucketing (both SAST) or the console's
        # "FY26" pill and a hand-typed 2025-07-01..2026-06-30 range would disagree.
        fy = {r['sha256'] for r in self.list_(fy='FY26', page_size=200)['results']}
        rng = {r['sha256'] for r in self.list_(date_from='2025-07-01', date_to='2026-06-30', page_size=200)['results']}
        self.assertEqual(fy, rng)

    def test_null_slip_ts_does_not_crash_and_is_excluded_from_fy(self):
        data = self.list_(page_size=200)
        null_row = next(r for r in data['results'] if r['sha256'] == sha(108))
        self.assertIsNone(null_row['slip_ts'])
        self.assertIsNone(null_row['fy'])
        for fy in ('FY25', 'FY26', 'FY27'):
            self.assertNotIn(sha(108), {r['sha256'] for r in self.list_(fy=fy)['results']})
        detail = self.client.get(reverse('receipts:detail', args=[sha(108)]))
        self.assertEqual(detail.status_code, 200)
        self.assertIsNone(detail.json()['fy'])

    def test_bad_fy_label_is_ignored_not_500(self):
        for bad in ('2026', 'FY', 'FY2026-27', "FY26'; drop table x; --", 'fy 26'):
            self.assertEqual(self.list_(fy=bad)['count'], 8, f'fy={bad!r}')


# --------------------------------------------------------------------------- #
# 3. Journal join: duplication + cross-tenant collision
# --------------------------------------------------------------------------- #
class JournalJoinTests(ReceiptsFixtureMixin, TestCase):
    def test_five_line_journal_yields_exactly_one_row(self):
        data = self.list_(page_size=200)
        hits = [r for r in data['results'] if r['sha256'] == self.NAMES['a_fy26']]
        self.assertEqual(len(hits), 1, 'journal join multiplied the slip row')
        self.assertEqual(data['count'], 12)
        # and the export agrees
        rows = self.csv_rows(self.export())
        self.assertEqual(sum(1 for r in rows[1:] if r[-2] == self.NAMES['a_fy26']), 1)

    def test_journal_summary_is_the_expense_side_not_the_bank_side(self):
        row = self.client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).json()
        j = row['journal']
        self.assertIsNotNone(j)
        self.assertEqual(j['journal_number'], 697)
        self.assertEqual(j['account_code'], '429')
        self.assertEqual(j['account_name'], 'Hardware & Consumables')
        self.assertEqual(j['description'], 'Drill bits')
        self.assertEqual(j['contact_name'], 'Builders Warehouse')
        self.assertEqual(Decimal(j['debit']), Decimal('259.00'))
        self.assertEqual(Decimal(j['credit']), Decimal('-259.00'))
        self.assertEqual(Decimal(j['amount']), Decimal('259.00'))
        self.assertTrue(j['date'].startswith('2025-08-10'))

    def test_cross_tenant_journal_697_resolves_to_the_slips_org_only(self):
        # Tenant B also has #697 (812.40 Woolworths groceries). Slip is Klikk -> must see Builders 259.
        row = self.client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).json()
        j = row['journal']
        self.assertEqual(j['contact_name'], 'Builders Warehouse')
        self.assertNotEqual(j['contact_name'], 'Woolworths')
        self.assertEqual(Decimal(j['amount']), Decimal('259.00'), 'amount must not include tenant B lines')
        self.assertNotEqual(Decimal(j['amount']), Decimal('1071.40'), 'tenant A + tenant B summed together')
        self.assertEqual(j['description'], 'Drill bits')

    def test_slip_pointing_at_other_tenant_sees_other_tenants_lines(self):
        insert_slip(sha(201), slip_ts=ts(2024, 2, 2, 10, 0), xero_status='MATCHED', synced=True,
                    journal_number=697, xero_org='Dippenaar Family', ocr={'supplier': 'Woolworths', 'total': '812.40'})
        j = self.client.get(reverse('receipts:detail', args=[sha(201)])).json()['journal']
        self.assertEqual(j['contact_name'], 'Woolworths')
        self.assertEqual(j['account_code'], '450')
        self.assertEqual(Decimal(j['amount']), Decimal('812.40'))

    def test_unknown_named_org_gives_no_journal_not_a_guess(self):
        # NULL/blank xero_org now falls back to the Klikk tenant (see JournalTypeAndOrgScopeTests);
        # an unknown *named* org must still resolve to nothing — never to another tenant's lines.
        insert_slip(sha(202), slip_ts=ts(2025, 1, 1), xero_status='MATCHED', journal_number=697, xero_org='Nonexistent Org',
                    ocr={'total': '1.00'})
        row = self.client.get(reverse('receipts:detail', args=[sha(202)])).json()
        self.assertIsNone(row['journal'], 'journal must not be resolved from a different tenant')
        self.assertEqual(row['journal_number'], 697)

    def test_orphan_journal_number_and_no_journal_number(self):
        rows = {r['sha256']: r for r in self.list_(page_size=200)['results']}
        self.assertIsNone(rows[self.NAMES['l_orphan_jn']]['journal'])
        self.assertEqual(rows[self.NAMES['l_orphan_jn']]['journal_number'], 555)
        self.assertIsNone(rows[self.NAMES['c_fy25_pend']]['journal'])
        self.assertIsNone(rows[self.NAMES['c_fy25_pend']]['journal_number'])

    def test_same_org_journal_number_shared_by_journal_and_manual_journal_is_not_blended(self):
        # The join is discriminated by journal_type = 'journal' (services.JOURNAL_TYPE): when a slip's
        # number exists as BOTH a `journal` and a `manual_journal` in the same org, the endpoint must
        # return exactly the 'journal' one — never the manual one, never a blend of the two.
        k_loan = make_account(self.klikk, 'acc-k-loan', '880', 'Loan - MC', 'CURRLIAB')
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_meals = XeroAccount.objects.get(account_id='acc-k-meals')
        d = ts(2026, 1, 8)
        # regular journal #777: Spur 300 (2 lines)
        make_line(self.klikk, 777, k_meals, '300.00', journal_id='k777-j1', date=d, description='Spur lunch')
        make_line(self.klikk, 777, k_bank, '-300.00', journal_id='k777-j2', date=d, description='Spur lunch')
        # manual journal #777 (same org, same number): loan 1500 (3 lines)
        make_line(self.klikk, 777, k_loan, '1000.00', journal_id='k777-m0', date=d,
                  description='Loan repayment A', journal_type='manual_journal')
        make_line(self.klikk, 777, k_loan, '500.00', journal_id='k777-m1', date=d,
                  description='Loan repayment B', journal_type='manual_journal')
        make_line(self.klikk, 777, k_bank, '-1500.00', journal_id='k777-m2', date=d,
                  description='Loan repayment', journal_type='manual_journal')
        insert_slip(sha(210), slip_ts=d, xero_status='MATCHED (auto-recon 2026-08-17)', xero_detail='journal_number=777',
                    synced=True, journal_number=777, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Spur', 'total': '300.00', 'category': 'Meals'})
        j = self.client.get(reverse('receipts:detail', args=[sha(210)])).json()['journal']
        self.assertIsNotNone(j)
        self.assertEqual(j['journal_number'], 777)
        self.assertEqual(Decimal(j['amount']), Decimal('300.00'),
                         f"journal amount {j['amount']} must be exactly the 'journal' #777, not the manual/blend")
        self.assertNotEqual(Decimal(j['amount']), Decimal('1800.00'),
                            'amount is the blended sum of the journal and the manual_journal')
        self.assertNotEqual(Decimal(j['amount']), Decimal('1500.00'), "amount is the manual_journal's, not the journal's")
        self.assertEqual(j['description'], 'Spur lunch')
        self.assertEqual(j['account_code'], '420')
        self.assertEqual(Decimal(j['debit']), Decimal('300.00'))
        self.assertEqual(Decimal(j['credit']), Decimal('-300.00'))

    def test_two_line_journal(self):
        j = next(r for r in self.list_(page_size=200)['results'] if r['sha256'] == self.NAMES['b_fy26_auto'])['journal']
        self.assertEqual(j['journal_number'], 962018)
        self.assertEqual(j['account_code'], '420')
        self.assertEqual(j['contact_name'], 'Jonnys Mozambican Restaurant')
        self.assertEqual(Decimal(j['amount']), Decimal('1234.56'))


# --------------------------------------------------------------------------- #
# 4. Detail
# --------------------------------------------------------------------------- #
class DetailTests(ReceiptsFixtureMixin, TestCase):
    def test_detail_returns_ocr_items_and_comments_in_created_order(self):
        s = self.NAMES['a_fy26']
        c1 = SlipComment.objects.create(sha256=s, text='first', author='mc')
        c2 = SlipComment.objects.create(sha256=s, text='second', author='mc')
        # force c1 to be newer than c2 in created_at terms? No — assert by created_at ascending.
        resp = self.client.get(reverse('receipts:detail', args=[s]))
        self.assertEqual(resp.status_code, 200)
        row = resp.json()
        self.assertEqual(row['ocr']['supplier'], 'Builders Warehouse')
        self.assertEqual(row['items'], [{'description': 'Drill bits', 'amount': '200.00'}, {'description': 'Screws', 'amount': '25.22'}])
        self.assertEqual([c['id'] for c in row['comments']], [c1.id, c2.id])
        self.assertEqual([c['text'] for c in row['comments']], ['first', 'second'])
        self.assertEqual(row['comment_count'], 2)
        self.assertNotIn('file_bytes', row)
        self.assertNotIn('file_bytes', row['ocr'])

    def test_detail_with_null_ocr_and_missing_items(self):
        row = self.client.get(reverse('receipts:detail', args=[self.NAMES['i_ocr_null']])).json()
        self.assertEqual(row['ocr'], {})
        self.assertEqual(row['items'], [])
        row = self.client.get(reverse('receipts:detail', args=[self.NAMES['h_ocr_nototal']])).json()
        self.assertEqual(row['items'], [])
        self.assertIsNone(row['total'])

    def test_detail_with_non_object_ocr_jsonb(self):
        # jsonb can legally hold a string / array / number; the API must not 500 on them.
        insert_slip(sha(301), slip_ts=ts(2025, 5, 5), ocr_raw='"just a string"')
        insert_slip(sha(302), slip_ts=ts(2025, 5, 5), ocr_raw='[1, 2, 3]')
        insert_slip(sha(303), slip_ts=ts(2025, 5, 5), ocr_raw='{"items": "not-a-list", "total": 12}')
        insert_slip(sha(304), slip_ts=ts(2025, 5, 5), ocr_raw='{"items": [1, "x", null, {"description": "ok", "amount": "1.00"}], "total": "12"}')
        for s in (sha(301), sha(302), sha(303), sha(304)):
            resp = self.client.get(reverse('receipts:detail', args=[s]))
            self.assertEqual(resp.status_code, 200, f'{s}: {resp.content[:300]}')
        self.assertEqual(self.client.get(reverse('receipts:detail', args=[sha(304)])).json()['items'],
                         [{'description': 'ok', 'amount': '1.00'}])
        # numeric json total (12, not "12") — ->> renders it as '12' so it is numeric
        self.assertEqual(self.client.get(reverse('receipts:detail', args=[sha(303)])).json()['total'], '12.00')
        # the list must also survive these rows
        self.assertEqual(self.list_(page_size=200)['count'], 16)

    def test_detail_404_and_odd_keys(self):
        self.assertEqual(self.client.get(reverse('receipts:detail', args=['0' * 64])).status_code, 404)
        self.assertEqual(self.client.get(reverse('receipts:detail', args=['nope'])).status_code, 404)
        self.assertEqual(self.client.get(reverse('receipts:detail', args=["x'--"])).status_code, 404)
        self.assertEqual(self.client.get(reverse('receipts:detail', args=['a' * 300])).status_code, 404)

    def test_view_url_in_detail_matches_list(self):
        s = self.NAMES['a_fy26']
        d = self.client.get(reverse('receipts:detail', args=[s])).json()['view_url']
        l = next(r for r in self.list_(page_size=200)['results'] if r['sha256'] == s)['view_url']
        self.assertEqual(d, l)


# --------------------------------------------------------------------------- #
# 5. Review PATCH
# --------------------------------------------------------------------------- #
class ReviewPatchTests(ReceiptsFixtureMixin, TestCase):
    def url(self, s=None):
        return reverse('receipts:review', args=[s or self.NAMES['a_fy26']])

    def test_unauthenticated_patch_is_401(self):
        self.client.credentials()  # setUp authenticates by default; this test probes the anonymous path
        resp = self.client.patch(self.url(), {'decision': 'CAPTURE'}, format='json')
        self.assertEqual(resp.status_code, 401, resp.content)
        self.assertFalse(SlipReview.objects.exists())
        # Even for an unknown slip the gate must come before the lookup.
        self.assertEqual(self.client.patch(self.url('0' * 64), {'decision': 'CAPTURE'}, format='json').status_code, 401)

    def test_garbage_bearer_token_is_401(self):
        self.client.credentials(HTTP_AUTHORIZATION='Bearer not.a.jwt')
        self.assertEqual(self.client.patch(self.url(), {'decision': 'CAPTURE'}, format='json').status_code, 401)
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + 'a' * 40)
        self.assertEqual(self.client.patch(self.url(), {'decision': 'CAPTURE'}, format='json').status_code, 401)
        self.assertFalse(SlipReview.objects.exists())

    def test_authenticated_patch_upserts_and_sets_updated_by(self):
        self.auth()
        resp = self.client.patch(self.url(), {'to_process': True, 'decision': 'capture', 'note': 'buy it'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        self.assertEqual(body['decision'], 'CAPTURE')
        self.assertTrue(body['to_process'])
        self.assertEqual(body['note'], 'buy it')
        self.assertEqual(body['updated_by'], 'reviewer')
        self.assertIsNotNone(body['updated_at'])
        rv = SlipReview.objects.get(sha256=self.NAMES['a_fy26'])
        self.assertEqual(rv.updated_by, 'reviewer')
        # second PATCH is an update, not a duplicate; untouched fields persist
        resp = self.client.patch(self.url(), {'note': 'changed'}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(SlipReview.objects.count(), 1)
        rv.refresh_from_db()
        self.assertEqual(rv.note, 'changed')
        self.assertEqual(rv.decision, 'CAPTURE')
        self.assertTrue(rv.to_process)

    def test_every_valid_decision_including_blank(self):
        self.auth()
        for d in ('', 'CAPTURE', 'MEAL_SKIP', 'PERSONAL', 'DUPLICATE', 'ALREADY_IN_XERO'):
            resp = self.client.patch(self.url(), {'decision': d}, format='json')
            self.assertEqual(resp.status_code, 200, (d, resp.content))
            self.assertEqual(resp.json()['decision'], d)
        # null decision == undecided
        resp = self.client.patch(self.url(), {'decision': None}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['decision'], '')

    def test_invalid_decision_is_400_and_writes_nothing(self):
        self.auth()
        for d in ('BANANA', 'CAPTURE;DROP', 'SKIP', 'MEAL SKIP', 'MEAL-SKIP', 1, ['CAPTURE'], {'x': 1}):
            resp = self.client.patch(self.url(), {'decision': d}, format='json')
            self.assertEqual(resp.status_code, 400, (d, resp.content))
        self.assertFalse(SlipReview.objects.exists())
        # case/whitespace are normalised, not rejected (views.py strip().upper())
        resp = self.client.patch(self.url(), {'decision': ' capture '}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['decision'], 'CAPTURE')

    def test_invalid_decision_with_valid_sibling_field_still_400_atomically(self):
        self.auth()
        resp = self.client.patch(self.url(), {'to_process': True, 'decision': 'NOPE'}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(SlipReview.objects.exists(), 'partial write on a 400')

    def test_unknown_sha_is_404(self):
        self.auth()
        self.assertEqual(self.client.patch(self.url('0' * 64), {'decision': 'CAPTURE'}, format='json').status_code, 404)
        self.assertFalse(SlipReview.objects.exists())

    def test_empty_body_is_400(self):
        self.auth()
        self.assertEqual(self.client.patch(self.url(), {}, format='json').status_code, 400)
        self.assertEqual(self.client.patch(self.url()).status_code, 400)
        self.assertEqual(self.client.patch(self.url(), {'unknown_field': 1}, format='json').status_code, 400)
        self.assertFalse(SlipReview.objects.exists())

    def test_to_process_coercions(self):
        self.auth()
        for raw, want in ((True, True), ('true', True), ('1', True), (1, True), ('yes', True), ('on', True),
                          (False, False), ('false', False), ('0', False), (0, False), (None, False)):
            resp = self.client.patch(self.url(), {'to_process': raw}, format='json')
            self.assertEqual(resp.status_code, 200, (raw, resp.content))
            self.assertEqual(resp.json()['to_process'], want, f'to_process={raw!r}')

    def test_to_process_uppercase_TRUE_is_coerced_true(self):
        # BUG-3: views.py receipt_review_view — `data['to_process'] in (True, 1, '1', 'true', 'True', 'yes', 'on')`
        # means 'TRUE' / 'Yes' / 'ON' silently become False: the client is told 200 but the flag is off.
        # Compare with services._bool() which lowercases first; the two coercions disagree.
        self.auth()
        resp = self.client.patch(self.url(), {'to_process': 'TRUE'}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()['to_process'], "to_process='TRUE' silently coerced to False")

    def test_non_object_json_body_does_not_500(self):
        # BUG-1: views.py receipt_review_view does `'note' in data` / `data['note']` on whatever JSON
        # was posted. A JSON *string* body containing the substring 'note' (or 'decision') passes the
        # `in` test and then `data['note']` -> TypeError -> 500; a JSON number body -> TypeError on `in`
        # -> 500; a JSON list ['note'] -> `data['note']` TypeError -> 500. Authenticated caller, but
        # still a 5xx from client input. Fix: `if not isinstance(request.data, dict): return 400`.
        self.auth()
        client = APIClient(raise_request_exception=False)
        client.credentials(**self.client._credentials)
        for raw in ('"a note"', '["note"]', '42', 'null', '"decision"'):
            resp = client.generic('PATCH', self.url(), data=raw, content_type='application/json')
            self.assertIn(resp.status_code, (200, 400), f'body={raw!r} -> {resp.status_code}')
        self.assertFalse(SlipReview.objects.exists())

    def test_note_is_stringified_not_rejected(self):
        self.auth()
        resp = self.client.patch(self.url(), {'note': 123}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['note'], '123')
        resp = self.client.patch(self.url(), {'note': None}, format='json')
        self.assertEqual(resp.json()['note'], '')

    def test_review_visible_in_list_and_export_after_patch(self):
        self.auth()
        self.client.patch(self.url(), {'to_process': True, 'decision': 'DUPLICATE', 'note': 'dup of x'}, format='json')
        row = next(r for r in self.list_(page_size=200)['results'] if r['sha256'] == self.NAMES['a_fy26'])
        self.assertEqual(row['review']['decision'], 'DUPLICATE')
        self.assertTrue(row['review']['to_process'])
        rows = self.csv_rows(self.export(decision='DUPLICATE'))
        self.assertEqual(len(rows), 2)
        hdr = rows[0]
        rec = dict(zip(hdr, rows[1]))
        self.assertEqual(rec['decision'], 'DUPLICATE')
        self.assertEqual(rec['to_process'], 'True')
        self.assertEqual(rec['note'], 'dup of x')

    def test_form_encoded_patch_also_works(self):
        self.auth()
        resp = self.client.patch(self.url(), data='decision=PERSONAL&to_process=true',
                                 content_type='application/x-www-form-urlencoded')
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()['decision'], 'PERSONAL')
        self.assertTrue(resp.json()['to_process'])


# --------------------------------------------------------------------------- #
# 6. Comment POST
# --------------------------------------------------------------------------- #
class CommentPostTests(ReceiptsFixtureMixin, TestCase):
    def url(self, s=None):
        return reverse('receipts:comments', args=[s or self.NAMES['a_fy26']])

    def test_unauthenticated_post_is_401(self):
        self.client.credentials()  # setUp authenticates by default; this test probes the anonymous path
        resp = self.client.post(self.url(), {'text': 'hi'}, format='json')
        self.assertEqual(resp.status_code, 401, resp.content)
        self.assertFalse(SlipComment.objects.exists())
        self.assertEqual(self.client.post(self.url('0' * 64), {'text': 'hi'}, format='json').status_code, 401)

    def test_authenticated_post_201_sets_author(self):
        self.auth()
        resp = self.client.post(self.url(), {'text': '  needs VAT invoice  '}, format='json')
        self.assertEqual(resp.status_code, 201, resp.content)
        body = resp.json()
        self.assertEqual(body['text'], 'needs VAT invoice')
        self.assertEqual(body['author'], 'reviewer')
        self.assertIsNotNone(body['created_at'])
        self.assertIn('id', body)
        c = SlipComment.objects.get()
        self.assertEqual(c.author, 'reviewer')
        self.assertEqual(c.sha256, self.NAMES['a_fy26'])

    def test_blank_or_whitespace_text_is_400(self):
        self.auth()
        for t in ('', '   ', '\n\t ', None):
            resp = self.client.post(self.url(), {'text': t}, format='json')
            self.assertEqual(resp.status_code, 400, (t, resp.content))
        self.assertEqual(self.client.post(self.url(), {}, format='json').status_code, 400)
        self.assertEqual(self.client.post(self.url()).status_code, 400)
        self.assertFalse(SlipComment.objects.exists())

    def test_unknown_sha_is_404(self):
        self.auth()
        self.assertEqual(self.client.post(self.url('0' * 64), {'text': 'x'}, format='json').status_code, 404)
        self.assertFalse(SlipComment.objects.exists())

    def test_comment_count_and_detail_order(self):
        self.auth()
        for t in ('one', 'two', 'three'):
            self.assertEqual(self.client.post(self.url(), {'text': t}, format='json').status_code, 201)
        row = next(r for r in self.list_(page_size=200)['results'] if r['sha256'] == self.NAMES['a_fy26'])
        self.assertEqual(row['comment_count'], 3)
        others = [r['comment_count'] for r in self.list_(page_size=200)['results'] if r['sha256'] != self.NAMES['a_fy26']]
        self.assertEqual(set(others), {0})
        detail = self.client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).json()
        self.assertEqual([c['text'] for c in detail['comments']], ['one', 'two', 'three'])
        created = [c['created_at'] for c in detail['comments']]
        self.assertEqual(created, sorted(created))

    def test_non_object_json_body_does_not_500(self):
        # BUG-2: views.py receipt_comments_view `(request.data or {}).get('text')` — a non-empty JSON
        # list / number / string body is truthy and has no .get -> AttributeError -> 500.
        # Fix: `if not isinstance(request.data, dict): return 400`.
        self.auth()
        client = APIClient(raise_request_exception=False)
        client.credentials(**self.client._credentials)
        for raw in ('["text"]', '"text"', '42', '[]'):
            resp = client.generic('POST', self.url(), data=raw, content_type='application/json')
            self.assertIn(resp.status_code, (400, 415), f'body={raw!r} -> {resp.status_code}')
        self.assertFalse(SlipComment.objects.exists())

    def test_long_and_unicode_text_is_stored_verbatim(self):
        self.auth()
        text = 'Kwitansie — R1 234,56 — “aanhalings” 🧾 ' + 'x' * 5000
        resp = self.client.post(self.url(), {'text': text}, format='json')
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(SlipComment.objects.get().text, text.strip())

    def test_get_on_comments_url_is_405_not_a_listing(self):
        # unauthenticated GET is gated first (401); authenticated GET must be 405, never a comment dump
        self.assertEqual(self.anon.get(self.url()).status_code, 401)
        self.assertEqual(self.client.get(self.url()).status_code, 405)


# --------------------------------------------------------------------------- #
# 7. Export
# --------------------------------------------------------------------------- #
class ExportTests(ReceiptsFixtureMixin, TestCase):
    EXPECTED_HEADER = ['date', 'supplier', 'total', 'category', 'xero_status', 'status_group', 'journal_number',
                       'synced', 'to_process', 'decision', 'note', 'filename', 'sha256', 'view_url']

    def test_csv_default_headers_and_row_count(self):
        resp = self.export()
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp['Content-Type'].startswith('text/csv'), resp['Content-Type'])
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertRegex(resp['Content-Disposition'], r'filename="receipts-\d{4}-\d{2}-\d{2}\.csv"')
        rows = self.csv_rows(resp)
        self.assertEqual(rows[0], self.EXPECTED_HEADER)
        self.assertEqual(len(rows), 12 + 1)
        self.assertEqual(len({r[12] for r in rows[1:]}), 12, 'duplicate sha rows in export')

    def test_csv_view_url_populated_and_valid(self):
        rows = self.csv_rows(self.export())
        idx = rows[0].index('view_url')
        sidx = rows[0].index('sha256')
        for r in rows[1:]:
            self.assertTrue(r[idx].startswith('http'), r[idx])
            self.assertIn(f'/audit/slip/{r[sidx]}/?s=', r[idx])
            sig = parse_qs(urlparse(r[idx]).query)['s'][0]
            self.assertEqual(sig, slip_signature(r[sidx]))

    def test_csv_respects_filters_and_ordering(self):
        rows = self.csv_rows(self.export(status='MATCHED', ordering='total'))
        self.assertEqual(len(rows), 3 + 1)
        self.assertEqual([r[2] for r in rows[1:]], ['77.00', '259.00', '1234.56'])
        rows = self.csv_rows(self.export(fy='FY26', category='Hardware'))
        self.assertEqual(len(rows), 3 + 1)
        rows = self.csv_rows(self.export(q='nothing-matches-this-zzz'))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0], self.EXPECTED_HEADER)

    def test_csv_ignores_pagination_params(self):
        rows = self.csv_rows(self.export(page=2, page_size=3))
        self.assertEqual(len(rows), 13, 'export must not be paginated')

    def test_csv_never_contains_file_bytes(self):
        resp = self.export()
        text = resp.content.decode('utf-8', 'replace')
        self.assertFalse(body_has_blob(text))
        self.assertNotIn('file_bytes', text)
        # 13 short CSV lines; a leaked 900-byte blob (or its base64/escaped form) would blow past this.
        self.assertLess(len(resp.content), 13 * 400)

    def test_csv_review_columns(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True, decision='CAPTURE', note='line1\nline2, "quoted"')
        rows = self.csv_rows(self.export(to_process='true'))
        self.assertEqual(len(rows), 2)
        rec = dict(zip(rows[0], rows[1]))
        self.assertEqual(rec['to_process'], 'True')
        self.assertEqual(rec['decision'], 'CAPTURE')
        self.assertEqual(rec['note'], 'line1\nline2, "quoted"')
        self.assertEqual(rec['synced'], 'True')
        self.assertEqual(rec['journal_number'], '697')
        self.assertEqual(rec['status_group'], 'MATCHED')

    def test_csv_null_values(self):
        rows = self.csv_rows(self.export(q='no-timestamp'))
        rec = dict(zip(rows[0], rows[1]))
        self.assertEqual(rec['date'], '')
        self.assertEqual(rec['total'], '15.00')
        self.assertEqual(rec['to_process'], 'False')
        self.assertEqual(rec['decision'], '')

    def test_xlsx_export(self):
        from openpyxl import load_workbook
        resp = self.export(format='xlsx', fy='FY26')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertRegex(resp['Content-Disposition'], r'attachment; filename="receipts-\d{4}-\d{2}-\d{2}\.xlsx"')
        self.assertNotIn('X-Export-Note', resp, 'openpyxl is installed; must not degrade to csv')
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, 'Receipts')
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), self.EXPECTED_HEADER)
        fy26 = self.list_(fy='FY26', page_size=200)['count']
        self.assertEqual(len(rows) - 1, fy26)
        self.assertEqual(fy26, 9)
        view_idx = self.EXPECTED_HEADER.index('view_url')
        self.assertTrue(all(str(r[view_idx]).startswith('http') for r in rows[1:]))
        # no blob in the zip payload either
        self.assertFalse(body_has_blob(resp.content.decode('latin-1')))

    def test_xlsx_with_blob_row_and_weird_cells(self):
        from openpyxl import load_workbook
        # The blob row + unicode note + newline note must all serialise.
        SlipReview.objects.create(sha256=self.NAMES['c_fy25_pend'], note='ünï\ncode =SUM(1)')
        resp = self.export(format='xlsx')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        ws = load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws.max_row, 13)

    def test_invalid_format_is_400(self):
        for f in ('pdf', 'json', 'CSV;DROP', 'xls', '../../etc/passwd'):
            resp = self.export(format=f)
            self.assertEqual(resp.status_code, 400, f)
            self.assertEqual(resp['Content-Type'], 'application/json')
        self.assertEqual(self.export(format='CSV').status_code, 200, 'format is case-insensitive')
        self.assertEqual(self.export(format='XLSX').status_code, 200)

    def test_export_rejects_non_get(self):
        self.assertEqual(self.client.post(reverse('receipts:export')).status_code, 405)
        self.assertEqual(self.client.delete(reverse('receipts:export')).status_code, 405)

    def test_export_junk_filters_do_not_500(self):
        resp = self.export(ordering='file_bytes', min_total='NaN', fy='zzz', q="'; drop table x; --", page='x', status='x', decision='y')
        self.assertEqual(resp.status_code, 200)


# --------------------------------------------------------------------------- #
# 8. view_url signature + base URL
# --------------------------------------------------------------------------- #
class ViewUrlTests(ReceiptsFixtureMixin, TestCase):
    def _row(self, name='a_fy26'):
        return next(r for r in self.list_(page_size=200)['results'] if r['sha256'] == self.NAMES[name])

    def test_signature_matches_slip_signature_and_viewer_accepts_it(self):
        from django.conf import settings
        s = self.NAMES['c_fy25_pend']  # the blob row, so the viewer has bytes to serve
        row = self._row('c_fy25_pend')
        sig = slip_signature(s)
        self.assertEqual(len(sig), 32)
        # Exact shape: <SLIP_VIEW_BASE_URL>/audit/slip/<sha>/?s=<sig>  (base may carry the nginx /backend prefix)
        self.assertEqual(row['view_url'], f'{settings.SLIP_VIEW_BASE_URL}/audit/slip/{s}/?s={sig}')
        url = urlparse(row['view_url'])
        self.assertEqual(parse_qs(url.query)['s'][0], sig)
        self.assertTrue(url.path.endswith(f'/audit/slip/{s}/'))
        # Signed viewer honours it (hits the real Django route; the base prefix is nginx's).
        resp = self.client.get(f'/audit/slip/{s}/', {'s': sig})
        self.assertEqual(resp.status_code, 200, resp.content[:200])
        self.assertEqual(resp.content, BLOB)
        self.assertEqual(resp['Content-Type'], 'image/jpeg')
        self.assertIn('inline', resp['Content-Disposition'])

    def test_tampered_signature_is_rejected(self):
        s = self.NAMES['c_fy25_pend']
        good = slip_signature(s)
        path = f'/audit/slip/{s}/'
        bad = ('0' if good[0] != '0' else '1') + good[1:]
        self.assertEqual(self.client.get(path, {'s': bad}).status_code, 403)
        self.assertEqual(self.client.get(path, {'s': good[:-1]}).status_code, 403)
        self.assertEqual(self.client.get(path, {'s': good + 'a'}).status_code, 403)
        self.assertEqual(self.client.get(path, {'s': ''}).status_code, 403)
        self.assertEqual(self.client.get(path).status_code, 403)
        if good != good.upper():
            self.assertEqual(self.client.get(path, {'s': good.upper()}).status_code, 403)
        # A valid signature for slip X must not open slip Y.
        other = self.NAMES['a_fy26']
        self.assertEqual(self.client.get(f'/audit/slip/{other}/', {'s': good}).status_code, 403)

    def test_signature_for_slip_without_bytes_is_valid_but_404(self):
        s = self.NAMES['a_fy26']  # no file_bytes in fixture
        self.assertEqual(self.client.get(f'/audit/slip/{s}/', {'s': slip_signature(s)}).status_code, 404)

    def test_default_base_url(self):
        with override_settings(SLIP_VIEW_BASE_URL='https://console.8-bit.space/backend'):
            self.assertTrue(self._row()['view_url'].startswith('https://console.8-bit.space/backend/audit/slip/'))

    def test_base_url_honours_setting(self):
        with override_settings(SLIP_VIEW_BASE_URL='https://example.test/xyz'):
            row = self._row()
            self.assertTrue(row['view_url'].startswith('https://example.test/xyz/audit/slip/'), row['view_url'])
            rows = self.csv_rows(self.export(q='Builders'))
            self.assertTrue(all(r[-1].startswith('https://example.test/xyz/') for r in rows[1:]))
            d = self.client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).json()['view_url']
            self.assertTrue(d.startswith('https://example.test/xyz/'))
        with override_settings(SLIP_VIEW_BASE_URL='http://localhost:8000'):
            self.assertTrue(self._row()['view_url'].startswith('http://localhost:8000/audit/slip/'))

    def test_signature_changes_with_secret_key(self):
        s = self.NAMES['a_fy26']
        before = slip_signature(s)
        with override_settings(SECRET_KEY='a-completely-different-secret-key-for-the-test'):
            after = slip_signature(s)
            self.assertNotEqual(before, after)
            self.assertEqual(parse_qs(urlparse(self._row()['view_url']).query)['s'][0], after)


# --------------------------------------------------------------------------- #
# 9. Method / auth matrix (every endpoint is now IsAuthenticated)
# --------------------------------------------------------------------------- #
class MethodMatrixTests(ReceiptsFixtureMixin, TestCase):
    def test_reads_are_gated_like_the_writes(self):
        # Reads and the export were AllowAny; they now require authentication like the writes.
        for url in (reverse('receipts:list'), reverse('receipts:detail', args=[self.NAMES['a_fy26']]),
                    reverse('receipts:export')):
            self.assertEqual(self.anon.get(url).status_code, 401, url)
            self.assertEqual(self.client.get(url).status_code, 200, url)
        self.assertEqual(self.anon.patch(reverse('receipts:review', args=[self.NAMES['a_fy26']]), {'note': 'x'}, format='json').status_code, 401)
        self.assertEqual(self.anon.post(reverse('receipts:comments', args=[self.NAMES['a_fy26']]), {'text': 'x'}, format='json').status_code, 401)

    def test_wrong_methods_are_405_not_writes(self):
        s = self.NAMES['a_fy26']
        # Anonymous wrong-method is 401, not 405: DRF runs authentication/permissions (initial())
        # before method routing, so an anonymous caller cannot even probe the allowed-method map.
        self.assertEqual(self.anon.post(reverse('receipts:list'), {}, format='json').status_code, 401)
        self.assertEqual(self.anon.delete(reverse('receipts:detail', args=[s])).status_code, 401)
        # Authenticated wrong methods are 405 and never write.
        self.assertEqual(self.client.post(reverse('receipts:list'), {}, format='json').status_code, 405)
        self.assertEqual(self.client.delete(reverse('receipts:detail', args=[s])).status_code, 405)
        self.assertEqual(self.client.put(reverse('receipts:detail', args=[s]), {}, format='json').status_code, 405)
        self.assertEqual(self.client.post(reverse('receipts:review', args=[s]), {'note': 'x'}, format='json').status_code, 405)
        self.assertEqual(self.client.put(reverse('receipts:review', args=[s]), {'note': 'x'}, format='json').status_code, 405)
        self.assertEqual(self.client.delete(reverse('receipts:review', args=[s])).status_code, 405)
        self.assertEqual(self.client.patch(reverse('receipts:comments', args=[s]), {'text': 'x'}, format='json').status_code, 405)
        self.assertEqual(self.client.delete(reverse('receipts:comments', args=[s])).status_code, 405)
        self.assertFalse(SlipReview.objects.exists())
        self.assertFalse(SlipComment.objects.exists())

    def test_export_and_list_urls_do_not_collide_with_a_sha_named_export(self):
        # /audit/receipts/export/ must be the export, never the detail of a slip literally named "export"
        insert_slip('export', slip_ts=ts(2025, 1, 1), ocr={'total': '1.00'})
        resp = self.client.get('/audit/receipts/export/')
        self.assertTrue(resp['Content-Type'].startswith('text/csv'))


# --------------------------------------------------------------------------- #
# 10. Auth gate — every endpoint requires authentication (change 1)
# --------------------------------------------------------------------------- #
class AuthGateTests(ReceiptsFixtureMixin, TestCase):
    """
    Adversarial coverage of the authentication gate. The whole point of gating the
    reads is that an anonymous caller can no longer harvest signed ``view_url``s
    (permanent links to receipt images), supplier names, or OCR money fields.
    The export runs a SEPARATE auth code path (the ``drf_login_required`` decorator,
    not a DRF permission class), so it is exercised independently throughout.
    """

    def endpoints(self, client, *, body=True):
        """(name, response) for all five endpoints, called with ``client``."""
        s = self.NAMES['a_fy26']
        return [
            ('list', client.get(reverse('receipts:list'))),
            ('detail', client.get(reverse('receipts:detail', args=[s]))),
            ('export_csv', client.get(reverse('receipts:export'))),
            ('export_xlsx', client.get(reverse('receipts:export'), {'format': 'xlsx'})),
            ('review', client.patch(reverse('receipts:review', args=[s]),
                                    {'to_process': True, 'note': 'authed'} if body else None, format='json')),
            ('comments', client.post(reverse('receipts:comments', args=[s]),
                                     {'text': 'authed comment'} if body else None, format='json')),
        ]

    def test_anonymous_caller_gets_401_on_all_five_endpoints(self):
        for name, resp in self.endpoints(self.anon):
            self.assertEqual(resp.status_code, 401, f'{name}: anonymous request must be 401')
        # and the gated writes wrote nothing
        self.assertFalse(SlipReview.objects.exists())
        self.assertFalse(SlipComment.objects.exists())

    def test_anonymous_401_bodies_leak_nothing(self):
        # An anonymous 401 must not hand out anything harvestable: no signed view_url,
        # no slip sha256, no supplier, no CSV header row, no signature, no blob bytes.
        sig = slip_signature(self.NAMES['c_fy25_pend'])
        markers = ['view_url', '/audit/slip/', 'Builders Warehouse', 'Engen Stellenbosch',
                   'date,supplier,total', sig, *self.NAMES.values()]
        for name, resp in self.endpoints(self.anon):
            self.assertEqual(resp.status_code, 401, name)
            text = resp.content.decode('utf-8', 'replace')
            for marker in markers:
                self.assertNotIn(marker, text, f'{name}: anonymous 401 body leaks {marker!r}')
            self.assertFalse(body_has_blob(text), f'{name}: anonymous 401 body leaks file bytes')

    def test_valid_jwt_succeeds_on_all_five_endpoints(self):
        by_name = dict(self.endpoints(self.client))
        self.assertEqual(by_name['list'].status_code, 200)
        self.assertEqual(by_name['list'].json()['count'], 12)
        self.assertEqual(by_name['detail'].status_code, 200)
        self.assertEqual(by_name['detail'].json()['sha256'], self.NAMES['a_fy26'])
        self.assertEqual(by_name['export_csv'].status_code, 200)
        self.assertTrue(by_name['export_csv']['Content-Type'].startswith('text/csv'))
        self.assertEqual(by_name['export_xlsx'].status_code, 200)
        self.assertEqual(by_name['export_xlsx']['Content-Type'],
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertEqual(by_name['review'].status_code, 200)
        self.assertEqual(by_name['comments'].status_code, 201)

    def test_review_write_stamps_the_authenticated_username(self):
        resp = self.client.patch(reverse('receipts:review', args=[self.NAMES['a_fy26']]),
                                 {'decision': 'CAPTURE'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated_by'], 'reviewer', 'updated_by must be the JWT user, not blank')
        self.assertEqual(SlipReview.objects.get(sha256=self.NAMES['a_fy26']).updated_by, 'reviewer')

    def test_garbage_bearer_token_is_401_on_drf_view_and_export(self):
        # The export's auth is a decorator, not a permission class — exercise both code paths.
        for garbage in ('not.a.jwt', 'a' * 40, 'ey.ey.ey', 'null'):
            client = APIClient()
            client.credentials(HTTP_AUTHORIZATION=f'Bearer {garbage}')
            self.assertEqual(client.get(reverse('receipts:list')).status_code, 401, garbage)
            resp = client.get(reverse('receipts:export'))
            self.assertEqual(resp.status_code, 401, garbage)
            self.assertNotIn('date,supplier,total', resp.content.decode('utf-8', 'replace'))

    def test_expired_token_is_401_on_drf_view_and_export(self):
        token = AccessToken.for_user(self.user)
        token.set_exp(lifetime=-dt.timedelta(seconds=1))  # already expired when serialised
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        self.assertEqual(client.get(reverse('receipts:list')).status_code, 401)
        self.assertEqual(client.get(reverse('receipts:detail', args=[self.NAMES['a_fy26']])).status_code, 401)
        resp = client.get(reverse('receipts:export'))
        self.assertEqual(resp.status_code, 401, 'export must reject an expired token via its own auth path')
        self.assertNotIn('date,supplier,total', resp.content.decode('utf-8', 'replace'))

    def test_export_401_carries_www_authenticate(self):
        resp = self.anon.get(reverse('receipts:export'))
        self.assertEqual(resp.status_code, 401)
        self.assertIn('Bearer', resp.headers.get('WWW-Authenticate', ''),
                      'export 401 must carry a WWW-Authenticate challenge')

    def test_export_auth_is_checked_before_format_validation(self):
        # An anonymous caller must not be able to distinguish a valid format from a bogus
        # one (401 for both, identical body) — auth runs before the 400 format check.
        good = self.anon.get(reverse('receipts:export'), {'format': 'xlsx'})
        bogus = self.anon.get(reverse('receipts:export'), {'format': 'exe'})
        self.assertEqual(good.status_code, 401)
        self.assertEqual(bogus.status_code, 401, 'anonymous bogus format must 401, not 400')
        self.assertEqual(good.content, bogus.content, 'anonymous caller can distinguish valid from invalid formats')

    def test_anonymous_non_get_export_never_returns_a_file(self):
        url = reverse('receipts:export')
        for method in ('post', 'put', 'delete'):
            resp = getattr(self.anon, method)(url)
            self.assertEqual(resp.status_code, 401, f'{method}: auth gate must run before method routing')
            self.assertNotEqual(resp.status_code, 200)
            self.assertFalse(resp['Content-Type'].startswith('text/csv'), method)
            self.assertNotIn('attachment', resp.headers.get('Content-Disposition', ''), method)
            self.assertNotIn(b'date,supplier,total', resp.content, method)

    def test_signed_slip_viewer_stays_public(self):
        # DELIBERATELY ungated: the console's <img>/<iframe> cannot send a Bearer token and
        # exported spreadsheets link here. A regression to 401/403 with a VALID signature
        # silently breaks the console's receipt modal — pin the contract both ways.
        s = self.NAMES['c_fy25_pend']  # the fixture row with file_bytes set
        good = slip_signature(s)
        resp = self.anon.get(f'/audit/slip/{s}/', {'s': good})
        self.assertEqual(resp.status_code, 200, 'anonymous signed viewer must stay reachable')
        self.assertEqual(resp.content, BLOB)
        bad = ('0' if good[0] != '0' else '1') + good[1:]
        self.assertEqual(self.anon.get(f'/audit/slip/{s}/', {'s': bad}).status_code, 403)
        self.assertEqual(self.anon.get(f'/audit/slip/{s}/').status_code, 403)

    def test_authenticated_export_still_never_leaks_file_bytes(self):
        # The blob-leak guarantee must hold on the now-authenticated export path too.
        csv_resp = self.client.get(reverse('receipts:export'))
        self.assertEqual(csv_resp.status_code, 200)
        self.assertFalse(body_has_blob(csv_resp.content.decode('utf-8', 'replace')))
        xlsx_resp = self.client.get(reverse('receipts:export'), {'format': 'xlsx'})
        self.assertEqual(xlsx_resp.status_code, 200)
        self.assertFalse(body_has_blob(xlsx_resp.content.decode('latin-1')))


# --------------------------------------------------------------------------- #
# 11. Journal join — journal_type discrimination + org fallback (change 2)
# --------------------------------------------------------------------------- #
class JournalTypeAndOrgScopeTests(ReceiptsFixtureMixin, TestCase):
    """
    The lateral join now requires journal_type = services.JOURNAL_TYPE ('journal') and
    resolves the org as: blank/NULL xero_org -> services.KLIKK_TENANT_ID, else tenant
    name lookup (unknown name -> no journal). The summary aggregates DEBIT lines only,
    falling back to the whole journal when there is no debit line.
    """

    def detail(self, s):
        resp = self.client.get(reverse('receipts:detail', args=[s]))
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()

    def test_number_existing_only_as_other_types_resolves_no_journal(self):
        # No fallback to manual_journal / system_journal / transaction when the number
        # does not exist under type 'journal' in the slip's org.
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_hw = XeroAccount.objects.get(account_id='acc-k-hw')
        d = ts(2026, 2, 1)
        for i, jtype in enumerate(('manual_journal', 'system_journal', 'transaction')):
            num = 810 + i
            make_line(self.klikk, num, k_hw, '250.00', journal_id=f'k{num}-1', date=d,
                      description=f'{jtype} expense', journal_type=jtype)
            make_line(self.klikk, num, k_bank, '-250.00', journal_id=f'k{num}-2', date=d,
                      description=f'{jtype} bank', journal_type=jtype)
            s = sha(400 + i)
            insert_slip(s, slip_ts=d, xero_status='MATCHED', xero_detail=f'journal_number={num}',
                        journal_number=num, xero_org='Klikk (Pty) Ltd',
                        ocr={'supplier': 'Some Shop', 'total': '250.00', 'category': 'Hardware'})
            row = self.detail(s)
            self.assertIsNone(row['journal'], f'{jtype}: endpoint must not fall back to a non-journal type')
            self.assertEqual(row['journal_number'], num)
        # and the list agrees
        rows = {r['sha256']: r for r in self.list_(page_size=200)['results']}
        for i in range(3):
            self.assertIsNone(rows[sha(400 + i)]['journal'])

    def test_journal_type_does_not_cross_orgs(self):
        # #901 is a 'journal' in Klikk and a 'manual_journal' in Dippenaar Family. A dfam slip
        # must resolve NOTHING (its org has no type-'journal' #901), never Klikk's lines.
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_meals = XeroAccount.objects.get(account_id='acc-k-meals')
        d_groc = XeroAccount.objects.get(account_id='acc-d-groc')
        d_bank = XeroAccount.objects.get(account_id='acc-d-bank')
        d = ts(2026, 2, 10)
        make_line(self.klikk, 901, k_meals, '210.00', journal_id='k901-1', date=d, description='Klikk 901 expense')
        make_line(self.klikk, 901, k_bank, '-210.00', journal_id='k901-2', date=d, description='Klikk 901 bank')
        make_line(self.dfam, 901, d_groc, '999.00', journal_id='d901-1', date=d,
                  description='Dfam 901 manual', journal_type='manual_journal')
        make_line(self.dfam, 901, d_bank, '-999.00', journal_id='d901-2', date=d,
                  description='Dfam 901 manual', journal_type='manual_journal')
        insert_slip(sha(405), slip_ts=d, journal_number=901, xero_org='Dippenaar Family',
                    ocr={'supplier': 'Dfam Shop', 'total': '999.00'})
        insert_slip(sha(406), slip_ts=d, journal_number=901, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Klikk Shop', 'total': '210.00'})
        self.assertIsNone(self.detail(sha(405))['journal'],
                          "dfam slip must not resolve Klikk's 'journal' #901 across orgs")
        j = self.detail(sha(406))['journal']
        self.assertIsNotNone(j)
        self.assertEqual(Decimal(j['amount']), Decimal('210.00'))
        self.assertEqual(j['description'], 'Klikk 901 expense')

    def test_null_and_blank_org_both_fall_back_to_the_klikk_tenant(self):
        # The fallback resolves by tenant_id == services.KLIKK_TENANT_ID (the constant), not by
        # name: the fixture's 'Klikk (Pty) Ltd' tenant deliberately has a DIFFERENT tenant_id.
        live = make_tenant(KLIKK_TENANT_ID, 'Klikk Live (Pty) Ltd')
        acc_exp = make_account(live, 'acc-live-exp', '460', 'Consumables', 'EXPENSE')
        acc_bank = make_account(live, 'acc-live-bank', '091', 'Live Bank', 'BANK')
        d = ts(2026, 3, 3)
        make_line(live, 950, acc_exp, '120.00', journal_id='live950-1', date=d, description='Fallback coffee')
        make_line(live, 950, acc_bank, '-120.00', journal_id='live950-2', date=d, description='Fallback coffee')
        insert_slip(sha(410), slip_ts=d, journal_number=950, xero_org=None,
                    ocr={'supplier': 'Coffee Shop', 'total': '120.00'})
        insert_slip(sha(411), slip_ts=d, journal_number=950, xero_org='',
                    ocr={'supplier': 'Coffee Shop', 'total': '120.00'})
        for s, label in ((sha(410), 'NULL xero_org'), (sha(411), 'empty-string xero_org')):
            j = self.detail(s)['journal']
            self.assertIsNotNone(j, f'{label} must fall back to the Klikk tenant')
            self.assertEqual(Decimal(j['amount']), Decimal('120.00'), label)
            self.assertEqual(j['description'], 'Fallback coffee', label)
            self.assertEqual(j['account_code'], '460', label)
        # A slip NAMING a tenant never falls back: dfam has no #950, so no journal.
        insert_slip(sha(412), slip_ts=d, journal_number=950, xero_org='Dippenaar Family',
                    ocr={'supplier': 'Coffee Shop', 'total': '120.00'})
        self.assertIsNone(self.detail(sha(412))['journal'],
                          'a named org must not fall back to the Klikk tenant')

    def test_summary_comes_from_debit_lines_only_never_the_credit_or_bank_side(self):
        # Debit and credit sides carry different descriptions/accounts/contacts. The bank
        # debit is the LARGEST debit, so if the BANK demotion were dropped the summary would
        # pick it; and if credit lines leaked in, amount would exceed the debit sum.
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_hw = XeroAccount.objects.get(account_id='acc-k-hw')
        sub = make_account(self.klikk, 'acc-k-sub', '431', 'Subscriptions', 'EXPENSE')
        liab = make_account(self.klikk, 'acc-k-liab', '881', 'Sundry Creditors', 'CURRLIAB')
        deb_contact = make_contact(self.klikk, 'con-k-deb', 'Debit Side Contact')
        cred_contact = make_contact(self.klikk, 'con-k-cred', 'Credit Side Contact')
        d = ts(2026, 4, 20)
        make_line(self.klikk, 920, sub, '150.00', journal_id='k920-1', date=d,
                  description='Debit main line', contact=deb_contact)
        make_line(self.klikk, 920, k_hw, '50.00', journal_id='k920-2', date=d,
                  description='Debit minor line', contact=deb_contact)
        make_line(self.klikk, 920, k_bank, '300.00', journal_id='k920-3', date=d,
                  description='Bank debit line')
        make_line(self.klikk, 920, liab, '-500.00', journal_id='k920-4', date=d,
                  description='Credit side line', contact=cred_contact)
        insert_slip(sha(420), slip_ts=d, journal_number=920, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Sub Shop', 'total': '500.00'})
        j = self.detail(sha(420))['journal']
        self.assertIsNotNone(j)
        self.assertEqual(Decimal(j['amount']), Decimal('500.00'), 'amount must be the sum of DEBIT lines only')
        self.assertEqual(Decimal(j['debit']), Decimal('500.00'))
        self.assertEqual(Decimal(j['credit']), Decimal('-500.00'))
        self.assertEqual(j['description'], 'Debit main line', 'main line must be the largest NON-BANK debit')
        self.assertNotEqual(j['description'], 'Bank debit line')
        self.assertNotEqual(j['description'], 'Credit side line')
        self.assertEqual(j['account_code'], '431')
        self.assertNotIn(j['account_code'], ('090', '881'))
        self.assertEqual(j['contact_name'], 'Debit Side Contact')
        self.assertNotEqual(j['contact_name'], 'Credit Side Contact')

    def test_credit_only_journal_falls_back_to_whole_journal_not_nulls(self):
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_hw = XeroAccount.objects.get(account_id='acc-k-hw')
        ref_contact = make_contact(self.klikk, 'con-k-ref', 'Refund Contact')
        d = ts(2026, 5, 5)
        make_line(self.klikk, 930, k_hw, '-100.00', journal_id='k930-1', date=d,
                  description='Refund line', contact=ref_contact)
        make_line(self.klikk, 930, k_bank, '-50.00', journal_id='k930-2', date=d,
                  description='Bank refund line')
        insert_slip(sha(425), slip_ts=d, journal_number=930, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Refund Shop', 'total': '100.00'})
        j = self.detail(sha(425))['journal']
        self.assertIsNotNone(j, 'a credit-only journal must still resolve')
        self.assertIsNotNone(j['description'], 'no-debit journal must fall back, not return null description')
        self.assertIsNotNone(j['account_code'], 'no-debit journal must fall back, not return null account')
        self.assertEqual(j['description'], 'Refund line')  # non-bank line first in the fallback too
        self.assertEqual(j['account_code'], '429')
        self.assertEqual(j['contact_name'], 'Refund Contact')
        self.assertEqual(Decimal(j['credit']), Decimal('-150.00'))
        self.assertEqual(Decimal(j['amount']), Decimal('0.00'), 'no debit lines: amount falls back to sum(debit) = 0')

    def test_multi_line_journal_yields_one_row_in_list_and_export(self):
        k_bank = XeroAccount.objects.get(account_id='acc-k-bank')
        k_hw = XeroAccount.objects.get(account_id='acc-k-hw')
        k_vat = XeroAccount.objects.get(account_id='acc-k-vat')
        d = ts(2026, 5, 20)
        for i, (acc, amt) in enumerate(((k_hw, '400.00'), (k_hw, '35.00'), (k_vat, '65.25'), (k_bank, '-500.25'))):
            make_line(self.klikk, 940, acc, amt, journal_id=f'k940-{i}', date=d, description='Big basket')
        insert_slip(sha(430), slip_ts=d, journal_number=940, xero_org='Klikk (Pty) Ltd',
                    ocr={'supplier': 'Basket Shop', 'total': '500.25'})
        data = self.list_(page_size=200)
        self.assertEqual(data['count'], 13, 'list count must not fan out on the 4-line journal')
        self.assertEqual(data['totals']['count'], 13)
        hits = [r for r in data['results'] if r['sha256'] == sha(430)]
        self.assertEqual(len(hits), 1)
        self.assertEqual(Decimal(hits[0]['journal']['amount']), Decimal('500.25'))
        rows = self.csv_rows(self.export())
        self.assertEqual(len(rows), 13 + 1, 'export must have one row per slip, no fan-out')
        self.assertEqual(sum(1 for r in rows[1:] if r[12] == sha(430)), 1)

    def test_export_journal_number_column_agrees_with_detail(self):
        rows = self.csv_rows(self.export())
        by_sha = {r[12]: dict(zip(rows[0], r)) for r in rows[1:]}
        # resolving slip: export column == detail journal.journal_number == slip journal_number
        detail_a = self.detail(self.NAMES['a_fy26'])
        self.assertEqual(detail_a['journal']['journal_number'], 697)
        self.assertEqual(by_sha[self.NAMES['a_fy26']]['journal_number'], '697')
        # orphan number: journal is null in detail but the register's number still exports
        detail_l = self.detail(self.NAMES['l_orphan_jn'])
        self.assertIsNone(detail_l['journal'])
        self.assertEqual(by_sha[self.NAMES['l_orphan_jn']]['journal_number'], '555')
        # no number at all -> empty cell
        self.assertEqual(by_sha[self.NAMES['c_fy25_pend']]['journal_number'], '')


# --------------------------------------------------------------------------- #
# 12. Bulk endpoint — auth gate + request shape (POST /audit/receipts/bulk/)
# --------------------------------------------------------------------------- #
def register_snapshot():
    """Every column of every register row, normalised for bytewise comparison."""
    with connection.cursor() as cur:
        cur.execute(
            'select sha256, slip_ts, filename, source, mime_ext, byte_size, file_bytes, xero_status, '
            'xero_detail, imported_at, synced_to_xero, ocr::text, search_tsv::text, journal_number, xero_org '
            'from whatsapp.klikk_slips order by sha256'
        )
        return [tuple(bytes(v) if isinstance(v, memoryview) else v for v in row) for row in cur.fetchall()]


class BulkMixin(ReceiptsFixtureMixin):
    def bulk_url(self):
        return reverse('receipts:bulk')

    def bulk(self, payload, client=None):
        return (client or self.client).post(self.bulk_url(), payload, format='json')

    def raw_client(self):
        client = APIClient(raise_request_exception=False)
        client.credentials(**self.client._credentials)
        return client

    def assert_nothing_written(self):
        self.assertFalse(SlipReview.objects.exists(), 'a rejected bulk call wrote a SlipReview')
        self.assertFalse(SlipComment.objects.exists(), 'a rejected bulk call wrote a SlipComment')


class BulkAuthAndShapeTests(BulkMixin, TestCase):
    def test_anonymous_post_is_401_and_writes_nothing(self):
        payload = {'sha256s': [self.NAMES['a_fy26']], 'decision': 'CAPTURE', 'comment': 'anon'}
        resp = self.anon.post(self.bulk_url(), payload, format='json')
        self.assertEqual(resp.status_code, 401, resp.content[:300])
        self.assertNotIn('updated', resp.content.decode('utf-8', 'replace'))
        self.assert_nothing_written()

    def test_malformed_or_expired_bearer_is_401_and_writes_nothing(self):
        payload = {'sha256s': [self.NAMES['a_fy26']], 'note': 'x'}
        for garbage in ('not.a.jwt', 'a' * 40, 'ey.ey.ey', ''):
            client = APIClient()
            client.credentials(HTTP_AUTHORIZATION=f'Bearer {garbage}')
            self.assertEqual(client.post(self.bulk_url(), payload, format='json').status_code, 401, repr(garbage))
        token = AccessToken.for_user(self.user)
        token.set_exp(lifetime=-dt.timedelta(seconds=1))  # already expired when serialised
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        self.assertEqual(client.post(self.bulk_url(), payload, format='json').status_code, 401, 'expired token')
        self.assert_nothing_written()

    def test_token_for_deleted_or_deactivated_user_is_401(self):
        ghost = User.objects.create_user(username='ghost', email='ghost@example.com', password='pw-irrelevant')
        token = str(AccessToken.for_user(ghost))
        ghost.delete()
        payload = {'sha256s': [self.NAMES['a_fy26']], 'note': 'x'}
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        self.assertEqual(client.post(self.bulk_url(), payload, format='json').status_code, 401,
                         'a token for a since-deleted user must be rejected')
        frozen = User.objects.create_user(username='frozen', email='frozen@example.com',
                                          password='pw-irrelevant', is_active=False)
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(frozen)}')
        self.assertEqual(client.post(self.bulk_url(), payload, format='json').status_code, 401,
                         'a token for a deactivated user must be rejected')
        self.assert_nothing_written()

    def test_get_bulk_is_405_and_never_the_detail_of_a_slip_named_bulk(self):
        # THE URLCONF-ORDER CANARY: '<str:sha256>/' sits one line below 'bulk/' in urls.py. Insert a
        # register row literally keyed 'bulk'; if anyone reorders the URLconf, GET /audit/receipts/bulk/
        # resolves that slip (200 + signed view_url) instead of 405ing — this is the test that screams.
        insert_slip('bulk', slip_ts=ts(2025, 6, 1), filename='bulk-trap.jpg',
                    ocr={'supplier': 'BulkTrap Supplier', 'total': '666.00'})
        resp = self.client.get(self.bulk_url())
        self.assertEqual(resp.status_code, 405, 'GET bulk/ must be method-not-allowed, not a slip detail')
        text = resp.content.decode('utf-8', 'replace')
        for marker in ('BulkTrap', 'view_url', 'xero_status', '666.00', '/audit/slip/'):
            self.assertNotIn(marker, text, f'GET bulk/ was routed to receipt_detail_view (leaked {marker!r})')
        # POSTing FOR the slip named 'bulk' proves the POST reached the bulk view through the same URL.
        resp = self.bulk({'sha256s': ['bulk'], 'note': 'trap'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 1, 'commented': 0, 'unknown': []})
        # anonymous GET is gated before method routing; other methods are 405 and never write
        self.assertEqual(self.anon.get(self.bulk_url()).status_code, 401)
        self.assertEqual(self.client.put(self.bulk_url(), {'sha256s': ['bulk'], 'note': 'x'}, format='json').status_code, 405)
        self.assertEqual(self.client.patch(self.bulk_url(), {'sha256s': ['bulk'], 'note': 'x'}, format='json').status_code, 405)
        self.assertEqual(self.client.delete(self.bulk_url()).status_code, 405)
        self.assertEqual(SlipReview.objects.count(), 1, 'only the POST may write')

    def test_non_object_json_bodies_are_400_never_500(self):
        client = self.raw_client()
        for raw in ('[]', '["sha256s"]', '"sha256s"', '42', 'null', 'true', '{bad json', ''):
            resp = client.generic('POST', self.bulk_url(), data=raw, content_type='application/json')
            self.assertEqual(resp.status_code, 400, f'body={raw!r} -> {resp.status_code}: {resp.content[:200]}')
        self.assert_nothing_written()

    def test_sha256s_shape_garbage_is_400_and_writes_nothing(self):
        a = self.NAMES['a_fy26']
        bad = [
            {'note': 'x'},                          # sha256s missing entirely
            {'sha256s': None, 'note': 'x'},
            {'sha256s': 'abc', 'note': 'x'},        # a string, not a list
            {'sha256s': 123, 'note': 'x'},
            {'sha256s': {'0': a}, 'note': 'x'},
            {'sha256s': [], 'note': 'x'},
            {'sha256s': [''], 'note': 'x'},
            {'sha256s': ['   '], 'note': 'x'},
            {'sha256s': ['\t\n'], 'note': 'x'},
            {'sha256s': [None], 'note': 'x'},
            {'sha256s': [123], 'note': 'x'},
            {'sha256s': [True], 'note': 'x'},       # bool is not a sha string
            {'sha256s': [[a]], 'note': 'x'},        # nested list
            {'sha256s': [a, ''], 'note': 'x'},      # one bad entry poisons the whole batch
            {'sha256s': [a, None], 'note': 'x'},
            {'sha256s': [a, 42], 'note': 'x'},
        ]
        for payload in bad:
            resp = self.bulk(payload)
            self.assertEqual(resp.status_code, 400, (payload, resp.content[:200]))
        self.assert_nothing_written()

    def test_form_encoded_body_is_400_not_500(self):
        resp = self.client.post(self.bulk_url(), data=f'sha256s={self.NAMES["a_fy26"]}&note=x',
                                content_type='application/x-www-form-urlencoded')
        self.assertEqual(resp.status_code, 400, resp.content[:300])
        self.assert_nothing_written()


# --------------------------------------------------------------------------- #
# 13. Bulk endpoint — the 500 cap is counted AFTER de-duplication
# --------------------------------------------------------------------------- #
class BulkCapAndDedupTests(BulkMixin, TestCase):
    def test_exactly_500_distinct_sha256s_are_accepted(self):
        known = [self.NAMES['a_fy26'], self.NAMES['b_fy26_auto'], self.NAMES['c_fy25_pend']]
        unknown = [sha(20000 + i) for i in range(497)]
        resp = self.bulk({'sha256s': known + unknown, 'note': 'cap check'})
        self.assertEqual(resp.status_code, 200, f'exactly 500 distinct must be accepted: {resp.content[:300]}')
        body = resp.json()
        self.assertEqual(body['updated'], 3)
        self.assertEqual(body['unknown'], unknown)
        self.assertEqual(SlipReview.objects.count(), 3)

    def test_501_distinct_sha256s_are_400_and_write_nothing(self):
        shas_ = [sha(21000 + i) for i in range(500)] + [self.NAMES['a_fy26']]
        resp = self.bulk({'sha256s': shas_, 'note': 'x'})
        self.assertEqual(resp.status_code, 400, resp.content[:300])
        self.assertIn('500', resp.json().get('detail', ''))
        self.assert_nothing_written()

    def test_600_entries_deduplicating_to_3_are_accepted_cap_is_post_dedup(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        u = sha(22000)
        raw = [a, b, u] * 200  # 600 entries, 3 distinct
        self.assertEqual(len(raw), 600)
        resp = self.bulk({'sha256s': raw, 'note': 'dedup first'})
        self.assertEqual(resp.status_code, 200,
                         f'600 entries -> 3 distinct must be accepted (cap counts AFTER de-dup): {resp.content[:300]}')
        body = resp.json()
        self.assertEqual(body['updated'], 2)
        self.assertEqual(body['unknown'], [u], 'unknown must be de-duplicated too')
        self.assertEqual(SlipReview.objects.count(), 2)
        # whitespace-padded variants of one sha de-duplicate with the clean one (750 raw entries, 1 distinct)
        resp = self.bulk({'sha256s': [f'  {a}  ', a, f'{a} '] * 250, 'note': 'padded'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 1)
        self.assertEqual(resp.json()['unknown'], [])

    def test_duplicate_sha256s_produce_exactly_one_comment_per_distinct_receipt(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        resp = self.bulk({'sha256s': [a, a, b, a, b, b, a], 'comment': 'dup guard'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['commented'], 2)
        self.assertEqual(SlipComment.objects.count(), 2,
                         'one comment per INPUT ENTRY was written instead of one per distinct receipt')
        self.assertEqual(SlipComment.objects.filter(sha256=a).count(), 1)
        self.assertEqual(SlipComment.objects.filter(sha256=b).count(), 1)


# --------------------------------------------------------------------------- #
# 14. Bulk endpoint — actions (key presence, coercion, upsert semantics)
# --------------------------------------------------------------------------- #
class BulkActionTests(BulkMixin, TestCase):
    def test_no_action_key_is_400_nothing_to_do(self):
        a = self.NAMES['a_fy26']
        resp = self.bulk({'sha256s': [a]})
        self.assertEqual(resp.status_code, 400, resp.content[:300])
        self.assertIn('nothing to do', resp.json()['detail'])
        # the per-slip endpoints' field spellings are NOT bulk actions
        for payload in ({'sha256s': [a], 'to_process': True},      # bulk spells it set_to_process
                        {'sha256s': [a], 'text': 'a comment'},     # the comments endpoint's spelling
                        {'sha256s': [a], 'Decision': 'CAPTURE'}):  # keys are case-sensitive
            resp = self.bulk(payload)
            self.assertEqual(resp.status_code, 400, (payload, resp.content[:200]))
        self.assert_nothing_written()

    def test_note_empty_string_clears_a_previous_note(self):
        # Key PRESENCE decides, not truthiness: {"note": ""} is a legitimate clear.
        a = self.NAMES['a_fy26']
        SlipReview.objects.create(sha256=a, to_process=True, decision='CAPTURE', note='old note', updated_by='before')
        resp = self.bulk({'sha256s': [a], 'note': ''})
        self.assertEqual(resp.status_code, 200, f'note="" must be accepted as a clear: {resp.content[:300]}')
        self.assertEqual(resp.json()['updated'], 1)
        rv = SlipReview.objects.get(sha256=a)
        self.assertEqual(rv.note, '', 'the note was not cleared')
        self.assertEqual(rv.decision, 'CAPTURE', 'clearing the note must not touch the decision')
        self.assertTrue(rv.to_process, 'clearing the note must not touch to_process')
        self.assertEqual(rv.updated_by, 'reviewer')

    def test_set_to_process_coercion_stores_the_right_value(self):
        # A 200 must never mean "silently stored False" — assert the STORED value every time.
        a = self.NAMES['a_fy26']
        for raw in (True, 'true', 'TRUE', 'True', 1, '1', 'yes', 'Yes', 'on', 'ON'):
            SlipReview.objects.all().delete()
            resp = self.bulk({'sha256s': [a], 'set_to_process': raw})
            self.assertEqual(resp.status_code, 200, (raw, resp.content[:200]))
            self.assertEqual(resp.json()['updated'], 1, raw)
            self.assertTrue(SlipReview.objects.get(sha256=a).to_process,
                            f'set_to_process={raw!r}: 200 returned but False stored — the silent-False bug')
        for raw in (False, 'false', 'FALSE', 0, '0', 'no', 'No', 'off', 'OFF'):
            SlipReview.objects.update_or_create(sha256=a, defaults={'to_process': True})
            resp = self.bulk({'sha256s': [a], 'set_to_process': raw})
            self.assertEqual(resp.status_code, 200, (raw, resp.content[:200]))
            self.assertFalse(SlipReview.objects.get(sha256=a).to_process, f'set_to_process={raw!r} must store False')

    def test_set_to_process_unrecognised_values_are_400_and_never_silently_false(self):
        a = self.NAMES['a_fy26']
        SlipReview.objects.create(sha256=a, to_process=True)
        for raw in ('banana', [], {}, 'yess', '2', 2, 'null', [1], {'v': True}, 'true false'):
            resp = self.bulk({'sha256s': [a], 'set_to_process': raw})
            self.assertEqual(resp.status_code, 400, (raw, resp.content[:200]))
        rv = SlipReview.objects.get(sha256=a)
        self.assertTrue(rv.to_process, 'a rejected set_to_process value silently flipped the stored flag')
        self.assertEqual(SlipReview.objects.count(), 1)

    def test_set_to_process_json_null_is_an_explicit_clear_to_false(self):
        a = self.NAMES['a_fy26']
        SlipReview.objects.create(sha256=a, to_process=True)
        resp = self.bulk({'sha256s': [a], 'set_to_process': None})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertFalse(SlipReview.objects.get(sha256=a).to_process, 'set_to_process: null must clear to False')

    def test_decision_every_enum_value_lowercase_clear_and_invalid(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['d_fy26_nix']
        enum = ('CAPTURE', 'MEAL_SKIP', 'PERSONAL', 'DUPLICATE', 'ALREADY_IN_XERO')
        self.assertEqual(set(enum) | {''}, set(DECISION_VALUES), 'the decision enum drifted from the contract')
        for d in enum:
            resp = self.bulk({'sha256s': [a, b], 'decision': d})
            self.assertEqual(resp.status_code, 200, (d, resp.content[:200]))
            self.assertEqual(resp.json()['updated'], 2, d)
            self.assertEqual(SlipReview.objects.get(sha256=a).decision, d)
            self.assertEqual(SlipReview.objects.get(sha256=b).decision, d)
        resp = self.bulk({'sha256s': [a], 'decision': 'capture'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(SlipReview.objects.get(sha256=a).decision, 'CAPTURE', 'lower-case decision must be upper-cased')
        resp = self.bulk({'sha256s': [a], 'decision': ''})
        self.assertEqual(resp.status_code, 200, 'decision="" is a legitimate clear')
        self.assertEqual(SlipReview.objects.get(sha256=a).decision, '')
        self.assertEqual(SlipReview.objects.get(sha256=b).decision, 'ALREADY_IN_XERO', "clearing a must not touch b's decision")
        for d in ('NOPE', 'CAPTURE;DROP', 'MEAL SKIP', 'MEAL-SKIP', 7, ['CAPTURE'], {'d': 'CAPTURE'}):
            resp = self.bulk({'sha256s': [b], 'decision': d})
            self.assertEqual(resp.status_code, 400, (d, resp.content[:200]))
        self.assertEqual(SlipReview.objects.get(sha256=b).decision, 'ALREADY_IN_XERO',
                         'an invalid decision must not clobber the stored one')

    def test_combined_actions_apply_every_field_to_every_target(self):
        targets = [self.NAMES['a_fy26'], self.NAMES['j_ts_null'], self.NAMES['i_ocr_null']]
        resp = self.bulk({'sha256s': targets, 'set_to_process': 'true', 'decision': 'personal',
                          'note': 'bulk note', 'comment': 'bulk comment'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 3, 'commented': 3, 'unknown': []})
        for s in targets:
            rv = SlipReview.objects.get(sha256=s)
            self.assertTrue(rv.to_process, s)
            self.assertEqual(rv.decision, 'PERSONAL', s)
            self.assertEqual(rv.note, 'bulk note', s)
            self.assertEqual(rv.updated_by, 'reviewer', s)
            c = SlipComment.objects.get(sha256=s)
            self.assertEqual(c.text, 'bulk comment', s)
            self.assertEqual(c.author, 'reviewer', s)
        self.assertEqual(SlipComment.objects.count(), 3)
        rows = {r['sha256']: r for r in self.list_(page_size=200)['results']}
        self.assertEqual(rows[targets[0]]['review']['decision'], 'PERSONAL')
        self.assertEqual(rows[targets[0]]['comment_count'], 1)

    def test_comment_only_call_creates_no_review_rows(self):
        targets = [self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']]
        resp = self.bulk({'sha256s': targets, 'comment': 'just a comment'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 0, 'comment-only must not count review upserts')
        self.assertEqual(resp.json()['commented'], 2)
        self.assertFalse(SlipReview.objects.exists(),
                         'a comment-only bulk call materialised SlipReview rows — that is a real bug')
        self.assertEqual(SlipComment.objects.count(), 2)

    def test_review_only_call_creates_no_comments(self):
        targets = [self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']]
        resp = self.bulk({'sha256s': targets, 'set_to_process': True, 'decision': 'CAPTURE', 'note': 'n'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 2)
        self.assertEqual(resp.json()['commented'], 0)
        self.assertFalse(SlipComment.objects.exists(), 'a review-only bulk call wrote comments')

    def test_updated_by_and_author_are_the_jwt_user_not_a_default(self):
        booker = User.objects.create_user(username='bookkeeper', email='bookkeeper@example.com',
                                          password='pw-irrelevant')
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {RefreshToken.for_user(booker).access_token}')
        a = self.NAMES['a_fy26']
        resp = self.bulk({'sha256s': [a], 'note': 'from the bookkeeper', 'comment': 'bk comment'}, client=client)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(SlipReview.objects.get(sha256=a).updated_by, 'bookkeeper',
                         'updated_by must be the calling JWT user, not a default or another user')
        self.assertEqual(SlipComment.objects.get(sha256=a).author, 'bookkeeper')

    def test_upsert_replaces_only_the_sent_fields_and_never_duplicates(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        SlipReview.objects.create(sha256=a, to_process=True, decision='MEAL_SKIP', note='keep me', updated_by='earlier')
        SlipReview.objects.create(sha256=b, to_process=False, decision='', note='keep me too', updated_by='earlier')
        resp = self.bulk({'sha256s': [a, b], 'decision': 'DUPLICATE'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 2)
        self.assertEqual(SlipReview.objects.count(), 2, 'bulk upsert duplicated review rows')
        ra, rb = SlipReview.objects.get(sha256=a), SlipReview.objects.get(sha256=b)
        self.assertEqual(ra.decision, 'DUPLICATE')
        self.assertEqual(rb.decision, 'DUPLICATE')
        self.assertEqual(ra.note, 'keep me', 'a field NOT in the request was clobbered by the upsert')
        self.assertEqual(rb.note, 'keep me too')
        self.assertTrue(ra.to_process, 'to_process was not in the request and must be untouched')
        self.assertFalse(rb.to_process)
        self.assertEqual(ra.updated_by, 'reviewer', 'updated_by must move to the caller on every upsert')

    def test_blank_comment_is_400_even_with_valid_siblings_and_writes_nothing(self):
        a = self.NAMES['a_fy26']
        for c in ('', '   ', '\n\t', None):
            resp = self.bulk({'sha256s': [a], 'note': 'valid', 'comment': c})
            self.assertEqual(resp.status_code, 400, (c, resp.content[:200]))
        self.assert_nothing_written()


# --------------------------------------------------------------------------- #
# 15. Bulk endpoint — unknown sha256s (the register is the source of truth)
# --------------------------------------------------------------------------- #
class BulkUnknownShaTests(BulkMixin, TestCase):
    def test_mixed_known_and_unknown_processes_known_reports_unknown_in_input_order(self):
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        u1, u2, u3 = sha(7001), sha(7002), sha(7003)
        resp = self.bulk({'sha256s': [u1, a, u2, b, u3], 'note': 'mixed', 'comment': 'mixed c'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['unknown'], [u1, u2, u3], 'unknown must mirror INPUT order')
        self.assertEqual(body['updated'], 2)
        self.assertEqual(body['commented'], 2)
        self.assertEqual(set(SlipReview.objects.values_list('sha256', flat=True)), {a, b},
                         'only register-backed sha256s may be written')
        self.assertEqual(set(SlipComment.objects.values_list('sha256', flat=True)), {a, b})
        self.assertEqual(SlipReview.objects.get(sha256=a).note, 'mixed')

    def test_every_sha_unknown_is_200_not_404(self):
        unknowns = [sha(7101), sha(7102)]
        resp = self.bulk({'sha256s': unknowns, 'set_to_process': True, 'decision': 'CAPTURE',
                          'note': 'n', 'comment': 'c'})
        self.assertEqual(resp.status_code, 200, f'all-unknown must be 200, never 404: {resp.content[:300]}')
        self.assertEqual(resp.json(), {'updated': 0, 'commented': 0, 'unknown': unknowns})
        self.assert_nothing_written()

    def test_review_row_orphaned_from_the_register_is_unknown(self):
        # A slip the sync deleted: SlipReview row exists, register row does not. The register is
        # the source of truth — the sha must be reported unknown and the orphan left untouched.
        ghost = sha(7201)
        SlipReview.objects.create(sha256=ghost, decision='CAPTURE', note='orphan', updated_by='sync-victim')
        resp = self.bulk({'sha256s': [ghost], 'note': 'resurrect?', 'comment': 'hello?'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json(), {'updated': 0, 'commented': 0, 'unknown': [ghost]},
                         'a SlipReview row alone must not make a sha known')
        orphan = SlipReview.objects.get(sha256=ghost)
        self.assertEqual(orphan.note, 'orphan', 'the orphaned review row was modified')
        self.assertEqual(orphan.updated_by, 'sync-victim')
        self.assertFalse(SlipComment.objects.exists())

    def test_hostile_unknown_values_are_reported_not_500(self):
        a = self.NAMES['a_fy26']
        hostiles = ["'; drop table whatsapp.klikk_slips; --", '%_%', 'ωμέγα-slip', 'z' * 300,
                    self.NAMES['j_ts_null'].upper(),  # case variant of a KNOWN sha — keys are case-sensitive
                    'Robert"); DROP TABLE receipts_slipreview;--']
        resp = self.bulk({'sha256s': hostiles + [a], 'note': 'hostile'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['unknown'], hostiles, 'hostile strings must round-trip as unknown, in order')
        self.assertEqual(body['updated'], 1)
        self.assertEqual(self.list_()['count'], 12, 'the register did not survive the hostile input')
        self.assertEqual(SlipReview.objects.count(), 1)

    def test_nul_byte_sha_entry_is_never_a_5xx(self):
        # JSON legally carries \x00; Postgres text cannot. The contract says unknown ids are
        # reported back with a 200 — a driver-level NUL explosion (500) from pure client input is
        # a bug (a 400 would also be acceptable). No DB assertions after the request on purpose.
        client = self.raw_client()
        resp = client.post(self.bulk_url(), {'sha256s': ['abc\x00def', self.NAMES['a_fy26']], 'note': 'nul'},
                           format='json')
        self.assertLess(resp.status_code, 500,
                        f'NUL byte in a sha256s entry blew up the endpoint: {resp.status_code} {resp.content[:300]}')

    def test_nul_byte_in_note_or_comment_is_never_a_5xx(self):
        client = self.raw_client()
        a = self.NAMES['a_fy26']
        resp = client.post(self.bulk_url(), {'sha256s': [a], 'note': 'x\x00y'}, format='json')
        self.assertLess(resp.status_code, 500, f'NUL in note: {resp.status_code} {resp.content[:300]}')
        resp = client.post(self.bulk_url(), {'sha256s': [a], 'comment': 'x\x00y'}, format='json')
        self.assertLess(resp.status_code, 500, f'NUL in comment: {resp.status_code} {resp.content[:300]}')


# --------------------------------------------------------------------------- #
# 16. Bulk endpoint — register isolation (never writes whatsapp.klikk_slips)
# --------------------------------------------------------------------------- #
class BulkRegisterIsolationTests(BulkMixin, TestCase):
    def test_bulk_never_writes_to_the_register(self):
        before = register_snapshot()
        self.assertEqual(len(before), 12)
        a, b = self.NAMES['a_fy26'], self.NAMES['b_fy26_auto']
        # a full write, a clear, an all-unknown call and a rejected call
        self.assertEqual(self.bulk({'sha256s': [a, b], 'set_to_process': True, 'decision': 'CAPTURE',
                                    'note': 'isolation', 'comment': 'isolation'}).status_code, 200)
        self.assertEqual(self.bulk({'sha256s': [a], 'note': '', 'decision': ''}).status_code, 200)
        self.assertEqual(self.bulk({'sha256s': [sha(7301)], 'note': 'ghost'}).status_code, 200)
        self.assertEqual(self.bulk({'sha256s': [a], 'decision': 'NOPE'}).status_code, 400)
        after = register_snapshot()
        self.assertEqual(before, after,
                         'bulk mutated whatsapp.klikk_slips — the register must be byte-identical')
        # belt-and-braces on the columns the recon depends on
        self.assertEqual([r[10] for r in after], [r[10] for r in before], 'synced_to_xero changed')
        self.assertEqual([r[7] for r in after], [r[7] for r in before], 'xero_status changed')
        self.assertEqual([r[6] for r in after], [r[6] for r in before], 'file_bytes changed')


# --------------------------------------------------------------------------- #
# 17. ids_only list mode (GET /audit/receipts/?ids_only=1)
# --------------------------------------------------------------------------- #
class IdsOnlyTests(ReceiptsFixtureMixin, TestCase):
    def ids(self, **params):
        resp = self.client.get(reverse('receipts:list'), {'ids_only': '1', **params})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()

    def all_row_shas(self, **params):
        """Page the ROW mode to exhaustion (deliberately tiny pages) and return the sha order."""
        out, page = [], 1
        while True:
            data = self.list_(page=page, page_size=3, **params)
            out += self.shas(data)
            if page >= data['num_pages']:
                return out
            page += 1

    def test_shape_count_and_no_row_data_leaks(self):
        resp = self.client.get(reverse('receipts:list'), {'ids_only': '1'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        data = resp.json()
        self.assertEqual(set(data), {'count', 'sha256s', 'truncated'}, 'ids mode must carry NOTHING else')
        self.assertEqual(data['count'], 12)
        self.assertFalse(data['truncated'])
        self.assertEqual(len(data['sha256s']), 12)
        self.assertEqual(set(data['sha256s']), set(self.NAMES.values()))
        text = resp.content.decode('utf-8', 'replace')
        for marker in ('results', 'view_url', 'file_bytes', 'ocr', 'journal', 'supplier', 'total',
                       'review', 'filename', '/audit/slip/', 'Builders', 'page_size'):
            self.assertNotIn(marker, text, f'ids_only leaked {marker!r}')
        self.assertFalse(body_has_blob(text), 'ids_only leaked file bytes')
        self.assertNotIn(slip_signature(self.NAMES['c_fy25_pend']), text,
                         'ids_only leaked a signed-viewer signature')

    def test_ids_only_requires_auth(self):
        resp = self.anon.get(reverse('receipts:list'), {'ids_only': '1'})
        self.assertEqual(resp.status_code, 401)
        text = resp.content.decode('utf-8', 'replace')
        for s in self.NAMES.values():
            self.assertNotIn(s, text, 'anonymous ids_only 401 leaked sha256s')

    def test_ids_only_matches_row_mode_for_every_filter(self):
        SlipReview.objects.create(sha256=self.NAMES['a_fy26'], to_process=True, decision='CAPTURE')
        SlipReview.objects.create(sha256=self.NAMES['d_fy26_nix'], to_process=True)
        SlipReview.objects.create(sha256=self.NAMES['f_ocr_abc'], decision='PERSONAL')
        filter_sets = [
            {}, {'status': 'MATCHED'}, {'status': 'PENDING'}, {'status': 'SKIPPED'},
            {'to_process': 'true'}, {'to_process': 'false'},
            {'decision': 'NONE'}, {'decision': 'UNDECIDED'}, {'decision': 'CAPTURE'}, {'decision': 'PERSONAL'},
            {'fy': 'FY26'}, {'fy': 'FY25'}, {'fy': 'not-a-year'},
            {'date_from': '2026-04-01', 'date_to': '2026-04-04'},
            {'date_from': '2025-07-01'}, {'date_to': '2025-12-31'},
            {'q': 'Builders'}, {'q': 'jonnys mozambican'},
            {'synced': 'true'}, {'category': 'Hardware'}, {'min_total': '100'}, {'max_total': '40'},
            {'fy': 'FY26', 'status': 'PENDING'}, {'q': 'Builders', 'status': 'SKIPPED'},
            {'to_process': 'true', 'decision': 'UNDECIDED'},
        ]
        for params in filter_sets:
            expected = self.all_row_shas(**params)
            data = self.ids(**params)
            self.assertEqual(data['sha256s'], expected,
                             f'ids_only disagrees with paging the row mode to exhaustion for {params!r}')
            self.assertEqual(data['count'], len(expected), f'count vs ids length for {params!r}')
            self.assertFalse(data['truncated'], params)

    def test_ids_only_honours_ordering(self):
        num_asc = [self.NAMES[n] for n in ('j_ts_null', 'c_fy25_pend', 'l_orphan_jn', 'd_fy26_nix',
                                           'a_fy26', 'k_fy27', 'b_fy26_auto')]
        self.assertEqual(self.ids(ordering='total')['sha256s'][:7], num_asc,
                         'ids_only ordering=total must sort by the numeric total ascending')
        self.assertEqual(self.ids(ordering='-total')['sha256s'][:7], list(reversed(num_asc)))
        for o in ('slip_ts', '-slip_ts', 'total', '-total', 'supplier', '-supplier',
                  'xero_status', '-xero_status'):
            row_order = self.shas(self.list_(ordering=o, page_size=200))
            self.assertEqual(self.ids(ordering=o)['sha256s'], row_order, f'ordering={o!r}')
        # junk ordering falls back to the default, same as the row mode
        self.assertEqual(self.ids(ordering='file_bytes; drop')['sha256s'],
                         self.shas(self.list_(page_size=200)))

    def test_ids_only_ignores_page_and_page_size(self):
        full = self.ids()['sha256s']
        self.assertEqual(len(full), 12)
        data = self.ids(page='99', page_size='1')
        self.assertEqual(data['sha256s'], full, 'ids_only must ignore page/page_size')
        self.assertEqual(data['count'], 12)
        self.assertEqual(self.ids(page='2', page_size='5')['sha256s'], full)

    def test_ids_only_falsy_values_fall_through_to_row_mode(self):
        for v in ('0', 'false', 'no', 'off', '', 'banana', 'ids_only'):
            resp = self.client.get(reverse('receipts:list'), {'ids_only': v})
            self.assertEqual(resp.status_code, 200, (v, resp.content[:200]))
            data = resp.json()
            self.assertIn('results', data, f'ids_only={v!r} must fall through to the ROW mode')
            self.assertIn('totals', data, v)
            self.assertNotIn('sha256s', data, v)
            self.assertNotIn('truncated', data, v)
        data = self.list_()  # absent entirely
        self.assertIn('results', data)
        self.assertNotIn('sha256s', data)

    def test_ids_only_truthy_variants_trigger_ids_mode(self):
        for v in ('1', 'true', 'TRUE', 'yes', 'on'):
            resp = self.client.get(reverse('receipts:list'), {'ids_only': v})
            self.assertEqual(resp.status_code, 200, v)
            self.assertEqual(set(resp.json()), {'count', 'sha256s', 'truncated'}, f'ids_only={v!r}')

    def test_ids_only_q_with_nul_byte_is_never_a_5xx(self):
        client = APIClient(raise_request_exception=False)
        client.credentials(**self.client._credentials)
        resp = client.get(reverse('receipts:list'), {'ids_only': '1', 'q': 'a\x00b'})
        self.assertLess(resp.status_code, 500, f'NUL in q: {resp.status_code} {resp.content[:300]}')


# --------------------------------------------------------------------------- #
# 18. ids_only truncation boundary (MAX_IDS)
# --------------------------------------------------------------------------- #
class IdsOnlyTruncationTests(ReceiptsFixtureMixin, TestCase):
    """
    views.py binds the name at import time (``from .services import MAX_IDS``), so the patch
    must target ``apps.receipts.views.MAX_IDS`` — patching services.MAX_IDS would be a no-op.
    """

    def ids(self, **params):
        resp = self.client.get(reverse('receipts:list'), {'ids_only': '1', **params})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()

    def test_truncated_false_when_the_filter_matches_exactly_the_cap(self):
        with mock.patch('apps.receipts.views.MAX_IDS', 3):
            data = self.ids(status='MATCHED')  # exactly 3 fixture rows
        self.assertEqual(data['count'], 3)
        self.assertEqual(len(data['sha256s']), 3)
        self.assertFalse(data['truncated'], 'exactly-at-cap must NOT be flagged truncated')
        self.assertEqual(set(data['sha256s']),
                         {self.NAMES['a_fy26'], self.NAMES['b_fy26_auto'], self.NAMES['l_orphan_jn']})

    def test_truncated_true_when_the_filter_exceeds_the_cap(self):
        row_order = self.shas(self.list_(fy='FY26', status='PENDING', page_size=200))
        self.assertEqual(len(row_order), 4)  # f, g, h, i
        with mock.patch('apps.receipts.views.MAX_IDS', 3):
            data = self.ids(fy='FY26', status='PENDING')
        self.assertTrue(data['truncated'], 'over-cap must be flagged truncated')
        self.assertEqual(data['count'], 4, 'count reports the FULL filter match, not the cap')
        self.assertEqual(data['sha256s'], row_order[:3], 'truncated ids must be the ordered prefix')


# --------------------------------------------------------------------------- #
# 19. page_size clamp — proven against MORE than MAX_PAGE_SIZE rows
# --------------------------------------------------------------------------- #
class PageSizeClampTests(TestCase):
    """
    page_size is clamped into 1..200, never rejected. The fixture holds 205 rows so
    'page_size=200 honoured' means an actual 200-row page, not a vacuous echo on 12 rows.
    """
    N = 205

    @classmethod
    def setUpTestData(cls):
        create_slips_table()
        base = ts(2025, 8, 1, 0, 0)
        for i in range(cls.N):
            insert_slip(sha(5000 + i), slip_ts=base + dt.timedelta(minutes=i), filename=f'page-{i:03d}.jpg')
        cls.user = User.objects.create_user(username='pager', email='pager@example.com', password='pw-irrelevant')

    def setUp(self):
        self.client = APIClient()
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {RefreshToken.for_user(self.user).access_token}')

    def list_(self, **params):
        resp = self.client.get(reverse('receipts:list'), params)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        return resp.json()

    def test_contract_constants(self):
        # If any of these moves, the console contract moved with it — scream, don't drift.
        from . import services, views
        self.assertEqual(views.BULK_MAX, 500)
        self.assertEqual(views.MAX_IDS, 2000)
        self.assertEqual(services.MAX_PAGE_SIZE, 200)
        self.assertEqual(services.DEFAULT_PAGE_SIZE, 50)

    def test_page_size_200_is_honoured_with_a_real_200_row_page(self):
        data = self.list_(page_size=200)
        self.assertEqual(data['page_size'], 200)
        self.assertEqual(len(data['results']), 200, 'page_size=200 must return an actual 200-row page')
        self.assertEqual(data['count'], self.N)
        self.assertEqual(data['num_pages'], 2)
        page2 = self.list_(page_size=200, page=2)
        self.assertEqual(len(page2['results']), 5)
        all_shas = {r['sha256'] for r in data['results']} | {r['sha256'] for r in page2['results']}
        self.assertEqual(len(all_shas), self.N, 'the two pages must partition the set without overlap/loss')

    def test_oversize_page_size_is_clamped_to_200_not_rejected(self):
        for oversize in (201, 500, 2000, 10 ** 9):
            data = self.list_(page_size=oversize)
            self.assertEqual(data['page_size'], 200, f'page_size={oversize} must clamp to 200, not reject')
            self.assertEqual(len(data['results']), 200, oversize)
            self.assertEqual(data['num_pages'], 2, oversize)

    def test_zero_negative_and_junk_page_size(self):
        data = self.list_(page_size=0)
        self.assertEqual(data['page_size'], 1, 'page_size=0 must clamp to 1')
        self.assertEqual(len(data['results']), 1)
        self.assertEqual(data['num_pages'], self.N)
        data = self.list_(page_size=-5)
        self.assertEqual(data['page_size'], 1, 'page_size=-5 must clamp to 1')
        self.assertEqual(len(data['results']), 1)
        for junk in ('abc', '', '200.5', '2e2', None):
            params = {} if junk is None else {'page_size': junk}
            data = self.list_(**params)
            self.assertEqual(data['page_size'], 50, f'page_size={junk!r} must fall back to the default 50')
            self.assertEqual(len(data['results']), 50, junk)
            self.assertEqual(data['num_pages'], 5, junk)

    def test_ids_only_at_scale_ignores_pagination_and_keeps_order(self):
        expected = [sha(5000 + i) for i in reversed(range(self.N))]  # default ordering is -slip_ts
        resp = self.client.get(reverse('receipts:list'), {'ids_only': '1', 'page': '99', 'page_size': '1'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        data = resp.json()
        self.assertEqual(data['count'], self.N)
        self.assertFalse(data['truncated'])
        self.assertEqual(data['sha256s'], expected,
                         'ids_only must return ALL matching ids in order, ignoring page/page_size')
