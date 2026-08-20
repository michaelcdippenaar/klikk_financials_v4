"""
Adversarial tests for the Audit Findings register API (apps.audit.findings_*).

Written against CONTRACT.md (senior-dev, 2026-08-20, incl. the FY RESOLUTION
amendment) — deliberately NOT against the implementation, which was not read
before these tests were authored. Where the contract and the implementation
disagree, the FAILING TEST is the deliverable, not a softened assertion.

Date-robustness: the register's read default is "the most recent FY that has
findings" while the write default is the strict current FY. Rather than
hard-coding 2026/2027 (which would rot on 1 Jul 2027), the fixture seeds
CUR-1 / CUR-2 where CUR is independently derived from today's date with the
1-July rule — under today's clock (Aug 2026) that IS 2026/2027, so the
contract's concrete examples are exercised verbatim. The 30 Jun / 1 Jul
boundary itself is pinned with a frozen clock, never the real one.

Run:  manage.py test apps.audit
"""
from __future__ import annotations

import contextlib
import csv
import datetime as dt
import hashlib
import hmac
import io
import threading
import types
from decimal import Decimal
from unittest import mock
from urllib.parse import parse_qs, urlparse

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connections
from django.test import TestCase, TransactionTestCase, override_settings
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from apps.audit import services as audit_services
from apps.audit.findings_services import (
    BULK_MAX,
    DEFAULT_PAGE_SIZE,
    MAX_PAGE_SIZE,
    SEVERITIES,
    STATUSES,
    current_fy,
)

# The contract (FY RESOLUTION — AMENDED 2026-08-20) mandates this export. Importing it
# tolerantly here so ONE missing name cannot kill the whole module's collection — the
# tests that depend on it fail loudly instead.
try:
    from apps.audit.findings_services import resolve_default_fy
except ImportError:  # pragma: no cover — this is itself a contract violation
    resolve_default_fy = None
from apps.audit.models import AuditFinding, AuditFindingComment

User = get_user_model()

LIST_URL = '/audit/findings/'
SUMMARY_URL = '/audit/findings/summary/'
EXPORT_URL = '/audit/findings/export/'
BULK_URL = '/audit/findings/bulk/'


def detail_url(pk):
    return f'/audit/findings/{pk}/'


def comments_url(pk):
    return f'/audit/findings/{pk}/comments/'


def attachments_url(pk):
    return f'/audit/findings/{pk}/attachments/'


def _fy_for(d: dt.date) -> int:
    """Independent re-derivation of the contract's FY rule (1 Jul start, named by ending year)."""
    return d.year + 1 if d.month >= 7 else d.year


CUR = _fy_for(dt.date.today())    # 2027 when run in Aug 2026
FY_A = CUR - 1                    # the "just closed" FY the register holds (2026 today)
FY_OLD = CUR - 2                  # a prior FY (2025 today)


def ref_for(fy: int, seq: int) -> str:
    return f'FY{fy % 100:02d}-{seq:03d}'


EXPECTED_EXPORT_HEADER = [
    'ref', 'fy', 'title', 'severity', 'status', 'category', 'amount', 'currency',
    'owner', 'due_date', 'source', 'check_code', 'asana_gid', 'description',
    'evidence', 'comments', 'created_by', 'created_at', 'updated_at',
]

FINDING_DICT_KEYS = {
    'id', 'fy', 'ref', 'title', 'severity', 'status', 'category', 'amount',
    'currency', 'description', 'evidence', 'owner', 'due_date', 'source',
    'check_code', 'asana_gid', 'created_by', 'updated_by', 'created_at',
    'updated_at', 'comment_count', 'attachment_count',
}

# A production-shaped CSV-injection title: leading '=', comma, double-quote, newline.
INJECTION_TITLE = '=HYPERLINK("http://evil.example/x"), "Aurras R138,000"\npaid with no invoice'


def attachment_sig(pk: int) -> str:
    """The contract's signature formula, computed independently of the implementation."""
    return hmac.new(settings.SECRET_KEY.encode(), f'auditfinding:{pk}'.encode(),
                    hashlib.sha256).hexdigest()[:32]


def _frozen_date(y, m, d):
    class Frozen(dt.date):
        @classmethod
        def today(cls):
            return dt.date(y, m, d)
    return Frozen


@contextlib.contextmanager
def clock_at(y, m, d):
    """Freeze apps.audit.services' view of 'today' without touching the real datetime module."""
    ns = types.SimpleNamespace(
        date=_frozen_date(y, m, d), timedelta=dt.timedelta, datetime=dt.datetime,
        time=dt.time, timezone=dt.timezone,
    )
    with mock.patch.object(audit_services, 'dt', ns):
        yield


def mk_finding(fy, seq, title, *, severity='MEDIUM', status='OPEN', category='OTHER',
               amount=None, owner='', source='internal-audit run 13', check_code='',
               asana_gid='', description='', evidence=None, due_date=None,
               currency='ZAR', created_by='seed'):
    return AuditFinding.objects.create(
        fy=fy, ref=ref_for(fy, seq), title=title, severity=severity, status=status,
        category=category, amount=None if amount is None else Decimal(amount),
        currency=currency, description=description, evidence=evidence or [],
        owner=owner, due_date=due_date, source=source, check_code=check_code,
        asana_gid=asana_gid, created_by=created_by,
    )


class AuthedTestBase(TestCase):
    maxDiff = None

    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user(
            username='auditor-mc', email='mc@tremly.com', password='pw-irrelevant')

    def setUp(self):
        self.client = APIClient()
        self.auth()
        self.anon = APIClient()
        self.tolerant = APIClient(raise_request_exception=False)
        self.tolerant.credentials(**self.client._credentials)

    def auth(self, user=None):
        token = str(RefreshToken.for_user(user or self.user).access_token)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

    # -- helpers -------------------------------------------------------------
    def list_(self, **params):
        resp = self.client.get(LIST_URL, params)
        self.assertEqual(resp.status_code, 200, resp.content[:500])
        return resp.json()

    def refs(self, data):
        return [r['ref'] for r in data['results']]

    def create(self, **over):
        body = {
            'title': 'Payments made before supplier bill captured',
            'severity': 'HIGH',
            'category': 'SUP',
            'source': 'internal-audit run 13',
            'description': 'Cash left the bank before a supplier bill existed.',
        }
        body.update(over)
        body = {k: v for k, v in body.items() if v is not ...}
        return self.client.post(LIST_URL, body, format='json')


class FindingsFixtureMixin(AuthedTestBase):
    """
    11 production-shaped findings: 10 in FY_A (one of every severity and status,
    NULL amounts, evidence-only search bait, a CSV-injection title) + 1 in FY_OLD.
    Refs are pre-assigned exactly as the allocator would have (FY26-001..010).
    """

    # (seq, severity, status, category, amount, owner)
    FY_A_ROWS = {
        1: ('CRITICAL', 'OPEN', 'SUP', '429110.39', 'bookkeeper'),
        2: ('HIGH', 'IN_PROGRESS', 'PRC', None, 'MC'),
        3: ('HIGH', 'OPEN', 'VAT', '45644.00', 'accountant'),
        4: ('MEDIUM', 'OPEN', 'BNK', '458498.00', 'bookkeeper'),
        5: ('MEDIUM', 'RESOLVED', 'DOC', '579160.00', 'bookkeeper'),
        6: ('LOW', 'ACCEPTED', 'SUP', '3166.42', 'bookkeeper'),
        7: ('INFO', 'WITHDRAWN', 'OTHER', None, ''),
        8: ('HIGH', 'OPEN', 'PAYROLL', '99810.00', 'accountant'),
        9: ('HIGH', 'OPEN', 'PAYROLL', '17420.01', 'MC'),
        10: ('HIGH', 'OPEN', 'SUP', '138000.00', 'MC'),
    }
    FY_A_AMOUNT_SUM = sum(Decimal(a) for (_s, _st, _c, a, _o) in FY_A_ROWS.values() if a)

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.f1 = mk_finding(
            FY_A, 1, 'Payments made before supplier bill captured',
            severity='CRITICAL', category='SUP', amount='429110.39', owner='bookkeeper',
            asana_gid='1217633700114593', due_date=dt.date(FY_A, 9, 30),
            description='Payments left the bank account before any bill was captured.',
            evidence=[{'type': 'journal', 'ref': 'JN-962018',
                       'note': 'payment posted 41 days before the bill'}])
        cls.f2 = mk_finding(
            FY_A, 2, 'SARS banking-details verification — Lia Dippenaar (related party)',
            severity='HIGH', status='IN_PROGRESS', category='PRC', owner='MC',
            asana_gid='1217628653069456',
            description='SARS eFiling flagged the banking-details change for verification.')
        cls.f3 = mk_finding(
            FY_A, 3, 'Input VAT recoverable ~R45,644 (bills captured VAT-inclusive, no split)',
            severity='HIGH', category='VAT', amount='45644.00', owner='accountant',
            source='input-VAT audit 2026-08-19', due_date=dt.date(FY_A, 8, 31),
            description='Bills were captured gross; the input VAT leg was never split out.')
        cls.f4 = mk_finding(
            FY_A, 4, 'R458,498 coded to loan accounts though cash left a Klikk bank account',
            severity='MEDIUM', category='BNK', amount='458498.00', owner='bookkeeper',
            check_code='BNK-05', due_date=dt.date(FY_A, 11, 30),
            description='BNK-05 MISPOSTED rows: loan-account coding with a real bank outflow.')
        cls.f5 = mk_finding(
            FY_A, 5, 'R579k FY spend with no attachment and no slip',
            severity='MEDIUM', status='RESOLVED', category='DOC', amount='579160.00',
            owner='bookkeeper', check_code='DOC-03',
            description='DOC-03: spend without any supporting document in either channel.')
        cls.f6 = mk_finding(
            FY_A, 6, 'Higgsfield R3,166.42 charged twice on card',
            severity='LOW', status='ACCEPTED', category='SUP', amount='3166.42',
            owner='bookkeeper', description='Duplicate card charge; refund pursued.')
        cls.f7 = mk_finding(
            FY_A, 7, 'Card statement cut-over documented for the record',
            severity='INFO', status='WITHDRAWN', category='OTHER',
            description='Informational only; withdrawn after review.')
        cls.f8 = mk_finding(
            FY_A, 8, 'Wandeli R99,810 paid as supplier — no payroll/UIF/SDL',
            severity='HIGH', category='PAYROLL', amount='99810.00', owner='accountant',
            source='codex worker audit 2026-08-20',
            description='Worker-vs-contractor classification exposure on recurring payments.')
        cls.f9 = mk_finding(
            FY_A, 9, 'Dayworker cash routing without payee detail',
            severity='HIGH', category='PAYROLL', amount='17420.01', owner='MC',
            source='codex worker audit 2026-08-20',
            description='Amounts routed through an intermediary with no payee breakdown.',
            evidence=[{'type': 'bank', 'ref': 'INVSTC-88421',
                       'note': 'Investec batch payment, no per-payee split'}])
        cls.f10 = mk_finding(
            FY_A, 10, INJECTION_TITLE,
            severity='HIGH', category='SUP', amount='138000.00', owner='MC',
            asana_gid='1217591235585694',
            description='Aurras paid with no invoice; title deliberately CSV-hostile.')
        cls.f_old = mk_finding(
            FY_OLD, 1, 'Aurras paid with no invoice (prior year)',
            severity='MEDIUM', category='SUP', amount='138000.00', owner='MC',
            description='Prior-year variant of the missing-invoice finding.')


class ExplicitFyMixin(FindingsFixtureMixin):
    """
    Pin ?fy=FY_A on fixture-dependent reads so every test below measures ITS OWN
    mechanism. The read-default rule (resolve_default_fy) is deliberately owned by
    FyHelperTests / FyResolution*Tests alone — without this pin, a read-default
    defect (which exists at the time of writing) cascades into 20+ unrelated tests.
    """

    def list_(self, **params):
        params.setdefault('fy', str(FY_A))
        return super().list_(**params)


# --------------------------------------------------------------------------- #
# 1. FY helpers + resolution (frozen clock, both resolve branches)
# --------------------------------------------------------------------------- #
class FyHelperTests(TestCase):
    def test_fy_bounds_2026(self):
        self.assertEqual(audit_services.fy_bounds(2026),
                         (dt.date(2025, 7, 1), dt.date(2026, 6, 30)))
        self.assertEqual(audit_services.fy_bounds(2027),
                         (dt.date(2026, 7, 1), dt.date(2027, 6, 30)))

    def test_current_fy_at_the_jun30_jul1_boundary(self):
        for (y, m, d), want in (((2026, 6, 30), 2026), ((2026, 7, 1), 2027),
                                ((2026, 12, 31), 2027), ((2027, 1, 1), 2027)):
            with clock_at(y, m, d):
                self.assertEqual(current_fy(), want, f'{y}-{m:02d}-{d:02d}')
                self.assertEqual(audit_services.fiscal_year_for_today(), want)

    def test_current_fy_matches_independent_rule_on_the_real_clock(self):
        self.assertEqual(current_fy(), _fy_for(dt.date.today()))

    def test_resolve_default_fy_is_exported(self):
        self.assertIsNotNone(
            resolve_default_fy,
            'CONTRACT VIOLATION: findings_services must export resolve_default_fy() '
            '(FY RESOLUTION amendment, 2026-08-20)')

    def test_resolve_default_fy_empty_register_is_current_fy(self):
        self.assertIsNotNone(resolve_default_fy, 'resolve_default_fy export missing')
        self.assertFalse(AuditFinding.objects.exists())
        self.assertEqual(resolve_default_fy(), current_fy())
        with clock_at(2026, 6, 30):
            self.assertEqual(resolve_default_fy(), 2026)
        with clock_at(2026, 7, 1):
            self.assertEqual(resolve_default_fy(), 2027)

    def test_resolve_default_fy_prefers_max_fy_with_findings(self):
        self.assertIsNotNone(resolve_default_fy, 'resolve_default_fy export missing')
        mk_finding(2026, 1, 'FY2026 row')
        self.assertGreaterEqual(current_fy(), 2027, 'test env sanity: today is after 30 Jun 2026')
        self.assertEqual(resolve_default_fy(), 2026,
                         'read default must be the latest FY WITH findings, not the current FY')
        mk_finding(2025, 1, 'FY2025 row')
        self.assertEqual(resolve_default_fy(), 2026)


class FyResolutionEmptyRegisterApiTests(AuthedTestBase):
    def test_list_defaults_to_current_fy_when_register_is_empty(self):
        data = self.list_()
        self.assertEqual(data['fy'], CUR)
        self.assertEqual(data['current_fy'], CUR)
        self.assertEqual(data['count'], 0)
        self.assertEqual(data['totals'], {'count': 0, 'amount': '0.00'})

    def test_summary_defaults_to_current_fy_and_offers_it_even_with_no_rows(self):
        resp = self.client.get(SUMMARY_URL)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['fy'], CUR)
        self.assertEqual(body['current_fy'], CUR)
        self.assertIn(CUR, body['fy_options'])
        self.assertEqual(body['count'], 0)


class FyResolutionApiTests(FindingsFixtureMixin):
    def test_list_default_is_latest_fy_with_findings_and_echoes_both_fys(self):
        data = self.list_()
        self.assertEqual(data['fy'], FY_A, 'default must resolve to the FY that has findings')
        self.assertEqual(data['current_fy'], CUR)
        self.assertEqual(data['count'], 10)
        self.assertNotIn(ref_for(FY_OLD, 1), self.refs(data))

    def test_fy_all_spans_every_fy_case_insensitively(self):
        for val in ('all', 'ALL', 'All'):
            data = self.list_(fy=val)
            self.assertEqual(data['count'], 11, val)
            self.assertIsNone(data['fy'], 'fy key must be null when the FY filter is removed')

    def test_explicit_fy_overrides(self):
        data = self.list_(fy=str(FY_OLD))
        self.assertEqual(data['count'], 1)
        self.assertEqual(self.refs(data), [ref_for(FY_OLD, 1)])
        self.assertEqual(data['fy'], FY_OLD)

    def test_bad_fy_values_are_400(self):
        # '' (present but empty) is not asserted: empty-vs-absent is contract-ambiguous.
        for bad in ('notanumber', '1999', '3000', '20.26', 'FY26'):
            resp = self.client.get(LIST_URL, {'fy': bad})
            self.assertEqual(resp.status_code, 400, f'fy={bad!r} -> {resp.status_code}')
            self.assertIn('detail', resp.json())

    def test_export_default_fy_matches_the_read_default(self):
        resp = self.client.get(EXPORT_URL)
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        rows = list(csv.reader(io.StringIO(resp.content.decode('utf-8'))))
        self.assertEqual(len(rows) - 1, 10,
                         'the export must default to the latest FY WITH findings, like the list')

    def test_post_without_fy_lands_in_the_current_fy_not_the_read_default(self):
        # THE ASYMMETRY (misfile guard): reads default to FY_A because it has rows,
        # but a bare POST must land in the strict current FY.
        resp = self.create(title='A brand-new current-year finding')
        self.assertEqual(resp.status_code, 201, resp.content[:500])
        body = resp.json()
        self.assertEqual(body['fy'], CUR,
                         'bare POST misfiled into the read-default FY instead of the current FY')
        self.assertEqual(body['ref'], ref_for(CUR, 1))
        # And the read default now moves to CUR, since CUR now has a finding.
        self.assertEqual(self.list_()['fy'], CUR)

    def test_post_with_explicit_fy_honours_it(self):
        resp = self.create(title='Explicit prior-year finding', fy=FY_OLD)
        self.assertEqual(resp.status_code, 201, resp.content[:500])
        self.assertEqual(resp.json()['fy'], FY_OLD)
        self.assertEqual(resp.json()['ref'], ref_for(FY_OLD, 2))

    def test_summary_default_matches_list_default(self):
        body = self.client.get(SUMMARY_URL).json()
        self.assertEqual(body['fy'], FY_A)
        self.assertEqual(body['current_fy'], CUR)


# --------------------------------------------------------------------------- #
# 2. Ref allocation
# --------------------------------------------------------------------------- #
class RefAllocationTests(AuthedTestBase):
    def test_sequential_refs_per_fy(self):
        for i in range(1, 11):
            resp = self.create(title=f'Finding number {i}', fy=2026)
            self.assertEqual(resp.status_code, 201, resp.content[:300])
            self.assertEqual(resp.json()['ref'], f'FY26-{i:03d}')
        # A different FY starts its own sequence.
        resp = self.create(title='First of the next year', fy=2027)
        self.assertEqual(resp.json()['ref'], 'FY27-001')
        resp = self.create(title='Second of the next year', fy=2027)
        self.assertEqual(resp.json()['ref'], 'FY27-002')

    def test_client_supplied_ref_is_ignored(self):
        resp = self.create(title='Ref smuggling attempt', fy=2026, ref='FY26-999')
        self.assertEqual(resp.status_code, 201, resp.content[:300])
        self.assertEqual(resp.json()['ref'], 'FY26-001',
                         'a client-supplied ref must be ignored, never honoured')

    def test_withdrawn_and_deleted_refs_are_never_reused(self):
        ids = []
        for i in (1, 2, 3):
            resp = self.create(title=f'Finding {i}', fy=2026)
            self.assertEqual(resp.status_code, 201)
            ids.append(resp.json()['id'])
        patch = self.client.patch(detail_url(ids[2]), {'status': 'WITHDRAWN'}, format='json')
        self.assertEqual(patch.status_code, 200, patch.content[:300])
        AuditFinding.objects.filter(pk=ids[1]).delete()   # hard-delete FY26-002 via the ORM
        resp = self.create(title='Finding 4 after churn', fy=2026)
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.json()['ref'], 'FY26-004',
                         'refs must be monotonic — FY26-002/003 must never be reused or renumbered')

    def test_unique_constraint_is_declared(self):
        names = {c.name for c in AuditFinding._meta.constraints}
        self.assertIn('uniq_auditfinding_ref_per_fy', names)


class RefConcurrencyTests(TransactionTestCase):
    """Real threads against the create endpoint: the advisory-lock allocator must serialise."""

    def test_parallel_creates_never_share_a_ref(self):
        user = User.objects.create_user(username='racer', email='r@example.com', password='x')
        token = str(RefreshToken.for_user(user).access_token)
        n = 6
        barrier = threading.Barrier(n, timeout=15)
        results, errors = [], []
        lock = threading.Lock()

        def worker(i):
            try:
                client = APIClient()
                client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
                barrier.wait()
                resp = client.post(LIST_URL, {
                    'title': f'Concurrent finding {i}', 'severity': 'MEDIUM',
                    'category': 'SUP', 'source': 'concurrency test', 'fy': 2026,
                }, format='json')
                with lock:
                    results.append((resp.status_code,
                                    resp.json() if resp.status_code == 201 else resp.content[:300]))
            except Exception as exc:  # noqa: BLE001 — surface to the assertion
                with lock:
                    errors.append(repr(exc))
            finally:
                connections.close_all()

        threads = [threading.Thread(target=worker, args=(i,)) for i in range(n)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=60)
        self.assertEqual(errors, [])
        self.assertEqual([s for s, _ in results], [201] * n,
                         f'every concurrent create must succeed: {results}')
        refs = [b['ref'] for _, b in results]
        self.assertEqual(len(set(refs)), n, f'duplicate refs allocated under concurrency: {refs}')
        self.assertEqual(sorted(refs), [f'FY26-{i:03d}' for i in range(1, n + 1)])
        self.assertEqual(AuditFinding.objects.filter(fy=2026).values('ref').distinct().count(), n)


# --------------------------------------------------------------------------- #
# 3. Filters
# --------------------------------------------------------------------------- #
class FilterTests(ExplicitFyMixin):
    def test_status_filter_comma_separated_case_insensitive(self):
        self.assertEqual(self.list_(status='open')['count'], 6)
        self.assertEqual({r['ref'] for r in self.list_(status='resolved,ACCEPTED')['results']},
                         {self.f5.ref, self.f6.ref})
        self.assertEqual(self.list_(status='In_Progress')['count'], 1)

    def test_severity_filter_comma_separated_case_insensitive(self):
        self.assertEqual(self.list_(severity='critical')['count'], 1)
        self.assertEqual(self.list_(severity='HIGH,critical')['count'], 6)
        self.assertEqual(self.list_(severity='info,LOW')['count'], 2)

    def test_bad_vocabulary_is_400(self):
        for params in ({'severity': 'NONSENSE'}, {'severity': 'HIGH,NONSENSE'},
                       {'status': 'nope'}, {'status': 'OPEN,nope'}):
            resp = self.client.get(LIST_URL, params)
            self.assertEqual(resp.status_code, 400, f'{params} -> {resp.status_code}')
            self.assertIn('detail', resp.json())

    def test_category_filter_exact_case_insensitive_comma_separated(self):
        self.assertEqual({r['ref'] for r in self.list_(category='payroll,vat')['results']},
                         {self.f3.ref, self.f8.ref, self.f9.ref})
        self.assertEqual(self.list_(category='sup')['count'], 3)
        self.assertEqual(self.list_(category='SU')['count'], 0, 'category is exact, not a prefix')

    def test_owner_icontains(self):
        self.assertEqual(self.list_(owner='keeper')['count'], 4)
        self.assertEqual(self.list_(owner='BOOKKEEPER')['count'], 4)
        self.assertEqual(self.list_(owner='mc')['count'], 3)

    def test_check_code_exact_uppercased(self):
        self.assertEqual([r['ref'] for r in self.list_(check_code='bnk-05')['results']],
                         [self.f4.ref])
        self.assertEqual([r['ref'] for r in self.list_(check_code='DOC-03')['results']],
                         [self.f5.ref])
        self.assertEqual(self.list_(check_code='BNK')['count'], 0)

    def test_q_hits_title_description_ref_owner_source_and_evidence(self):
        self.assertEqual([r['ref'] for r in self.list_(q='Higgsfield')['results']],
                         [self.f6.ref], 'q over title')
        self.assertEqual([r['ref'] for r in self.list_(q='eFiling')['results']],
                         [self.f2.ref], 'q over description')
        self.assertEqual([r['ref'] for r in self.list_(q=self.f4.ref)['results']],
                         [self.f4.ref], 'q over ref')
        self.assertEqual(self.list_(q='bookkeeper')['count'], 4, 'q over owner')
        self.assertEqual([r['ref'] for r in self.list_(q='input-VAT audit')['results']],
                         [self.f3.ref], 'q over source')
        # The ONLY place '88421' exists is inside f9's evidence JSON.
        self.assertEqual([r['ref'] for r in self.list_(q='88421')['results']],
                         [self.f9.ref], 'q must search the evidence JSON cast to text')

    def test_q_junk_and_sqli_do_not_500(self):
        for junk in ("'; DROP TABLE audit_auditfinding; --", '%', '_', '\\', "' or 1=1 --", '((('):
            data = self.list_(q=junk)
            self.assertIn('count', data)
        self.assertEqual(self.list_()['count'], 10, 'the register survived the junk q values')

    def test_no_match_totals_are_zero_strings(self):
        data = self.list_(q='zzz-nothing-matches-this')
        self.assertEqual(data['count'], 0)
        self.assertEqual(data['totals'], {'count': 0, 'amount': '0.00'})

    def test_amount_bounds_inclusive_and_null_amounts_drop_out(self):
        self.assertEqual({r['ref'] for r in self.list_(amount_min='100000')['results']},
                         {self.f1.ref, self.f4.ref, self.f5.ref, self.f10.ref})
        self.assertEqual({r['ref'] for r in self.list_(amount_max='50000')['results']},
                         {self.f3.ref, self.f6.ref, self.f9.ref})
        self.assertEqual({r['ref'] for r in self.list_(amount_min='45644.00',
                                                       amount_max='45644.00')['results']},
                         {self.f3.ref}, 'bounds are inclusive')
        # NULL-amount rows (f2, f7) must be excluded the moment either bound appears.
        data = self.list_(amount_min='0')
        self.assertEqual(data['count'], 8)
        got = {r['ref'] for r in data['results']}
        self.assertNotIn(self.f2.ref, got)
        self.assertNotIn(self.f7.ref, got)

    def test_non_numeric_amount_bounds_are_400(self):
        # NOTE: '' (param present but empty) is deliberately not asserted — the contract
        # doesn't distinguish empty-present from absent; both readings are defensible.
        for field in ('amount_min', 'amount_max'):
            for bad in ('abc', '1;drop'):
                resp = self.client.get(LIST_URL, {field: bad})
                self.assertEqual(resp.status_code, 400, f'{field}={bad!r} -> {resp.status_code}')

    def test_exotic_decimal_bounds_never_500(self):
        # Decimal() parses NaN/Infinity; Postgres NUMERIC comparison must not blow up.
        for weird in ('NaN', 'Infinity', '-Infinity', '1e400', 'sNaN'):
            resp = self.tolerant.get(LIST_URL, {'amount_min': weird})
            self.assertLess(resp.status_code, 500, f'amount_min={weird!r} -> {resp.status_code}')

    def test_filters_combine_with_the_fy_filter(self):
        # f_old is MEDIUM/SUP in FY_OLD; an explicit fy must exclude it, fy=all include it.
        self.assertNotIn(ref_for(FY_OLD, 1),
                         {r['ref'] for r in self.list_(category='SUP')['results']})
        self.assertIn(ref_for(FY_OLD, 1),
                      {r['ref'] for r in self.list_(category='SUP', fy='all')['results']})


# --------------------------------------------------------------------------- #
# 4. Ordering + pagination
# --------------------------------------------------------------------------- #
class OrderingPaginationTests(ExplicitFyMixin):
    SEV_RANK = {s: i for i, s in enumerate(SEVERITIES)}

    def test_default_ordering_is_minus_fy_then_ref(self):
        data = self.list_(fy='all', page_size=200)
        want = [ref_for(FY_A, i) for i in range(1, 11)] + [ref_for(FY_OLD, 1)]
        self.assertEqual(self.refs(data), want)

    def test_severity_orders_by_rank_not_alphabet(self):
        sevs = [r['severity'] for r in self.list_(ordering='severity', page_size=200)['results']]
        self.assertEqual(sevs, sorted(sevs, key=self.SEV_RANK.__getitem__),
                         'severity must sort by rank CRITICAL->INFO')
        self.assertEqual(sevs[0], 'CRITICAL')
        self.assertEqual(sevs[-1], 'INFO')
        self.assertNotEqual(sevs, sorted(sevs),
                            'this is the alphabetical ordering — severity must use rank, '
                            'which puts INFO/LOW after MEDIUM')
        rev = [r['severity'] for r in self.list_(ordering='-severity', page_size=200)['results']]
        self.assertEqual(rev[0], 'INFO')
        self.assertEqual(rev[-1], 'CRITICAL')

    def test_amount_ascending_puts_nulls_last(self):
        amounts = [r['amount'] for r in self.list_(ordering='amount', page_size=200)['results']]
        self.assertEqual(amounts[:8], ['3166.42', '17420.01', '45644.00', '99810.00',
                                       '138000.00', '429110.39', '458498.00', '579160.00'])
        self.assertEqual(amounts[8:], [None, None], 'NULL amounts must sort last ascending')

    def test_due_date_ascending_puts_nulls_last(self):
        rows = self.list_(ordering='due_date', page_size=200)['results']
        dues = [r['due_date'] for r in rows]
        self.assertEqual(dues[:3], [f'{FY_A}-08-31', f'{FY_A}-09-30', f'{FY_A}-11-30'])
        self.assertTrue(all(d is None for d in dues[3:]), 'NULL due_dates must sort last ascending')

    def test_unknown_ordering_falls_back_silently(self):
        default = self.refs(self.list_(page_size=200))
        for bad in ('file_bytes', 'ref; DROP TABLE x', '__proto__', 'evidence', 'amount desc', ' '):
            self.assertEqual(self.refs(self.list_(ordering=bad, page_size=200)), default,
                             f'ordering={bad!r} must silently fall back to the default')

    def test_page_size_clamps_never_400(self):
        self.assertEqual(self.list_(page_size=999)['page_size'], MAX_PAGE_SIZE)
        for raw in ('0', '-1'):
            data = self.list_(page_size=raw)
            self.assertGreaterEqual(data['page_size'], 1, f'page_size={raw}')
            self.assertLessEqual(data['page_size'], DEFAULT_PAGE_SIZE, f'page_size={raw}')
        self.assertEqual(self.list_(page_size='abc')['page_size'], DEFAULT_PAGE_SIZE)

    def test_page_clamps_never_400_or_500(self):
        self.assertEqual(self.list_(page='-3')['page'], 1)
        self.assertEqual(self.list_(page='0')['page'], 1)
        self.assertEqual(self.list_(page='abc')['page'], 1)
        far = self.list_(page='9999', page_size=3)
        self.assertEqual(far['count'], 10)
        self.assertIsInstance(far['results'], list)

    def test_pages_partition_without_overlap_or_loss(self):
        seen = []
        p = 1
        while True:
            data = self.list_(fy='all', page=p, page_size=4, ordering='ref')
            seen += self.refs(data)
            if p >= data['num_pages']:
                break
            p += 1
        self.assertEqual(len(seen), 11)
        self.assertEqual(len(set(seen)), 11)

    def test_totals_cover_the_whole_filter_not_the_page(self):
        data = self.list_(page_size=3, page=2)
        self.assertEqual(data['count'], 10)
        self.assertEqual(data['totals']['count'], 10)
        self.assertEqual(Decimal(data['totals']['amount']), self.FY_A_AMOUNT_SUM)


# --------------------------------------------------------------------------- #
# 5. Summary maths
# --------------------------------------------------------------------------- #
class SummaryTests(ExplicitFyMixin):
    def summary(self, **params):
        params.setdefault('fy', str(FY_A))
        resp = self.client.get(SUMMARY_URL, params)
        self.assertEqual(resp.status_code, 200, resp.content[:500])
        return resp.json()

    def test_count_and_amount_match_the_list_totals_under_the_same_filters(self):
        for params in ({}, {'severity': 'HIGH'}, {'status': 'open'}, {'owner': 'keeper'},
                       {'fy': 'all'}, {'q': '88421'}, {'amount_min': '100000'}):
            s = self.summary(**params)
            l = self.list_(**params)
            self.assertEqual(s['count'], l['totals']['count'], params)
            self.assertEqual(s['amount'], l['totals']['amount'], params)

    def test_null_amounts_are_counted_but_not_summed(self):
        s = self.summary()
        self.assertEqual(s['count'], 10, 'NULL-amount rows must still be counted')
        self.assertEqual(Decimal(s['amount']), self.FY_A_AMOUNT_SUM,
                         'NULL amounts must be excluded from the sum')
        # HIGH bucket: 5 rows but only 4 carry an amount — the classic off-by-one.
        high = next(b for b in s['by_severity'] if b['key'] == 'HIGH')
        self.assertEqual(high['count'], 5)
        self.assertEqual(Decimal(high['amount']),
                         Decimal('45644.00') + Decimal('99810.00')
                         + Decimal('17420.01') + Decimal('138000.00'))

    def test_bucket_counts_and_amounts_sum_to_the_totals(self):
        s = self.summary()
        for key in ('by_severity', 'by_status', 'by_category'):
            self.assertEqual(sum(b['count'] for b in s[key]), s['count'], key)
            self.assertEqual(sum(Decimal(b['amount']) for b in s[key]),
                             Decimal(s['amount']), key)

    def test_by_severity_is_rank_ordered_and_omits_absent_severities(self):
        s = self.summary()
        self.assertEqual([b['key'] for b in s['by_severity']],
                         ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO'])
        only_med = self.summary(status='resolved')
        self.assertEqual([b['key'] for b in only_med['by_severity']], ['MEDIUM'],
                         'absent severities must be omitted, not zero-filled')

    def test_by_status_follows_the_statuses_tuple_order(self):
        keys = [b['key'] for b in self.summary()['by_status']]
        self.assertEqual(keys, [st for st in STATUSES if st in keys])
        self.assertEqual(keys[0], 'OPEN')

    def test_by_category_and_by_owner_count_desc_then_key_asc(self):
        s = self.summary()
        cats = [(b['key'], b['count']) for b in s['by_category']]
        self.assertEqual(cats[0], ('SUP', 3))
        self.assertEqual(cats[1], ('PAYROLL', 2))
        singles = [k for k, c in cats[2:]]
        self.assertEqual(singles, sorted(singles), 'ties must break by key ascending')
        owners = [(b['key'], b['count']) for b in s['by_owner']]
        self.assertEqual(owners[0], ('bookkeeper', 4))
        self.assertEqual(owners[1], ('MC', 3))
        self.assertEqual(owners[2], ('accountant', 2))

    def test_open_count(self):
        self.assertEqual(self.summary()['open_count'], 6,
                         "open_count must be the number of status=='OPEN' rows")

    def test_fy_options_descending_and_includes_the_empty_current_fy(self):
        opts = self.summary()['fy_options']
        self.assertEqual(opts, sorted(opts, reverse=True))
        self.assertIn(CUR, opts, 'the current FY must be offered even with no rows in it')
        self.assertIn(FY_A, opts)
        self.assertIn(FY_OLD, opts)

    def test_summary_fy_all(self):
        s = self.summary(fy='all')
        self.assertEqual(s['count'], 11)
        self.assertEqual(Decimal(s['amount']), self.FY_A_AMOUNT_SUM + Decimal('138000.00'))

    def test_summary_bad_fy_is_400(self):
        self.assertEqual(self.client.get(SUMMARY_URL, {'fy': 'notanumber'}).status_code, 400)
        self.assertEqual(self.client.get(SUMMARY_URL, {'fy': '3000'}).status_code, 400)


# --------------------------------------------------------------------------- #
# 6. Auth gate — 401 everywhere, correct challenge, no leak
# --------------------------------------------------------------------------- #
class AuthGateTests(FindingsFixtureMixin):
    def endpoints(self, client):
        pk = self.f1.pk
        upload = SimpleUploadedFile('x.pdf', b'%PDF-1.4 anon', content_type='application/pdf')
        return [
            ('list', client.get(LIST_URL)),
            ('create', client.post(LIST_URL, {'title': 't', 'severity': 'HIGH',
                                              'category': 'SUP', 'source': 's'}, format='json')),
            ('detail', client.get(detail_url(pk))),
            ('patch', client.patch(detail_url(pk), {'status': 'RESOLVED'}, format='json')),
            ('comments', client.post(comments_url(pk), {'text': 'anon comment'}, format='json')),
            ('attachments', client.post(attachments_url(pk), {'file': upload}, format='multipart')),
            ('bulk', client.post(BULK_URL, {'ids': [pk], 'status': 'RESOLVED'}, format='json')),
            ('summary', client.get(SUMMARY_URL)),
            ('export_csv', client.get(EXPORT_URL)),
            ('export_xlsx', client.get(EXPORT_URL, {'format': 'xlsx'})),
        ]

    def test_anonymous_is_401_with_bearer_challenge_on_every_endpoint(self):
        before = AuditFinding.objects.count()
        for name, resp in self.endpoints(self.anon):
            self.assertEqual(resp.status_code, 401, f'{name}: anonymous must be 401, not '
                                                    f'{resp.status_code}')
            self.assertEqual(resp.headers.get('WWW-Authenticate'), 'Bearer realm="api"', name)
        self.assertEqual(AuditFinding.objects.count(), before, 'anonymous create/bulk wrote rows')
        self.assertFalse(AuditFindingComment.objects.exists())

    def test_anonymous_401_bodies_leak_nothing(self):
        markers = ['Payments made before supplier bill captured', 'Higgsfield', 'FY26-001',
                   'ref,fy,title', 'bookkeeper', '429110.39']
        for name, resp in self.endpoints(self.anon):
            text = resp.content.decode('utf-8', 'replace')
            for marker in markers:
                self.assertNotIn(marker, text, f'{name}: 401 body leaks {marker!r}')

    def test_garbage_and_expired_bearer_tokens_are_401(self):
        from rest_framework_simplejwt.tokens import AccessToken
        for garbage in ('not.a.jwt', 'a' * 40, 'ey.ey.ey'):
            client = APIClient()
            client.credentials(HTTP_AUTHORIZATION=f'Bearer {garbage}')
            self.assertEqual(client.get(LIST_URL).status_code, 401, garbage)
            self.assertEqual(client.get(EXPORT_URL).status_code, 401, garbage)
        token = AccessToken.for_user(self.user)
        token.set_exp(lifetime=-dt.timedelta(seconds=1))
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        self.assertEqual(client.get(LIST_URL).status_code, 401)
        self.assertEqual(client.get(EXPORT_URL).status_code, 401,
                         'the export runs its own auth path and must reject expired tokens too')

    def test_route_order_summary_export_bulk_are_not_swallowed_by_the_pk_route(self):
        # If '<int:pk>/' were mounted first these would 404; they must resolve.
        self.assertEqual(self.client.get(SUMMARY_URL).status_code, 200)
        self.assertEqual(self.client.get(EXPORT_URL).status_code, 200)
        resp = self.client.post(BULK_URL, {'ids': [self.f1.pk], 'comment': 'route check'},
                                format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])


# --------------------------------------------------------------------------- #
# 7. Create validation + adversarial input (400, never 500)
# --------------------------------------------------------------------------- #
class CreateValidationTests(AuthedTestBase):
    def assert_4xx(self, resp, label):
        self.assertGreaterEqual(resp.status_code, 400, f'{label}: expected 4xx')
        self.assertLess(resp.status_code, 500, f'{label}: got a 5xx — this is a defect')
        try:
            self.assertIn('detail', resp.json(), f'{label}: 4xx must carry a useful detail')
        except ValueError:
            self.fail(f'{label}: 4xx body is not JSON: {resp.content[:200]}')

    def test_created_finding_dict_shape_and_defaults(self):
        resp = self.create(amount='429110.39', evidence=[{'type': 'note'}],
                           due_date=f'{CUR}-01-31', asana_gid='1217633700114593')
        self.assertEqual(resp.status_code, 201, resp.content[:500])
        body = resp.json()
        self.assertEqual(set(body.keys()), FINDING_DICT_KEYS)
        self.assertEqual(body['status'], 'OPEN')
        self.assertEqual(body['currency'], 'ZAR')
        self.assertEqual(body['amount'], '429110.39')
        self.assertEqual(body['due_date'], f'{CUR}-01-31')
        self.assertEqual(body['evidence'], [{'type': 'note', 'ref': '', 'note': ''}],
                         "missing evidence keys must default to ''")
        self.assertEqual(body['comment_count'], 0)
        self.assertEqual(body['attachment_count'], 0)

    def test_severity_and_status_are_case_insensitive_and_stored_upper(self):
        resp = self.create(severity='critical', status='in_progress')
        self.assertEqual(resp.status_code, 201, resp.content[:300])
        self.assertEqual(resp.json()['severity'], 'CRITICAL')
        self.assertEqual(resp.json()['status'], 'IN_PROGRESS')

    def test_missing_required_fields_are_400(self):
        for field in ('title', 'severity', 'category', 'source'):
            resp = self.create(**{field: ...})
            self.assert_4xx(resp, f'missing {field}')
        self.assertFalse(AuditFinding.objects.exists())

    def test_non_object_json_bodies_are_400_with_the_contract_message(self):
        for raw in ('["title"]', '"a string"', 'null', '42'):
            resp = self.tolerant.generic('POST', LIST_URL, data=raw,
                                         content_type='application/json')
            self.assertEqual(resp.status_code, 400, f'body={raw!r} -> {resp.status_code}')
            self.assertEqual(resp.json()['detail'], 'request body must be a JSON object', raw)
        self.assertFalse(AuditFinding.objects.exists())

    def test_nul_bytes_are_400_in_every_text_field(self):
        for field in ('title', 'description', 'owner', 'source', 'category', 'check_code'):
            resp = self.tolerant.post(LIST_URL, {
                'title': 't', 'severity': 'HIGH', 'category': 'SUP', 'source': 's',
                field: 'bad\x00value',
            }, format='json')
            self.assert_4xx(resp, f'NUL in {field}')
        self.assertFalse(AuditFinding.objects.exists())

    def test_bad_evidence_shapes_are_400(self):
        for ev in ('a string', ['a string'], [{'type': 'unknown', 'ref': '', 'note': ''}],
                   [{'ref': 'no type'}], [123], {'type': 'note'}):
            resp = self.tolerant.post(LIST_URL, {
                'title': 't', 'severity': 'HIGH', 'category': 'SUP', 'source': 's',
                'evidence': ev,
            }, format='json')
            self.assert_4xx(resp, f'evidence={ev!r}')
        self.assertFalse(AuditFinding.objects.exists())

    def test_bad_amounts_are_4xx_never_500(self):
        for amt in ('abc', 'NaN', 'Infinity', '-Infinity', '9' * 40, {'a': 1}, [1]):
            with self.subTest(amount=amt):
                resp = self.tolerant.post(LIST_URL, {
                    'title': 't', 'severity': 'HIGH', 'category': 'SUP', 'source': 's',
                    'amount': amt,
                }, format='json')
                self.assert_4xx(resp, f'amount={amt!r}')
        self.assertFalse(AuditFinding.objects.exists())

    def test_bad_due_dates_are_400(self):
        for d in ('2026-13-45', 'yesterday', '20260101', '2026-02-30', 123):
            resp = self.tolerant.post(LIST_URL, {
                'title': 't', 'severity': 'HIGH', 'category': 'SUP', 'source': 's',
                'due_date': d,
            }, format='json')
            self.assert_4xx(resp, f'due_date={d!r}')

    def test_overlong_title_and_category_are_4xx_never_500(self):
        resp = self.tolerant.post(LIST_URL, {
            'title': 'x' * 301, 'severity': 'HIGH', 'category': 'SUP', 'source': 's',
        }, format='json')
        self.assert_4xx(resp, 'title > 300 chars')
        resp = self.tolerant.post(LIST_URL, {
            'title': 't', 'severity': 'HIGH', 'category': 'C' * 33, 'source': 's',
        }, format='json')
        self.assert_4xx(resp, 'category > 32 chars')
        self.assertFalse(AuditFinding.objects.exists())

    def test_bad_fy_in_post_is_400(self):
        for fy in (1999, 3000, 'abc', 20.5):
            resp = self.tolerant.post(LIST_URL, {
                'title': 't', 'severity': 'HIGH', 'category': 'SUP', 'source': 's', 'fy': fy,
            }, format='json')
            self.assert_4xx(resp, f'fy={fy!r}')

    def test_bad_severity_names_the_allowed_values(self):
        resp = self.create(severity='URGENT')
        self.assertEqual(resp.status_code, 400)
        detail = resp.json()['detail']
        for sev in SEVERITIES:
            self.assertIn(sev, detail, 'the 400 must name the allowed severities')


# --------------------------------------------------------------------------- #
# 8. Detail + PATCH
# --------------------------------------------------------------------------- #
class PatchTests(ExplicitFyMixin):
    def test_patch_edits_and_stamps_updated_by(self):
        resp = self.client.patch(detail_url(self.f1.pk), {
            'status': 'resolved', 'owner': 'accountant', 'amount': '100.5',
        }, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:500])
        body = resp.json()
        self.assertEqual(body['status'], 'RESOLVED')
        self.assertEqual(body['owner'], 'accountant')
        self.assertEqual(body['amount'], '100.50', 'amount must serialise as a 2dp string')
        self.assertEqual(body['updated_by'], 'auditor-mc')
        self.assertEqual(body['created_by'], 'seed', 'PATCH must not touch created_by')

    def test_fy_and_ref_are_immutable_and_silently_ignored(self):
        resp = self.client.patch(detail_url(self.f1.pk), {
            'fy': 2030, 'ref': 'FY30-999', 'title': 'Retitled finding',
        }, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['fy'], FY_A)
        self.assertEqual(body['ref'], ref_for(FY_A, 1))
        self.assertEqual(body['title'], 'Retitled finding')

    def test_patch_with_only_immutable_or_no_keys_is_400(self):
        self.assertEqual(self.client.patch(detail_url(self.f1.pk), {}, format='json').status_code,
                         400, 'empty PATCH must be 400')
        resp = self.client.patch(detail_url(self.f1.pk), {'fy': 2030}, format='json')
        self.assertEqual(resp.status_code, 400,
                         'fy is not an editable key, so a fy-only PATCH has no editable key')

    def test_patch_validation_matches_post(self):
        for body in ({'severity': 'URGENT'}, {'status': 'nope'}, {'amount': 'abc'},
                     {'due_date': 'yesterday'}, {'title': 'x\x00y'},
                     {'evidence': [{'type': 'bogus'}]}):
            resp = self.tolerant.patch(detail_url(self.f1.pk), body, format='json')
            self.assertEqual(resp.status_code, 400, f'{body} -> {resp.status_code}')

    def test_unknown_pk_is_404_with_the_contract_detail(self):
        resp = self.client.patch(detail_url(999999), {'status': 'RESOLVED'}, format='json')
        self.assertEqual(resp.status_code, 404)
        self.assertEqual(resp.json(), {'detail': 'finding not found'})
        self.assertEqual(self.client.get(detail_url(999999)).status_code, 404)

    def test_detail_get_carries_the_finding(self):
        resp = self.client.get(detail_url(self.f1.pk))
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        finding = body.get('finding', body)  # contract leaves the GET wrapper shape open
        self.assertEqual(finding['ref'], ref_for(FY_A, 1))
        self.assertEqual(finding['amount'], '429110.39')

    def test_patch_non_object_body_is_400_never_500(self):
        for raw in ('["status"]', '"x"', 'null'):
            resp = self.tolerant.generic('PATCH', detail_url(self.f1.pk), data=raw,
                                         content_type='application/json')
            self.assertEqual(resp.status_code, 400, f'body={raw!r} -> {resp.status_code}')


# --------------------------------------------------------------------------- #
# 9. Comments
# --------------------------------------------------------------------------- #
class CommentTests(ExplicitFyMixin):
    def test_comment_created_with_author_and_shape(self):
        resp = self.client.post(comments_url(self.f1.pk),
                                {'text': '  chase the bookkeeper for the bill  '}, format='json')
        self.assertEqual(resp.status_code, 201, resp.content[:300])
        body = resp.json()
        self.assertEqual(set(body.keys()), {'id', 'finding_id', 'text', 'author', 'created_at'})
        self.assertEqual(body['finding_id'], self.f1.pk)
        self.assertEqual(body['author'], 'auditor-mc')
        # comment_count reflected on the finding
        row = next(r for r in self.list_()['results'] if r['id'] == self.f1.pk)
        self.assertEqual(row['comment_count'], 1)

    def test_blank_whitespace_and_nul_text_are_400(self):
        for t in ('', '   ', '\n\t', None, 'bad\x00comment'):
            resp = self.tolerant.post(comments_url(self.f1.pk), {'text': t}, format='json')
            self.assertEqual(resp.status_code, 400, (t, resp.status_code))
        self.assertEqual(self.tolerant.post(comments_url(self.f1.pk), {}, format='json')
                         .status_code, 400)
        self.assertFalse(AuditFindingComment.objects.exists())

    def test_comment_on_unknown_finding_is_404(self):
        resp = self.client.post(comments_url(999999), {'text': 'x'}, format='json')
        self.assertEqual(resp.status_code, 404)


# --------------------------------------------------------------------------- #
# 10. Bulk
# --------------------------------------------------------------------------- #
class BulkTests(ExplicitFyMixin):
    def bulk(self, payload):
        return self.client.post(BULK_URL, payload, format='json')

    def test_basic_bulk_update_and_comment(self):
        resp = self.bulk({'ids': [self.f1.pk, self.f3.pk], 'status': 'RESOLVED',
                          'comment': 'closed at year-end review'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['updated'], 2)
        self.assertEqual(body['commented'], 2)
        self.assertEqual(body['unknown'], [])
        self.f1.refresh_from_db()
        self.assertEqual(self.f1.status, 'RESOLVED')
        self.assertEqual(AuditFindingComment.objects.filter(finding=self.f1).count(), 1)

    def test_cap_applies_after_deduplication(self):
        # 600 copies of one id is ONE action, not a 400.
        resp = self.bulk({'ids': [self.f1.pk] * 600, 'owner': 'bookkeeper'})
        self.assertEqual(resp.status_code, 200,
                         'cap must apply AFTER de-duplication: 600 copies of one id is one action')
        self.assertEqual(resp.json()['updated'], 1)
        # 500 distinct ids is fine; 501 is not.
        ok = self.bulk({'ids': list(range(100000, 100000 + BULK_MAX)), 'owner': 'x'})
        self.assertEqual(ok.status_code, 200, ok.content[:300])
        self.assertEqual(len(ok.json()['unknown']), BULK_MAX)
        over = self.bulk({'ids': list(range(100000, 100000 + BULK_MAX + 1)), 'owner': 'x'})
        self.assertEqual(over.status_code, 400, over.content[:300])
        self.assertIn('500', over.json()['detail'])

    def test_unknown_ids_reported_never_fatal(self):
        resp = self.bulk({'ids': [self.f1.pk, 999999, 888888], 'status': 'ACCEPTED'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['updated'], 1)
        self.assertEqual(set(body['unknown']), {999999, 888888})
        all_unknown = self.bulk({'ids': [999999], 'status': 'ACCEPTED'})
        self.assertEqual(all_unknown.status_code, 200, 'all-unknown ids is still a 200')
        self.assertEqual(all_unknown.json()['updated'], 0)

    def test_key_presence_not_truthiness_empty_owner_clears(self):
        AuditFinding.objects.filter(pk=self.f1.pk).update(owner='bookkeeper')
        resp = self.bulk({'ids': [self.f1.pk], 'owner': ''})
        self.assertEqual(resp.status_code, 200,
                         "{'owner': ''} is a legitimate clear-the-owner action, not a missing key")
        self.assertEqual(resp.json()['updated'], 1)
        self.f1.refresh_from_db()
        self.assertEqual(self.f1.owner, '')

    def test_string_ids_that_parse_are_accepted(self):
        resp = self.bulk({'ids': [str(self.f1.pk)], 'status': 'RESOLVED'})
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated'], 1)

    def test_bad_payload_shapes_are_400_never_500(self):
        for payload in ({'ids': [], 'status': 'RESOLVED'},          # empty ids
                        {'ids': 'notalist', 'status': 'RESOLVED'},  # not a list
                        {'ids': [self.f1.pk]},                      # no action key
                        {'status': 'RESOLVED'}):                    # no ids at all
            resp = self.tolerant.post(BULK_URL, payload, format='json')
            self.assertEqual(resp.status_code, 400, f'{payload} -> {resp.status_code}')
        for raw in ('[1,2]', '"ids"', 'null', '42'):
            resp = self.tolerant.generic('POST', BULK_URL, data=raw,
                                         content_type='application/json')
            self.assertEqual(resp.status_code, 400, f'body={raw!r} -> {resp.status_code}')

    def test_bulk_bad_field_values_never_500(self):
        for payload in ({'ids': [self.f1.pk], 'status': 'nope'},
                        {'ids': [self.f1.pk], 'due_date': '2026-13-45'},
                        {'ids': [self.f1.pk], 'comment': 'x\x00y'},
                        {'ids': [self.f1.pk, 'abc'], 'status': 'RESOLVED'}):
            resp = self.tolerant.post(BULK_URL, payload, format='json')
            self.assertGreaterEqual(resp.status_code, 400, f'{payload}')
            self.assertLess(resp.status_code, 500, f'{payload} -> {resp.status_code} (5xx defect)')


# --------------------------------------------------------------------------- #
# 11. Export
# --------------------------------------------------------------------------- #
class ExportTests(ExplicitFyMixin):
    def export(self, **params):
        params.setdefault('fy', str(FY_A))
        return self.client.get(EXPORT_URL, params)

    def csv_rows(self, resp):
        return list(csv.reader(io.StringIO(resp.content.decode('utf-8'))))

    def test_csv_header_exactly_matches_the_contract(self):
        resp = self.export()
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertTrue(resp['Content-Type'].startswith('text/csv'), resp['Content-Type'])
        self.assertRegex(resp['Content-Disposition'],
                         r'filename="audit-findings-\d{4}-\d{2}-\d{2}\.csv"')
        rows = self.csv_rows(resp)
        self.assertEqual(rows[0], EXPECTED_EXPORT_HEADER)
        self.assertEqual(len(rows) - 1, 10, 'export row count must equal the default-FY count')

    def test_csv_injection_title_survives_quoted(self):
        rows = self.csv_rows(self.export())
        title_idx = EXPECTED_EXPORT_HEADER.index('title')
        titles = [r[title_idx] for r in rows[1:]]
        self.assertIn(INJECTION_TITLE, titles,
                      'a title with =, comma, quote and newline must round-trip through the CSV')

    def test_csv_respects_filters(self):
        rows = self.csv_rows(self.export(severity='HIGH'))
        self.assertEqual(len(rows) - 1, 5)
        rows = self.csv_rows(self.export(fy='all'))
        self.assertEqual(len(rows) - 1, 11)
        rows = self.csv_rows(self.export(q='zzz-no-match'))
        self.assertEqual(len(rows), 1)

    def test_csv_evidence_and_comments_columns(self):
        AuditFindingComment.objects.create(finding=self.f1, text='chase it', author='mc')
        rows = self.csv_rows(self.export())
        rec = dict(zip(rows[0], next(r for r in rows[1:]
                                     if r[EXPECTED_EXPORT_HEADER.index('ref')] == self.f1.ref)))
        self.assertEqual(rec['evidence'],
                         'journal:JN-962018 — payment posted 41 days before the bill')
        self.assertEqual(rec['comments'], '1')
        self.assertEqual(rec['amount'], '429110.39')
        self.assertEqual(rec['created_by'], 'seed')

    def test_xlsx_export_opens_with_openpyxl(self):
        from openpyxl import load_workbook
        resp = self.export(format='xlsx')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp['Content-Type'],
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertRegex(resp['Content-Disposition'],
                         r'filename="audit-findings-\d{4}-\d{2}-\d{2}\.xlsx"')
        self.assertNotIn('X-Export-Note', resp, 'openpyxl is installed; must not degrade')
        ws = load_workbook(io.BytesIO(resp.content)).active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), EXPECTED_EXPORT_HEADER)
        self.assertEqual(len(rows) - 1, 10)

    def test_bad_format_is_400(self):
        for f in ('pdf', 'json', 'exe', '../../etc/passwd'):
            resp = self.export(format=f)
            self.assertEqual(resp.status_code, 400, f)

    def test_export_junk_filters_never_500(self):
        resp = self.tolerant.get(EXPORT_URL, {'ordering': 'evidence', 'amount_min': 'NaN',
                                              'q': "'; drop table x; --", 'page': 'x'})
        self.assertLess(resp.status_code, 500, resp.status_code)


# --------------------------------------------------------------------------- #
# 12. Identity stamping
# --------------------------------------------------------------------------- #
@override_settings(KLIKK_API_TOKEN='test-service-token-abc123')
class IdentityTests(AuthedTestBase):
    def test_service_token_caller_is_stamped_mcp(self):
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION='Bearer test-service-token-abc123')
        resp = client.post(LIST_URL, {
            'title': 'MCP-created finding', 'severity': 'HIGH', 'category': 'SUP',
            'source': 'mcp add_audit_finding', 'fy': 2026,
        }, format='json')
        self.assertEqual(resp.status_code, 201, resp.content[:500])
        self.assertEqual(resp.json()['created_by'], 'mcp')

    def test_jwt_caller_is_stamped_with_their_username(self):
        resp = self.create(title='Console-created finding')
        self.assertEqual(resp.status_code, 201, resp.content[:300])
        self.assertEqual(resp.json()['created_by'], 'auditor-mc')

    def test_updated_by_changes_on_patch_but_created_by_does_not(self):
        created = self.create(title='Ownership trail').json()
        second = User.objects.create_user(username='second-reviewer', password='x')
        self.auth(second)
        resp = self.client.patch(detail_url(created['id']), {'owner': 'MC'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        body = resp.json()
        self.assertEqual(body['created_by'], 'auditor-mc')
        self.assertEqual(body['updated_by'], 'second-reviewer')
        # service-token PATCH stamps updated_by = 'mcp'
        svc = APIClient()
        svc.credentials(HTTP_AUTHORIZATION='Bearer test-service-token-abc123')
        resp = svc.patch(detail_url(created['id']), {'owner': 'bookkeeper'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertEqual(resp.json()['updated_by'], 'mcp')
        self.assertEqual(resp.json()['created_by'], 'auditor-mc')

    def test_client_supplied_identity_fields_are_ignored(self):
        resp = self.create(title='Identity smuggling', created_by='evil', updated_by='evil')
        self.assertEqual(resp.status_code, 201, resp.content[:300])
        self.assertEqual(resp.json()['created_by'], 'auditor-mc',
                         'created_by must be stamped from the caller, never taken from the body')


# --------------------------------------------------------------------------- #
# 13. Attachments + signed viewer
# --------------------------------------------------------------------------- #
class AttachmentTests(ExplicitFyMixin):
    FILE_BYTES = b'%PDF-1.4\n% adversarial attachment body 0123456789\n' * 10

    def upload(self, finding=None, name='aurras-invoice.pdf', content=None,
               content_type='application/pdf'):
        f = SimpleUploadedFile(name, content or self.FILE_BYTES, content_type=content_type)
        return self.client.post(attachments_url((finding or self.f1).pk), {'file': f},
                                format='multipart')

    def _view_path(self, view_url):
        """Path + query only — never GET an absolute production URL through the test client."""
        parsed = urlparse(view_url)
        return parsed.path, parse_qs(parsed.query)

    @staticmethod
    def _local(path: str) -> str:
        """Strip any nginx base prefix so the Django test client hits the real route."""
        return path[path.index('/audit/'):] if '/audit/' in path else path

    def contract_viewer_path(self, att_id: int) -> str:
        return f'/audit/findings/attachments/{att_id}/file/'

    def test_upload_response_shape_signature_and_contract_path(self):
        with self.settings(MEDIA_ROOT='/tmp/test-audit-findings-media'):
            resp = self.upload()
            self.assertIn(resp.status_code, (200, 201), resp.content[:500])
            body = resp.json()
            for key in ('id', 'finding_id', 'original_name', 'content_type', 'size',
                        'uploaded_by', 'created_at', 'view_url'):
                self.assertIn(key, body)
            self.assertEqual(body['original_name'], 'aurras-invoice.pdf')
            self.assertEqual(body['size'], len(self.FILE_BYTES))
            self.assertEqual(body['uploaded_by'], 'auditor-mc')
            att_id = body['id']
            path, q = self._view_path(body['view_url'])
            self.assertEqual(q.get('s', [''])[0], attachment_sig(att_id),
                             "the view_url signature must match the contract's HMAC formula")
            self.assertTrue(
                path.endswith(f'/audit/findings/attachments/{att_id}/file/'),
                f"view_url path {path!r} is not the contract's "
                f"/audit/findings/attachments/<pk>/file/")

    def test_emitted_view_url_actually_serves_the_bytes(self):
        # The integration truth: whatever URL the API hands out MUST serve. A view_url
        # that 404s is a dead link in every export and in the console's viewer.
        with self.settings(MEDIA_ROOT='/tmp/test-audit-findings-media'):
            body = self.upload().json()
            path, q = self._view_path(body['view_url'])
            got = self.anon.get(self._local(path), {'s': q.get('s', [''])[0]})
            self.assertEqual(got.status_code, 200,
                             f'the emitted view_url {body["view_url"]!r} does not serve '
                             f'(got {got.status_code}) — dead link in exports/console')
            self.assertEqual(got.content, self.FILE_BYTES)
            # attachment_count reflected on the finding
            row = next(r for r in self.list_()['results'] if r['id'] == self.f1.pk)
            self.assertEqual(row['attachment_count'], 1)

    def test_contract_viewer_route_serves_with_a_valid_signature(self):
        # The viewer itself, addressed at the CONTRACT route, independent of the
        # (possibly wrong) URL the serialiser emits. Deliberately anonymous.
        with self.settings(MEDIA_ROOT='/tmp/test-audit-findings-media'):
            att_id = self.upload().json()['id']
            got = self.anon.get(self.contract_viewer_path(att_id),
                                {'s': attachment_sig(att_id)})
            self.assertEqual(got.status_code, 200, got.content[:200])
            self.assertEqual(got.content, self.FILE_BYTES)
            self.assertEqual(got['Content-Type'], 'application/pdf')

    def test_tampered_or_missing_signature_is_403(self):
        with self.settings(MEDIA_ROOT='/tmp/test-audit-findings-media'):
            att_id = self.upload().json()['id']
            local = self.contract_viewer_path(att_id)
            good = attachment_sig(att_id)
            bad = ('0' if good[0] != '0' else '1') + good[1:]
            self.assertEqual(self.anon.get(local, {'s': bad}).status_code, 403)
            self.assertEqual(self.anon.get(local, {'s': good[:-1]}).status_code, 403)
            self.assertEqual(self.anon.get(local).status_code, 403, 'missing ?s= must be 403')
            # A valid signature for attachment X must not open attachment Y.
            other_id = self.upload(name='second.pdf').json()['id']
            self.assertEqual(
                self.anon.get(self.contract_viewer_path(other_id), {'s': good}).status_code,
                403, "attachment X's signature must not open attachment Y")

    def test_unknown_attachment_pk_is_not_200_or_500(self):
        resp = self.anon.get('/audit/findings/attachments/999999/file/',
                             {'s': attachment_sig(999999)})
        self.assertIn(resp.status_code, (403, 404), resp.status_code)

    def test_oversized_file_is_400(self):
        with self.settings(MEDIA_ROOT='/tmp/test-audit-findings-media'):
            big = b'x' * (26 * 1024 * 1024)   # comfortably over the 25 MB cap
            resp = self.upload(content=big, name='huge.bin', content_type='application/octet-stream')
            self.assertEqual(resp.status_code, 400, f'{resp.status_code}: oversize must be 400')

    def test_missing_file_field_is_400(self):
        resp = self.tolerant.post(attachments_url(self.f1.pk), {}, format='multipart')
        self.assertEqual(resp.status_code, 400, resp.status_code)
